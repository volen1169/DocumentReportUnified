
# =============================================================================
# V4.4 FUNCTION DOCUMENTATION EDITION
#
# GOAL
#   ทำให้ค้นหาและแก้ไขโค้ดได้ง่ายที่สุดสำหรับมนุษย์และ AI
#
# SEARCH TAGS
#   FUNCTION :
#   CSS OWNER :
#   UI OWNER :
#   DATA FLOW :
#   SHAREPOINT :
#   SIDEBAR :
#   DASHBOARD :
#   HARDWARE :
#   REPORTS :
# =============================================================================


# =============================================================================
# V4.3 DEEP COMMENT EDITION
#
# COMMENT STANDARD
#   PURPOSE      = ทำหน้าที่อะไร
#   USED BY      = ถูกเรียกจากส่วนไหน
#   DEPENDENCY   = พึ่งพาอะไร
#   UI OWNER     = หน้าจอที่เกี่ยวข้อง
#   CSS OWNER    = CSS ที่เกี่ยวข้อง
# =============================================================================


# =============================================================================
# DOCUMENT REPORT UNIFIED
# VERSION : V4.1 COMMENTED EDITION
# =============================================================================
# TABLE OF CONTENTS
#
# SECTION 01 : CONFIGURATION
# SECTION 02 : AUTHENTICATION
# SECTION 03 : SHAREPOINT
# START SECTION 04 : THEME SYSTEM
# SECTION 04 : THEME SYSTEM
# END SECTION 04 : THEME SYSTEM

# START SECTION 05 : SIDEBAR
# SECTION 05 : SIDEBAR
# SECTION 06 : DASHBOARD
# SECTION 07 : ASSET MANAGEMENT
# SECTION 08 : PASSWORD MANAGER
# SECTION 09 : REPORTS & ANALYTICS
# SECTION 10 : ADMINISTRATION
#
# QUICK SEARCH TAGS
#   # THEME : LOGIN
#   # =============================================================================
# THEME : SIDEBAR
#
# PURPOSE
#   ควบคุมหน้าตา Sidebar ทั้งหมด
#
# COMPONENTS
#   - Background
#   - Navigation Button
#   - Active Menu
#   - Badge Counter
#   - Profile Card
#
# SEARCH TAG
#   SIDEBAR_THEME
# =============================================================================
# THEME : SIDEBAR
#   # THEME : DASHBOARD
#   # THEME : HARDWARE
#   # THEME : PASSWORD
# =============================================================================


# =============================================================================
# V3.1 DESIGN SYSTEM
# ใช้รวมสีหลักของระบบให้ Sidebar และ Login ใช้ Theme เดียวกัน
# =============================================================================
COLORS = {
    "primary": "#6366F1",     # Indigo
    "secondary": "#8B5CF6",   # Purple
    "accent": "#38BDF8",      # Sky
    "success": "#10B981",
    "warning": "#F59E0B",
    "danger": "#EF4444",
    "bg": "#F8FAFC",
    "card": "#FFFFFF",
}



# =============================================================================
# V3.2 SIDEBAR MODERN THEME
# Login Theme Inspired Sidebar
# =============================================================================
# =============================================================================
# THEME : SIDEBAR
# พื้นหลัง Sidebar / Active Menu / Profile Card / Badge
# =============================================================================
SIDEBAR_V32_THEME = """
<style>
/* LOGIN STYLE SIDEBAR */
section[data-testid="stSidebar"],
[data-testid="stSidebar"]{
    background:
        radial-gradient(circle at top left, rgba(56,189,248,.18), transparent 28%),
        radial-gradient(circle at bottom right, rgba(139,92,246,.15), transparent 30%),
        linear-gradient(135deg,#e0f2fe 0%,#dbeafe 35%,#ede9fe 100%) !important;
}

/* Sidebar inner container */
[data-testid="stSidebar"] > div:first-child{
    background:transparent !important;
}

/* Default text */
[data-testid="stSidebar"] *{
    color:#334155 !important;
}

/* Navigation buttons */
[data-testid="stSidebar"] .stButton > button{
    border-radius:16px !important;
}

/* Active menu */
[data-testid="stSidebar"] .stButton > button[kind="primary"]{
    background:transparent !important;
    color:#2563EB !important;
    border:none !important;
    border-left:4px solid #2563EB !important;
    box-shadow:none !important;
}

/* Badge */
.sidebar-badge{
    background:rgba(255,255,255,.65);
    backdrop-filter:blur(8px);
    border-radius:999px;
}

/* Profile card glass */
.profile-card{
    background:rgba(255,255,255,.30);
    backdrop-filter:blur(20px);
    border:1px solid rgba(255,255,255,.45);
    border-radius:20px;
}

/* ==========================================================
   FIX STREAMLIT SIDEBAR WHITE BOX
   ========================================================== */

[data-testid="stSidebar"]
[data-testid="stVerticalBlockBorderWrapper"]{
    background:transparent !important;
    border:none !important;
    box-shadow:none !important;
}

[data-testid="stSidebar"]
[data-testid="stVerticalBlockBorderWrapper"] > div{
    background:transparent !important;
    border:none !important;
    box-shadow:none !important;
    backdrop-filter:none !important;
}

[data-testid="stSidebar"]
[data-testid="stHorizontalBlock"]{
    background:transparent !important;
}

[data-testid="stSidebar"]
[data-testid="stHorizontalBlock"] > div{
    background:transparent !important;
    border:none !important;
    box-shadow:none !important;
}

[data-testid="stSidebar"]
div[data-testid="stVerticalBlock"]{
    background:transparent !important;
    border:none !important;
    box-shadow:none !important;
}

[data-testid="stSidebar"] .stButton{
    background:transparent !important;
    border:none !important;
    box-shadow:none !important;
}

/* ========================================================================= */
/* DEBUG REMOVE NAVIGATION WHITE BOX                                         */
/* ใช้ตรวจว่ากล่องขาวมาจากปุ่ม Secondary หรือไม่                          */
/* ========================================================================= */

[data-testid="stSidebar"] .stButton > button[kind="secondary"]{
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    backdrop-filter: none !important;
}

[data-testid="stSidebar"] .stButton > button[kind="secondary"]:hover{
    background: rgba(255,255,255,0.12) !important;
    border: 1px solid rgba(255,255,255,0.15) !important;
    box-shadow: none !important;
}

/* ล้าง wrapper รอบปุ่ม */
[data-testid="stSidebar"] .stButton{
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
}

/* ล้าง element-container ที่ Streamlit สร้าง */
[data-testid="stSidebar"] div[data-testid="stElementContainer"]{
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
}

/* FORCE KILL ALL SIDEBAR BUTTON BACKGROUND */

[data-testid="stSidebar"] .stButton > button{
    background: transparent !important;
    box-shadow: none !important;
}

[data-testid="stSidebar"] .stButton > button:hover{
    box-shadow: none !important;
}


[data-testid="stSidebar"] button p,
[data-testid="stSidebar"] button span{
    background: transparent !important;
    border-radius: 0 !important;
    box-shadow: none !important;
}



</style>
"""

# =============================================================================
# SIDEBAR V3.1 ROADMAP
# 1. ปรับ Sidebar ให้ใช้โทนเดียวกับ Login
# 2. ปรับ Active Menu เป็น Gradient
# 3. ปรับ Badge Counter
# 4. ปรับ Profile Card
# =============================================================================


# =============================================================================
# THEME ARCHITECTURE (PHASE 2A)
# =============================================================================
# GLOBAL_THEME      -> CSS ใช้ทั้งระบบ
# LOGIN_THEME       -> หน้า Login
# SIDEBAR_THEME     -> Sidebar และ Navigation
# DASHBOARD_THEME   -> Dashboard Cards / Metrics
# HARDWARE_THEME    -> Asset Cards
# PASSWORD_THEME    -> Password Manager
# COMMON_THEME      -> Utility Classes
# =============================================================================

import streamlit as st
import pandas as pd
import paramiko
import re
import textwrap
import datetime
import inspect
import html
from concurrent.futures import ThreadPoolExecutor, as_completed
import io
import msal
import requests
import urllib3
import extra_streamlit_components as stx
import plotly.express as px
from openpyxl import load_workbook
from copy import copy
try:
    from ldap3 import ALL, SUBTREE, Connection, Server
except Exception:
    ALL = SUBTREE = Connection = Server = None

# =============================================================================
# SECTION 01 : CONFIGURATION
# ค่าตั้งต้นระบบ / SharePoint / NAS / Admin / Dropdown ต่างๆ
# =============================================================================
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# SharePoint Config
TENANT_ID = st.secrets["TENANT_ID"]
CLIENT_ID = st.secrets["CLIENT_ID"]
CLIENT_SECRET = st.secrets["CLIENT_SECRET"]
SHAREPOINT_DOMAIN = "optimalcoth.sharepoint.com"
SITE_NAME = "InformationTechnology"
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
GRAPH_URL = "https://graph.microsoft.com/v1.0"


def _secret_bool(name: str, default=False) -> bool:
    value = st.secrets.get(name, default)
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("1", "true", "yes", "y", "on")

# AD / Firewall Policy Source
# ---------------------------------------------------------------------------
# AD_POLICY_SOURCE:
# - "ldap"  = query Domain Controller / AD Server directly via LDAP/LDAPS
# - "agent" = call internal AD Agent API
# - "graph" = Microsoft Graph / Entra ID
# - "auto"  = try ldap, then agent, then graph
AD_POLICY_SOURCE = st.secrets.get("AD_POLICY_SOURCE", "auto").lower().strip()

# LDAP / LDAPS direct AD settings
# Example:
# AD_LDAP_SERVER = "192.168.2.3"
# AD_LDAP_PORT = 389          # 636 for LDAPS
# AD_LDAP_USE_SSL = false     # true for LDAPS
# AD_DOMAIN = "optimalgroup.com"
# AD_BASE_DN = "DC=optimalgroup,DC=com"
# AD_BIND_USER = "OPTIMALGROUP\\svc_ad_reader" or "svc_ad_reader@optimalgroup.com"
# AD_BIND_PASSWORD = "..."
AD_LDAP_SERVER = st.secrets.get("AD_LDAP_SERVER", st.secrets.get("AD_SERVER", "")).strip()
AD_LDAP_USE_SSL = _secret_bool("AD_LDAP_USE_SSL", False)
AD_LDAP_PORT = int(st.secrets.get("AD_LDAP_PORT", 636 if AD_LDAP_USE_SSL else 389))
AD_DOMAIN = st.secrets.get("AD_DOMAIN", "").strip()
AD_BASE_DN = st.secrets.get("AD_BASE_DN", "").strip()
AD_BIND_USER = st.secrets.get("AD_BIND_USER", "").strip()
AD_BIND_PASSWORD = st.secrets.get("AD_BIND_PASSWORD", "")
AD_LDAP_TIMEOUT = int(st.secrets.get("AD_LDAP_TIMEOUT", 15))

# Optional internal AD Agent API settings
# Expected endpoints:
# GET {AD_AGENT_URL}/user-policy?user=<identity>
# Response: {"ok": true, "user": {...}, "groups": ["FW_Officer_D"], "policies": [...]}
AD_AGENT_URL = st.secrets.get("AD_AGENT_URL", "").rstrip("/")
AD_AGENT_TOKEN = st.secrets.get("AD_AGENT_TOKEN", "")

# NAS Config
# -----------------------------------------------------------------------------
# โหมดเดิม: SSH เข้า NAS โดยตรง (ใช้ได้เฉพาะกรณี NAS เปิด SSH จากภายนอก/VPN)
# โหมดใหม่: Synology DSM API ผ่าน Cloudflare Tunnel
#
# แนะนำสำหรับ Streamlit Cloud:
# NAS_AGENT_URL = "https://nas-agent.poonyaruk.co.th"
# NAS_USER     = "ชื่อผู้ใช้ NAS"
# NAS_PASSWORD = "รหัสผ่าน NAS"
# NAS_MODE     = "agent"
# -----------------------------------------------------------------------------
NAS_IP = st.secrets.get("NAS_IP", "")
NAS_PORT = int(st.secrets.get("NAS_PORT", 22))        # รองรับ SSH port custom เช่น 2222
SSH_USER = st.secrets.get("SSH_USER", "")
SSH_PWD = st.secrets.get("SSH_PWD", "")

NAS_BASE_URL = st.secrets.get("NAS_BASE_URL", st.secrets.get("NAS_URL", "")).rstrip("/")
NAS_USER = st.secrets.get("NAS_USER", SSH_USER)
NAS_PASSWORD = st.secrets.get("NAS_PASSWORD", SSH_PWD)
NAS_MODE = st.secrets.get("NAS_MODE", "api" if NAS_BASE_URL else "ssh").lower().strip()

# NAS Local API Agent Config
# ใช้สำหรับอ่าน ACL จริงจาก synoacltool ผ่าน Agent ที่รันอยู่บน NAS
# Streamlit Secrets ตัวอย่าง:
# NAS_AGENT_URL   = "https://nas-agent.poonyaruk.co.th"
# NAS_AGENT_TOKEN = "รหัสเดียวกับ NAS_AGENT_TOKEN ใน Docker"
NAS_AGENT_URL = st.secrets.get("NAS_AGENT_URL", "").rstrip("/")
NAS_AGENT_TOKEN = st.secrets.get("NAS_AGENT_TOKEN", "")
NAS_AGENT_MODE = st.secrets.get("NAS_AGENT_MODE", "agent" if NAS_AGENT_URL else "").lower().strip()

# Force Agent-first mode: ป้องกัน Streamlit ไปเรียก DSM domain เดิม (nas-api) ตอนใช้ NAS Agent
AGENT_ONLY_MODE = NAS_MODE in ("agent", "nas_agent", "nas-agent") or NAS_AGENT_MODE == "agent"
NAS_SHARES_SECRET = st.secrets.get("NAS_SHARES", "")

MAX_THREADS = 6
NAS_TIMEOUT = int(st.secrets.get("NAS_TIMEOUT", 30))
SYNOACL_PATH = ""  # Deprecated: Streamlit ไม่เรียก synoacltool เองแล้ว ใช้ NAS Agent แทน

# Password File Config
SHAREPOINT_FOLDER = "Update IT documents"
PASSWORD_FILE_NAME = "Password.xlsx"

# ✅ Admin List — เพิ่ม/ลด email ได้ที่นี่
ADMIN_EMAILS = [
    "itsupport@poonyaruk.co.th",
    "IT_Network@poonyaruk.co.th",
    "Teerapat.Po@optimal.co.th",

]

# Internet Policy Mapping
# -----------------------------------------------------------------------------
# ใช้สำหรับแสดงสิทธิ์ออก Internet จาก AD / Firewall Group
# แนวคิด: ให้ AD Group เป็น Source of Truth แล้ว Firewall และระบบนี้อ่านจาก Group เดียวกัน
# ถ้าบริษัทมี Policy เพิ่ม ให้เพิ่มชื่อ Group และคำอธิบายที่นี่ได้เลย
# -----------------------------------------------------------------------------
FW_POLICY_PREFIXES = ("FW_", "Firewall_", "Internet_")
FIREWALL_POLICY_MAPPING_LIST = st.secrets.get("FIREWALL_POLICY_MAPPING_LIST", "Firewall Policy Mapping")
FW_POLICY_MAP = {
    "FW_Officer_A": "Allow All Website",
    "FW_Officer_B": "Block Social Media",
    "FW_Officer_C": "Allow YouTube",
    "FW_Officer_D": "Allow Facebook",
    "FW_Officer_E": "Allow YouTube, Facebook",
    "FW_Manager": "Net True",
    "FW_IT": "IT Internet Policy",
    "FW_MD": "Management Internet Policy",
    "FW_Supervisor_B": "Supervisor Internet Policy",
    "FW_Conference": "Conference Room Internet Policy",
}
FW_POLICY_DEFAULT_DETAILS = {
    "FW_Officer_A": {
        "Policy Name": "Allow All Website",
        "Internet Level": "Full Access",
        "Allowed": "Web, Email, Business apps, YouTube, Facebook, Social media",
        "Blocked": "-",
        "Firewall Rule": "FW_Officer_A",
        "Description": "ใช้งาน Internet ได้เต็มตาม policy พนักงานกลุ่ม A",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_Officer_B": {
        "Policy Name": "Block Social Media",
        "Internet Level": "Standard Access",
        "Allowed": "Web, Email, Business apps",
        "Blocked": "Facebook, TikTok, Instagram, Social media",
        "Firewall Rule": "FW_Officer_B",
        "Description": "ใช้งาน Internet ทั่วไปได้ แต่บล็อก Social Media",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_Officer_C": {
        "Policy Name": "Allow YouTube",
        "Internet Level": "Standard + YouTube",
        "Allowed": "Web, Email, Business apps, YouTube",
        "Blocked": "Facebook, TikTok, Instagram",
        "Firewall Rule": "FW_Officer_C",
        "Description": "ใช้งานทั่วไปและ YouTube ได้ แต่ยังบล็อก Social Media อื่น",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_Officer_D": {
        "Policy Name": "Allow Facebook",
        "Internet Level": "Standard + Facebook",
        "Allowed": "Web, Email, Business apps, Facebook",
        "Blocked": "YouTube, TikTok, Instagram",
        "Firewall Rule": "FW_Officer_D",
        "Description": "ใช้งานทั่วไปและ Facebook ได้",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_Officer_E": {
        "Policy Name": "Allow YouTube, Facebook",
        "Internet Level": "Standard + Media",
        "Allowed": "Web, Email, Business apps, YouTube, Facebook",
        "Blocked": "TikTok, Instagram",
        "Firewall Rule": "FW_Officer_E",
        "Description": "ใช้งานทั่วไปพร้อม YouTube และ Facebook ได้",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_Manager": {
        "Policy Name": "Net True",
        "Internet Level": "Manager Access",
        "Allowed": "Web, Email, Business apps, approved management access",
        "Blocked": "ตาม policy firewall",
        "Firewall Rule": "FW_Manager",
        "Description": "Policy สำหรับระดับ Manager",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_IT": {
        "Policy Name": "IT Internet Policy",
        "Internet Level": "IT Admin Access",
        "Allowed": "All standard access, admin tools, remote support, vendor sites",
        "Blocked": "ตาม security baseline",
        "Firewall Rule": "FW_IT",
        "Description": "สิทธิ์ Internet สำหรับทีม IT",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_MD": {
        "Policy Name": "Management Internet Policy",
        "Internet Level": "Management Access",
        "Allowed": "Business apps, web, email, executive-approved services",
        "Blocked": "ตาม policy firewall",
        "Firewall Rule": "FW_MD",
        "Description": "สิทธิ์ Internet สำหรับผู้บริหาร",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_Supervisor_B": {
        "Policy Name": "Supervisor Internet Policy",
        "Internet Level": "Supervisor Access",
        "Allowed": "Web, Email, Business apps, approved team resources",
        "Blocked": "Social media ตามข้อกำหนด",
        "Firewall Rule": "FW_Supervisor_B",
        "Description": "สิทธิ์ Internet สำหรับ Supervisor",
        "Owner": "IT",
        "Last Updated": "",
    },
    "FW_Conference": {
        "Policy Name": "Conference Room Internet Policy",
        "Internet Level": "Meeting Room Access",
        "Allowed": "Meeting apps, web, presentation services",
        "Blocked": "High-risk categories ตาม policy firewall",
        "Firewall Rule": "FW_Conference",
        "Description": "สิทธิ์ Internet สำหรับอุปกรณ์ห้องประชุม",
        "Owner": "IT",
        "Last Updated": "",
    },
}


# Field mapping สำหรับ Computer Asset
COMPUTER_FIELDS = {
    "field_1": "บริษัท",
    "field_3": "ชื่อพนักงาน",
    "field_6": "Hostname",
    "field_7": "Model",
    "field_8": "Serial No.",
    "field_13": "RAM",
    "field_14": "Storage Type",
    "field_15": "Storage C:",
    "field_16": "Storage D:",
    "Status": "Status",
}

# Field mapping สำหรับ Monitor
MONITOR_FIELDS = {
    "field_1": "บริษัท",
    "field_3": "ชื่อพนักงาน",
    "field_2": "Brand/Model",
    "field_4": "Serial No.",
    "Status": "Status",
}

# Field mapping สำหรับ Printer
PRINTER_FIELDS = {
    "Company": "บริษัท",
    "User": "User",
    "Brand_x0020__x002f__x0020_Model": "Brand/Model",
    "S_x002f_N_x0020_No_x002e_": "Serial No.",
    "field_3": "IP Address",
}

COMPANY_OPTIONS = ["OPT", "SWI", "PRP", "PLC", "EGI", "THK"]
STATUS_OPTIONS = ["Active", "Inactive", "Spare", "Repair"]

# Field mapping สำหรับ Ink Stock
INK_STOCK_LIST   = "Ink Stock"        # ชื่อ SharePoint List สำหรับสต็อกหมึก
INK_HISTORY_LIST = "Ink History"      # ชื่อ SharePoint List สำหรับประวัติการเบิก
INK_LOW_THRESHOLD = 3                 # จำนวนต่ำสุดก่อนแจ้งเตือน (default)
INK_COLOR_OPTIONS = ["Black", "Cyan", "Magenta", "Yellow", "Color (Tri)", "Other"]
INK_STOCK_FIELDS = {
    "Title":         "รุ่นหมึก",
    "Color":         "สี",
    "Printer_Model": "รุ่นเครื่องพิมพ์",
    "Company":       "บริษัท",
    "Quantity":      "จำนวนคงเหลือ",
    "Min_Qty":       "จุดแจ้งเตือน",
    "Unit_Price":    "ราคา/ชิ้น (บาท)",
    "Notes":         "หมายเหตุ",
}
INK_HISTORY_FIELDS = {
    "Ink_Title":     "รุ่นหมึก",
    "Color":         "สี",
    "Qty_Change":    "จำนวน (+/-)",
    "Action":        "ประเภท",
    "Requester":     "ผู้เบิก/เพิ่ม",
    "Note":          "หมายเหตุ",
    "Timestamp":     "วันเวลา",
}

# =============================================================================
# SECTION 02 : AUTHENTICATION
# ระบบ Login Microsoft 365 / Cookie / ตรวจสอบสิทธิ์ Admin
# =============================================================================
def get_manager():
    """
    Singleton Cookie Manager

    Returns
    -------
    CookieManager
        ใช้จัดการ Login Cookie และ Session Persistence
    """
    if "cookie_manager" not in st.session_state:
        st.session_state["cookie_manager"] = stx.CookieManager(key="cookie_mgr_singleton")
    return st.session_state["cookie_manager"]

def is_admin(username: str) -> bool:
    """
    ตรวจสอบสิทธิ์ Admin จาก Email

    Parameters
    ----------
    username : str

    Returns
    -------
    bool
    """
    if not username:
        return False
    return any(username.lower() in email.lower() or email.lower() in username.lower()
               for email in ADMIN_EMAILS)

@st.cache_data(ttl=3600)
def get_access_token():
    app = msal.ConfidentialClientApplication(CLIENT_ID, authority=AUTHORITY, client_credential=CLIENT_SECRET)
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    return result.get("access_token")

def check_ms_login(username, password):
    app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY)
    result = app.acquire_token_by_username_password(username, password, scopes=["User.Read", "email"])
    if "access_token" in result:
        claims = result.get("id_token_claims", {})
        name = claims.get("name", username)
        email = claims.get("preferred_username", username)
        return True, name, email
    return False, result.get("error_description", "ล็อกอินไม่สำเร็จ"), ""

# =============================================================================
# SECTION 02.1 : AD / FIREWALL INTERNET POLICY
# ดึง Group Membership จาก Microsoft Entra ID / Active Directory ผ่าน Microsoft Graph
# แล้วแปลง Group ที่ขึ้นต้นด้วย FW_ เป็น Internet Policy
# =============================================================================
def _escape_ldap_filter_value(value: str) -> str:
    """Escape special characters for LDAP filter values."""
    text = str(value or "")
    return (
        text.replace("\\", r"\5c")
            .replace("*", r"\2a")
            .replace("(", r"\28")
            .replace(")", r"\29")
            .replace("\x00", r"\00")
    )


def _extract_cn_from_dn(dn: str) -> str:
    """Extract CN from a distinguishedName like CN=FW_Officer_D,OU=Groups,..."""
    text = str(dn or "").strip()
    if not text:
        return ""
    match = re.match(r"CN=((?:\\.|[^,])+)", text, flags=re.IGNORECASE)
    if not match:
        return text
    return match.group(1).replace(r"\,", ",").replace(r"\\", "\\").strip()


def _normalize_ad_identity(user_identity: str):
    ident = str(user_identity or "").strip()
    if "\\" in ident:
        ident = ident.split("\\")[-1].strip()
    return ident


def _ldap_enabled():
    return bool(AD_LDAP_SERVER and AD_BASE_DN and AD_BIND_USER and AD_BIND_PASSWORD)


def _ad_agent_enabled():
    return bool(AD_AGENT_URL)


def _source_order():
    if AD_POLICY_SOURCE in ("ldap", "agent", "graph"):
        return [AD_POLICY_SOURCE]
    return ["ldap", "agent", "graph"]


def _make_ldap_connection():
    if Server is None or Connection is None:
        raise Exception("ยังไม่ได้ติดตั้ง Python package 'ldap3' ใน environment ที่รันแอป")
    if not _ldap_enabled():
        raise Exception("ยังไม่ได้ตั้งค่า AD_LDAP_SERVER / AD_BASE_DN / AD_BIND_USER / AD_BIND_PASSWORD")

    server = Server(
        AD_LDAP_SERVER,
        port=AD_LDAP_PORT,
        use_ssl=AD_LDAP_USE_SSL,
        get_info=ALL,
        connect_timeout=AD_LDAP_TIMEOUT,
    )
    conn = Connection(
        server,
        user=AD_BIND_USER,
        password=AD_BIND_PASSWORD,
        auto_bind=True,
        receive_timeout=AD_LDAP_TIMEOUT,
    )
    return conn


@st.cache_data(ttl=1800, show_spinner=False)
def ldap_find_user(user_identity: str):
    """Find a user directly from AD LDAP and return core attributes."""
    ident = _normalize_ad_identity(user_identity)
    if not ident:
        return None

    local_part = ident.split("@")[0] if "@" in ident else ident
    safe_ident = _escape_ldap_filter_value(ident)
    safe_local = _escape_ldap_filter_value(local_part)

    filters = [
        f"(userPrincipalName={safe_ident})",
        f"(mail={safe_ident})",
        f"(sAMAccountName={safe_local})",
        f"(cn={safe_ident})",
        f"(displayName={safe_ident})",
        f"(displayName={_escape_ldap_filter_value(ident.replace('.', ' '))})",
    ]
    search_filter = "(&(objectCategory=person)(objectClass=user)(|" + "".join(filters) + "))"
    attributes = [
        "displayName",
        "userPrincipalName",
        "mail",
        "sAMAccountName",
        "distinguishedName",
        "memberOf",
        "department",
        "title",
        "company",
        "enabled",
    ]

    conn = _make_ldap_connection()
    try:
        ok = conn.search(
            search_base=AD_BASE_DN,
            search_filter=search_filter,
            search_scope=SUBTREE,
            attributes=attributes,
            size_limit=1,
        )
        if not ok or not conn.entries:
            return None

        entry = conn.entries[0]
        data = entry.entry_attributes_as_dict
        return {
            "id": str(data.get("distinguishedName", [""])[0] if isinstance(data.get("distinguishedName"), list) else data.get("distinguishedName", "")),
            "displayName": str(data.get("displayName", [""])[0] if isinstance(data.get("displayName"), list) and data.get("displayName") else data.get("displayName", "")),
            "userPrincipalName": str(data.get("userPrincipalName", [""])[0] if isinstance(data.get("userPrincipalName"), list) and data.get("userPrincipalName") else data.get("userPrincipalName", "")),
            "mail": str(data.get("mail", [""])[0] if isinstance(data.get("mail"), list) and data.get("mail") else data.get("mail", "")),
            "sAMAccountName": str(data.get("sAMAccountName", [""])[0] if isinstance(data.get("sAMAccountName"), list) and data.get("sAMAccountName") else data.get("sAMAccountName", "")),
            "department": str(data.get("department", [""])[0] if isinstance(data.get("department"), list) and data.get("department") else data.get("department", "")),
            "title": str(data.get("title", [""])[0] if isinstance(data.get("title"), list) and data.get("title") else data.get("title", "")),
            "company": str(data.get("company", [""])[0] if isinstance(data.get("company"), list) and data.get("company") else data.get("company", "")),
            "memberOf": data.get("memberOf", []) or [],
        }
    finally:
        conn.unbind()


@st.cache_data(ttl=1800, show_spinner=False)
def get_ldap_group_names_for_user(user_identity: str):
    """Return direct AD memberOf group names from Domain Controller."""
    user_obj = ldap_find_user(user_identity)
    if not user_obj:
        return []
    groups = [_extract_cn_from_dn(dn) for dn in user_obj.get("memberOf", [])]
    return sorted({g for g in groups if g}, key=lambda x: x.lower())


@st.cache_data(ttl=900, show_spinner=False)
def get_ad_agent_policy_summary(user_identity: str):
    """Read user/group/policy data from an internal AD Agent API if configured."""
    if not _ad_agent_enabled():
        raise Exception("ยังไม่ได้ตั้งค่า AD_AGENT_URL")

    headers = {}
    if AD_AGENT_TOKEN:
        headers["X-API-Token"] = AD_AGENT_TOKEN
        headers["Authorization"] = f"Bearer {AD_AGENT_TOKEN}"

    resp = requests.get(
        f"{AD_AGENT_URL}/user-policy",
        headers=headers,
        params={"user": user_identity},
        timeout=AD_LDAP_TIMEOUT,
    )
    try:
        data = resp.json()
    except Exception:
        data = {"error": resp.text[:300]}
    if resp.status_code >= 400:
        raise Exception(f"AD Agent HTTP {resp.status_code}: {data}")
    if not isinstance(data, dict):
        raise Exception("AD Agent response format ไม่ถูกต้อง")

    groups = data.get("groups", []) or []

    # รองรับ AD Agent ทั้ง 2 รูปแบบ:
    # 1) {"user": {...}, "groups": [...], "policies": [...]}
    # 2) {"displayName": "...", "mail": "...", "groups": [...], "internet_policy": [...]}
    user_obj = data.get("user") if isinstance(data.get("user"), dict) else {}
    if not user_obj:
        user_obj = {
            "displayName": data.get("displayName", ""),
            "userPrincipalName": data.get("userPrincipalName", ""),
            "mail": data.get("mail", ""),
            "sAMAccountName": data.get("sAMAccountName", ""),
            "department": data.get("department", ""),
            "title": data.get("title", ""),
            "company": data.get("company", ""),
        }

    policies = data.get("policies") or get_internet_policies_from_groups(groups)

    return {
        "ok": bool(data.get("ok", True)),
        "source": "AD Agent",
        "user": user_obj,
        "groups": groups,
        "policies": policies,
        "error": data.get("error", ""),
    }



@st.cache_data(ttl=900, show_spinner=False)
def get_ad_agent_policy_users(policy_name: str):
    """Read all users who are members of a firewall policy group from AD Agent.

    Expected AD Agent endpoint:
    GET {AD_AGENT_URL}/policy-users?policy=FW_Officer_A
    Response:
    {
      "ok": true,
      "policy": "FW_Officer_A",
      "description": "Allow All Website",
      "users": [{"displayName": "...", "mail": "...", "sAMAccountName": "..."}]
    }
    """
    if not _ad_agent_enabled():
        raise Exception("ยังไม่ได้ตั้งค่า AD_AGENT_URL")

    policy = str(policy_name or "").strip()
    if not policy:
        raise Exception("กรุณาระบุชื่อ Policy เช่น FW_Officer_A")

    headers = {}
    if AD_AGENT_TOKEN:
        headers["X-API-Token"] = AD_AGENT_TOKEN
        headers["Authorization"] = f"Bearer {AD_AGENT_TOKEN}"

    resp = requests.get(
        f"{AD_AGENT_URL}/policy-users",
        headers=headers,
        params={"policy": policy},
        timeout=AD_LDAP_TIMEOUT,
    )
    try:
        data = resp.json()
    except Exception:
        data = {"error": resp.text[:300]}

    if resp.status_code == 404:
        raise Exception("AD Agent ยังไม่มี endpoint /policy-users — กรุณาอัปเดต ad_agent.py บน NAS เป็นเวอร์ชันที่รองรับค้นหา Policy")
    if resp.status_code >= 400:
        raise Exception(f"AD Agent HTTP {resp.status_code}: {data}")
    if not isinstance(data, dict):
        raise Exception("AD Agent response format ไม่ถูกต้อง")

    users = data.get("users", []) or []
    if not isinstance(users, list):
        users = []

    return {
        "ok": bool(data.get("ok", True)),
        "source": "AD Agent",
        "policy": data.get("policy", policy),
        "description": data.get("description", FW_POLICY_MAP.get(policy, "")),
        "users": users,
        "count": data.get("count", len(users)),
        "error": data.get("error", ""),
    }


def get_policy_users_summary(policy_name: str):
    """คืนรายชื่อ User ทั้งหมดที่ได้ Policy นี้."""
    errors = []
    policy = str(policy_name or "").strip()

    if not policy:
        return {
            "ok": False,
            "source": "",
            "policy": "",
            "description": "",
            "users": [],
            "count": 0,
            "error": "กรุณาระบุชื่อ Policy",
        }

    # Streamlit Cloud ควรใช้ AD Agent เป็นหลัก เพราะเข้า IP ภายใน/LDAP ตรงไม่ได้
    try:
        if _ad_agent_enabled():
            result = get_ad_agent_policy_users(policy)
            if result.get("ok"):
                return result
            errors.append(f"AD Agent: {result.get('error', '')}")
        else:
            errors.append("AD Agent skipped: ยังไม่ได้ตั้งค่า AD_AGENT_URL")
    except Exception as e:
        errors.append(f"AD Agent: {e}")

    return {
        "ok": False,
        "source": "",
        "policy": policy,
        "description": FW_POLICY_MAP.get(policy, ""),
        "users": [],
        "count": 0,
        "error": " | ".join(errors) if errors else "ไม่พบข้อมูล Policy",
    }


def _escape_graph_filter_value(value: str) -> str:
    """Escape single quote for Microsoft Graph OData filter."""
    return str(value or "").replace("'", "''").strip()


def _graph_get(url, *, headers=None, params=None, timeout=30):
    """เรียก Microsoft Graph แบบรวม Error ให้ดูง่าย"""
    token = get_access_token()
    req_headers = {"Authorization": f"Bearer {token}"}
    if headers:
        req_headers.update(headers)
    res = requests.get(url, headers=req_headers, params=params, timeout=timeout)
    try:
        data = res.json()
    except Exception:
        data = {"error": {"message": res.text[:300]}}
    if res.status_code >= 400:
        msg = data.get("error", {}).get("message", str(data)) if isinstance(data, dict) else str(data)
        raise Exception(f"Graph HTTP {res.status_code}: {msg}")
    return data


@st.cache_data(ttl=1800, show_spinner=False)
def graph_find_user(user_identity: str):
    """หา User จาก UPN / Email / Display Name / Account name

    คืนค่า dict ที่มี id, displayName, userPrincipalName, mail
    """
    ident = str(user_identity or "").strip()
    if not ident:
        return None

    select_cols = "id,displayName,userPrincipalName,mail,mailNickname,jobTitle,department,companyName"

    # 1) ถ้าเป็น email/upn ให้เรียกตรงก่อน
    if "@" in ident:
        try:
            return _graph_get(
                f"{GRAPH_URL}/users/{ident}",
                params={"$select": select_cols},
            )
        except Exception:
            pass

    # 2) ถ้าเป็น account name เช่น Ratchaphruek.Ro ให้ลองเติม domain ที่ใช้ login อยู่
    login_email = st.session_state.get("user_email", "") if hasattr(st, "session_state") else ""
    login_domain = login_email.split("@")[-1] if "@" in login_email else ""
    if login_domain and "@" not in ident and " " not in ident:
        try:
            return _graph_get(
                f"{GRAPH_URL}/users/{ident}@{login_domain}",
                params={"$select": select_cols},
            )
        except Exception:
            pass

    # 3) ค้นหาจาก displayName / UPN / mailNickname
    safe = _escape_graph_filter_value(ident)
    filters = [
        f"startswith(displayName,'{safe}')",
        f"startswith(userPrincipalName,'{safe}')",
        f"startswith(mailNickname,'{safe}')",
    ]

    # ถ้าเป็นชื่อแบบมีจุด ให้ลองแปลงจุดเป็นเว้นวรรคด้วย
    if "." in ident:
        safe_space = _escape_graph_filter_value(ident.replace(".", " "))
        filters.append(f"startswith(displayName,'{safe_space}')")

    for flt in filters:
        try:
            data = _graph_get(
                f"{GRAPH_URL}/users",
                params={"$select": select_cols, "$top": 5, "$filter": flt},
            )
            users = data.get("value", []) if isinstance(data, dict) else []
            if users:
                return users[0]
        except Exception:
            continue

    return None


@st.cache_data(ttl=1800, show_spinner=False)
def get_ad_group_names_for_user(user_identity: str):
    """คืนรายชื่อ AD / Entra ID Groups ที่ User เป็นสมาชิกอยู่

    ต้องให้ App Registration มีสิทธิ์ Microsoft Graph อย่างน้อย:
    - User.Read.All
    - GroupMember.Read.All หรือ Directory.Read.All
    และกด Admin consent แล้ว
    """
    user_obj = graph_find_user(user_identity)
    if not user_obj or not user_obj.get("id"):
        return []

    groups = []
    url = f"{GRAPH_URL}/users/{user_obj['id']}/transitiveMemberOf/microsoft.graph.group"
    params = {"$select": "displayName", "$top": 999}

    while url:
        data = _graph_get(url, params=params)
        for item in data.get("value", []):
            name = str(item.get("displayName", "")).strip()
            if name:
                groups.append(name)
        url = data.get("@odata.nextLink")
        params = None

    return sorted(set(groups), key=lambda x: x.lower())


def _first_policy_value(row, aliases, default=""):
    """Return the first non-empty value from possible SharePoint column names."""
    for key in aliases:
        value = row.get(key, "") if hasattr(row, "get") else ""
        value = str(value or "").strip()
        if value and value.lower() not in ("nan", "none", "-"):
            return value
    return default


def _default_firewall_policy_rows():
    rows = []
    for group_name, description in FW_POLICY_MAP.items():
        detail = dict(FW_POLICY_DEFAULT_DETAILS.get(group_name, {}))
        rows.append({
            "AD Group": group_name,
            "Policy Name": detail.get("Policy Name", description),
            "Internet Level": detail.get("Internet Level", ""),
            "Allowed": detail.get("Allowed", ""),
            "Blocked": detail.get("Blocked", ""),
            "Firewall Rule": detail.get("Firewall Rule", group_name),
            "Description": detail.get("Description", description),
            "Owner": detail.get("Owner", "IT"),
            "Last Updated": detail.get("Last Updated", ""),
            "Source": "Default Mapping",
        })
    return rows


@st.cache_data(ttl=1800, show_spinner=False)
def load_firewall_policy_mapping():
    """Load detailed firewall policy mapping from SharePoint, fallback to defaults."""
    default_rows = _default_firewall_policy_rows()

    try:
        df_map = load_sp_data(FIREWALL_POLICY_MAPPING_LIST)
    except Exception:
        df_map = pd.DataFrame()

    if df_map is None or df_map.empty:
        return default_rows

    rows = []
    for _, row in df_map.iterrows():
        ad_group = _first_policy_value(row, ["AD Group", "ADGroup", "Group", "Group Name", "Title", "field_1"])
        if not ad_group:
            continue

        fallback_description = FW_POLICY_MAP.get(ad_group, "Firewall / Internet policy group")
        rows.append({
            "AD Group": ad_group,
            "Policy Name": _first_policy_value(row, ["Policy Name", "PolicyName", "Policy", "Title"], ad_group),
            "Internet Level": _first_policy_value(row, ["Internet Level", "InternetLevel", "Level", "Access Level", "field_2"]),
            "Allowed": _first_policy_value(row, ["Allowed", "Allow", "Allow List", "AllowList", "Can Access", "field_3"]),
            "Blocked": _first_policy_value(row, ["Blocked", "Block", "Block List", "BlockList", "Deny", "field_4"]),
            "Firewall Rule": _first_policy_value(row, ["Firewall Rule", "FirewallRule", "Rule", "Rule Name", "field_5"], ad_group),
            "Description": _first_policy_value(row, ["Description", "Policy Description", "PolicyDescription", "Detail", "field_6"], fallback_description),
            "Owner": _first_policy_value(row, ["Owner", "Responsible", "Managed By", "ManagedBy", "field_7"], "IT"),
            "Last Updated": _first_policy_value(row, ["Last Updated", "LastUpdated", "Modified", "Updated", "field_8"]),
            "Source": "SharePoint Mapping",
        })

    return rows or default_rows


def get_firewall_policy_mapping_dict():
    return {
        str(row.get("AD Group", "")).strip().lower(): row
        for row in load_firewall_policy_mapping()
        if str(row.get("AD Group", "")).strip()
    }


def get_internet_policies_from_groups(group_names):
    """แปลง AD Groups เป็น Internet Policy rows พร้อมรายละเอียดว่า Policy ทำอะไรได้บ้าง"""
    mapping = get_firewall_policy_mapping_dict()
    policies = []

    for group_name in group_names or []:
        g = str(group_name).strip()
        if not g:
            continue

        is_policy_group = g.lower() in mapping or any(g.upper().startswith(p.upper()) for p in FW_POLICY_PREFIXES)
        if not is_policy_group:
            continue

        detail = dict(mapping.get(g.lower(), {}))
        if not detail:
            detail = {
                "AD Group": g,
                "Policy Name": FW_POLICY_MAP.get(g, g),
                "Internet Level": "",
                "Allowed": "",
                "Blocked": "",
                "Firewall Rule": g,
                "Description": FW_POLICY_MAP.get(g, "Firewall / Internet policy group"),
                "Owner": "IT",
                "Last Updated": "",
                "Source": "AD Group (unmapped)",
            }

        policies.append({
            "Policy Internet": g,
            "AD Group": detail.get("AD Group", g),
            "Policy Name": detail.get("Policy Name", FW_POLICY_MAP.get(g, g)),
            "Internet Level": detail.get("Internet Level", ""),
            "Allowed": detail.get("Allowed", ""),
            "Blocked": detail.get("Blocked", ""),
            "Firewall Rule": detail.get("Firewall Rule", g),
            "Description": detail.get("Description", FW_POLICY_MAP.get(g, "Firewall / Internet policy group")),
            "Owner": detail.get("Owner", ""),
            "Last Updated": detail.get("Last Updated", ""),
            "Source": detail.get("Source", "AD Group"),
        })

    return policies


def get_user_internet_policy_summary(user_identity: str):
    """คืนสรุป Internet Policy ของ User จาก AD / Agent / Graph ตาม source ที่ตั้งค่าไว้"""
    errors = []

    for source in _source_order():
        try:
            if source == "ldap":
                if not _ldap_enabled():
                    errors.append("LDAP skipped: ยังไม่ได้ตั้งค่า LDAP secrets")
                    continue
                user_obj = ldap_find_user(user_identity)
                if not user_obj:
                    errors.append("LDAP: ไม่พบ User ใน AD Server")
                    continue
                groups = get_ldap_group_names_for_user(user_identity)
                policies = get_internet_policies_from_groups(groups)
                return {
                    "ok": True,
                    "source": "AD LDAP",
                    "user": user_obj,
                    "groups": groups,
                    "policies": policies,
                    "error": "",
                }

            if source == "agent":
                if not _ad_agent_enabled():
                    errors.append("AD Agent skipped: ยังไม่ได้ตั้งค่า AD_AGENT_URL")
                    continue
                result = get_ad_agent_policy_summary(user_identity)
                if result.get("ok"):
                    return result
                errors.append(f"AD Agent: {result.get('error', '')}")
                continue

            if source == "graph":
                user_obj = graph_find_user(user_identity)
                if not user_obj or not user_obj.get("id"):
                    errors.append("Graph: ไม่พบ User ใน Entra ID")
                    continue
                groups = get_ad_group_names_for_user(user_identity)
                policies = get_internet_policies_from_groups(groups)
                return {
                    "ok": True,
                    "source": "Microsoft Graph",
                    "user": user_obj,
                    "groups": groups,
                    "policies": policies,
                    "error": "",
                }
        except Exception as e:
            errors.append(f"{source.upper()}: {e}")

    return {
        "ok": False,
        "source": "",
        "user": {},
        "groups": [],
        "policies": [],
        "error": " | ".join(errors) if errors else "ไม่พบข้อมูลจากทุก source",
    }


def format_policy_names(policies):
    if not policies:
        return "-"
    return ", ".join(sorted({p.get("Policy Internet", "") for p in policies if p.get("Policy Internet")}))


def format_policy_descriptions(policies):
    if not policies:
        return "-"
    return ", ".join(sorted({p.get("Description", "") for p in policies if p.get("Description")}))


def format_policy_allowed(policies):
    if not policies:
        return "-"
    return ", ".join(sorted({p.get("Allowed", "") for p in policies if p.get("Allowed")}))


def format_policy_blocked(policies):
    if not policies:
        return "-"
    return ", ".join(sorted({p.get("Blocked", "") for p in policies if p.get("Blocked")}))


def get_asset_user_identity(row, asset_list_name: str = ""):
    """Return the best user identifier from an asset row for AD policy lookup."""
    if row is None:
        return ""

    candidates = []
    list_name = str(asset_list_name or "").lower()
    if "printer" in list_name:
        candidates.extend(["User", "field_3", "Title"])
    else:
        candidates.extend(["field_3", "User", "Title"])

    for key in candidates:
        value = row.get(key, "") if hasattr(row, "get") else ""
        value = str(value or "").strip()
        if value and value.lower() not in ("nan", "none", "-"):
            return value
    return ""

# =============================================================================
# SECTION 03 : SHAREPOINT CRUD
# โหลดข้อมูล / เพิ่ม / แก้ไข / ลบ ข้อมูลจาก SharePoint
# =============================================================================
@st.cache_data(ttl=3600)
def get_sp_site_id():
    token = get_access_token()
    headers = {'Authorization': f'Bearer {token}'}
    res = requests.get(f"{GRAPH_URL}/sites/{SHAREPOINT_DOMAIN}:/sites/{SITE_NAME}", headers=headers).json()
    return res.get('id')

@st.cache_data(ttl=3600)
def get_sp_list_id(list_name):
    token = get_access_token()
    headers = {'Authorization': f'Bearer {token}'}
    site_id = get_sp_site_id()
    res = requests.get(f"{GRAPH_URL}/sites/{site_id}/lists", headers=headers).json()
    target = next((l for l in res.get('value', []) if l['displayName'] == list_name), None)
    return target['id'] if target else None

@st.cache_data(ttl=3600)
def load_sp_data(target_display_name):
    """
    โหลดข้อมูลจาก SharePoint List

    Parameters
    ----------
    target_display_name : str

    Returns
    -------
    pandas.DataFrame
    """
    token = get_access_token()
    headers = {'Authorization': f'Bearer {token}'}
    try:
        site_id = get_sp_site_id()
        list_id = get_sp_list_id(target_display_name)
        if list_id:
            items_res = requests.get(
                f"{GRAPH_URL}/sites/{site_id}/lists/{list_id}/items?expand=fields&$top=999",
                headers=headers
            ).json()
            rows = []
            for item in items_res.get('value', []):
                fields = item['fields']
                fields['_item_id'] = item['id']  # เก็บ item ID สำหรับ CRUD
                rows.append(fields)
            return pd.DataFrame(rows)
    except Exception as e:
        print(f"Error fetching SP data: {e}")
    return pd.DataFrame()

def sp_create_item(list_name, fields_dict):
    token = get_access_token()
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    site_id = get_sp_site_id()
    list_id = get_sp_list_id(list_name)
    res = requests.post(
        f"{GRAPH_URL}/sites/{site_id}/lists/{list_id}/items",
        headers=headers,
        json= {"fields": fields_dict}
    )
    return res.status_code in (200, 201), res.json()

def sp_update_item(list_name, item_id, fields_dict):
    token = get_access_token()
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    site_id = get_sp_site_id()
    list_id = get_sp_list_id(list_name)
    res = requests.patch(
        f"{GRAPH_URL}/sites/{site_id}/lists/{list_id}/items/{item_id}/fields",
        headers=headers,
        json=fields_dict
    )
    return res.status_code in (200, 204), res.json() if res.content else {}

def sp_delete_item(list_name, item_id):
    token = get_access_token()
    headers = {'Authorization': f'Bearer {token}'}
    site_id = get_sp_site_id()
    list_id = get_sp_list_id(list_name)
    res = requests.delete(
        f"{GRAPH_URL}/sites/{site_id}/lists/{list_id}/items/{item_id}",
        headers=headers
    )
    return res.status_code in (200, 204, 200)

def clear_sp_cache():
    load_sp_data.clear()

# =============================================================================
# SECTION 04 : INK STOCK HELPERS
# ฟังก์ชันจัดการสต็อกหมึกและประวัติการเบิก
# =============================================================================
def ink_create(fields_dict):
    return sp_create_item(INK_STOCK_LIST, fields_dict)

def ink_update(item_id, fields_dict):
    return sp_update_item(INK_STOCK_LIST, item_id, fields_dict)

def ink_delete(item_id):
    return sp_delete_item(INK_STOCK_LIST, item_id)

def ink_adjust_quantity(item_id, current_qty, delta, title, color,
                        requester, note, action_label):
    """เพิ่ม/ลดสต็อกหมึกและบันทึก history ใน SharePoint"""
    new_qty = max(0, current_qty + delta)
    ok, _ = ink_update(item_id, {"Quantity": new_qty})
    if ok:
        from datetime import datetime
        history_fields = {
            "Ink_Title":  title,
            "Color":      color,
            "Qty_Change": delta,
            "Action":     action_label,
            "Requester":  requester,
            "Note":       note,
            "Timestamp":  datetime.now().strftime("%Y-%m-%d %H:%M"),
}
        sp_create_item(INK_HISTORY_LIST, history_fields)
        load_sp_data.clear()
    return ok, new_qty

# =============================================================================
# SECTION 05 : PASSWORD EXCEL
# อ่าน/เขียนไฟล์ Password.xlsx บน SharePoint
# =============================================================================
def parse_password_sheet(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return pd.DataFrame()
    raw_headers = list(all_rows[0])
    valid_col_idx = [i for i, h in enumerate(raw_headers) if h is not None]
    headers = [str(raw_headers[i]).strip() for i in valid_col_idx]
    data = []
    for row in all_rows[1:]:
        vals = [row[i] if i < len(row) else None for i in valid_col_idx]
        non_null = [v for v in vals if v is not None and str(v).strip() != '']
        if non_null:
            data.append([str(v).strip() if v is not None else None for v in vals])
    if not data:
        return pd.DataFrame()
    return pd.DataFrame(data, columns=headers)

@st.cache_data(ttl=1800)
def load_password_excel():
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    try:
        site_id = get_sp_site_id()
        drive_res = requests.get(f"{GRAPH_URL}/sites/{site_id}/drive", headers=headers).json()
        drive_id = drive_res.get("id")
        file_url = f"{GRAPH_URL}/drives/{drive_id}/root:/{SHAREPOINT_FOLDER}/{PASSWORD_FILE_NAME}:/content"
        file_res = requests.get(file_url, headers=headers)
        if file_res.status_code == 200:
            wb = load_workbook(io.BytesIO(file_res.content))
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                df = parse_password_sheet(ws)
                if not df.empty:
                    sheets[sheet_name] = df
            return sheets, drive_id
        return {"_error": f"HTTP {file_res.status_code} - ไม่พบไฟล์ที่ path: {SHAREPOINT_FOLDER}/{PASSWORD_FILE_NAME}", "_url": file_url}, drive_id
    except Exception as e:
        return {"_error": str(e)}, None

def upload_password_excel(drive_id, sheets_dict):
    """
    สร้าง Excel ใหม่จาก sheets_dict แล้วอัปโหลดทับไฟล์เดิมบน SharePoint
    """
    token = get_access_token()
    headers_auth = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/octet-stream'}

    # ดาวน์โหลดไฟล์เดิมมาก่อนเพื่อรักษา format
    site_id = get_sp_site_id()
    dl_headers = {'Authorization': f'Bearer {token}'}
    file_res = requests.get(
        f"{GRAPH_URL}/drives/{drive_id}/root:/{SHAREPOINT_FOLDER}/{PASSWORD_FILE_NAME}:/content",
        headers=dl_headers
    )
    wb = load_workbook(io.BytesIO(file_res.content))

    for sheet_name, df in sheets_dict.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # ลบ data rows เดิม (เก็บ header row 1)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        # เขียนข้อมูลใหม่
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val if val not in ('None', 'nan', '') else None)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    upload_res = requests.put(
        f"{GRAPH_URL}/drives/{drive_id}/root:/{SHAREPOINT_FOLDER}/{PASSWORD_FILE_NAME}:/content",
        headers=headers_auth,
        data=output.getvalue()
    )
    return upload_res.status_code in (200, 201)

# =============================================================================
# SECTION 06 : NAS CONNECTOR
# เชื่อมต่อ Synology NAS ได้ 2 แบบ
#   1) Synology DSM API ผ่าน Cloudflare Tunnel  ← แนะนำสำหรับ Streamlit Cloud
#   2) SSH แบบเดิม                                ← ใช้เมื่อเปิด SSH/VPN เท่านั้น
# =============================================================================

def _nas_api_enabled():
    """คืนค่า True เฉพาะเมื่อเลือกใช้ DSM API จริง ๆ

    หมายเหตุ: ถ้า NAS_MODE=agent ให้ปิด DSM API fallback เพื่อไม่ให้ไปเรียก
    nas-api.poonyaruk.co.th ซึ่งเป็น DSM route และทำให้ timeout บน Streamlit Cloud
    """
    if AGENT_ONLY_MODE:
        return False
    return NAS_MODE == "api" and bool(NAS_BASE_URL)


def _nas_agent_enabled():
    """คืนค่า True เมื่อกำหนด NAS_AGENT_URL เพื่ออ่าน ACL จริงผ่าน NAS Local API Agent"""
    return AGENT_ONLY_MODE or bool(NAS_AGENT_URL)


def _safe_json_response(resp, context="NAS API"):
    """แปลง Response เป็น JSON พร้อม Error ที่อ่านง่าย"""
    try:
        return resp.json()
    except Exception:
        preview = resp.text[:300] if getattr(resp, "text", None) else ""
        raise Exception(f"{context} ไม่ได้ตอบกลับเป็น JSON | HTTP {resp.status_code} | {preview}")


def synology_api_info():
    """ทดสอบว่า Synology WebAPI ผ่าน Cloudflare Tunnel ใช้งานได้หรือไม่"""
    if not NAS_BASE_URL:
        raise Exception("ยังไม่ได้ตั้งค่า NAS_BASE_URL ใน Streamlit secrets")

    resp = requests.get(
        f"{NAS_BASE_URL}/webapi/query.cgi",
        params={
            "api": "SYNO.API.Info",
            "version": "1",
            "method": "query",
            "query": "all",
        },
        timeout=NAS_TIMEOUT,
    )
    data = _safe_json_response(resp, "SYNO.API.Info")
    if not data.get("success"):
        raise Exception(f"SYNO.API.Info failed: {data}")
    return data


def synology_login(session="FileStation"):
    """Login DSM API แล้วคืนค่า SID"""
    if not NAS_BASE_URL:
        raise Exception("ยังไม่ได้ตั้งค่า NAS_BASE_URL ใน Streamlit secrets")
    if not NAS_USER or not NAS_PASSWORD:
        raise Exception("ยังไม่ได้ตั้งค่า NAS_USER / NAS_PASSWORD ใน Streamlit secrets")

    resp = requests.get(
        f"{NAS_BASE_URL}/webapi/auth.cgi",
        params={
            "api": "SYNO.API.Auth",
            "version": "7",
            "method": "login",
            "account": NAS_USER,
            "passwd": NAS_PASSWORD,
            "session": session,
            "format": "sid",
        },
        timeout=NAS_TIMEOUT,
    )
    data = _safe_json_response(resp, "SYNO.API.Auth")
    if not data.get("success"):
        raise Exception(f"NAS API Login failed: {data}")
    return data.get("data", {}).get("sid", "")


def synology_logout(sid, session="FileStation"):
    """Logout DSM API แบบ best-effort"""
    if not sid or not NAS_BASE_URL:
        return
    try:
        requests.get(
            f"{NAS_BASE_URL}/webapi/auth.cgi",
            params={
                "api": "SYNO.API.Auth",
                "version": "7",
                "method": "logout",
                "session": session,
                "_sid": sid,
            },
            timeout=10,
        )
    except Exception:
        pass


def synology_get_shares_api():
    """ดึงรายชื่อ Shared Folder ผ่าน Synology FileStation API"""
    sid = synology_login("FileStation")
    try:
        resp = requests.get(
            f"{NAS_BASE_URL}/webapi/entry.cgi",
            params={
                "api": "SYNO.FileStation.List",
                "version": "2",
                "method": "list_share",
                "_sid": sid,
            },
            timeout=NAS_TIMEOUT,
        )
        data = _safe_json_response(resp, "SYNO.FileStation.List")
        if not data.get("success"):
            raise Exception(f"list_share failed: {data}")

        shares = []
        for item in data.get("data", {}).get("shares", []):
            name = item.get("name") or item.get("path", "").strip("/")
            if name and not name.startswith("@"):
                shares.append(name)
        return sorted(set(shares))
    finally:
        synology_logout(sid, "FileStation")


def load_nas_data_api():
    """
    โหลดรายชื่อ Share ผ่าน DSM API เท่านั้น
    ใช้เป็น fallback เมื่อยังไม่ได้ตั้งค่า NAS Agent
    """
    synology_api_info()
    shares = synology_get_shares_api()

    rows = []
    for share in shares:
        rows.append({
            "Share": share,
            "ACL Tags (Raw)": "เชื่อมต่อผ่าน DSM API สำเร็จ — อ่าน Raw ACL ต้องใช้ NAS Agent หรือ SSH/synoacltool",
            "Matched Employees": "",
        })

    return pd.DataFrame(rows, columns=["Share", "ACL Tags (Raw)", "Matched Employees"]).sort_values("Share")


def nas_agent_health():
    """ตรวจสอบ NAS Local API Agent และยืนยันว่าเป็นเวอร์ชัน SSH"""
    if not NAS_AGENT_URL:
        raise Exception("ยังไม่ได้ตั้งค่า NAS_AGENT_URL ใน Streamlit secrets")

    resp = requests.get(
        f"{NAS_AGENT_URL}/health",
        timeout=NAS_TIMEOUT,
        headers={"Cache-Control": "no-cache"},
        params={"_": datetime.datetime.now().timestamp()},
    )
    data = _safe_json_response(resp, "NAS Agent /health")

    if data.get("status") != "ok":
        raise Exception(f"NAS Agent health failed: {data}")

    service_name = str(data.get("service", ""))
    if service_name and service_name != "nas-agent-ssh":
        raise Exception(
            "NAS Agent ยังเป็นเวอร์ชันเก่า "
            f"(service={service_name}) — กรุณา restart/recreate container ให้เป็น nas-agent-ssh"
        )

    return data


def nas_agent_get_shares():
    """ดึงรายชื่อ Shared Folder จาก NAS Agent (/shares)

    ถ้า Agent ยังไม่มี endpoint /shares สามารถใส่ Secrets เพิ่มได้:
    NAS_SHARES = "Share1,Share2,Share3"
    """
    if not NAS_AGENT_URL:
        raise Exception("ยังไม่ได้ตั้งค่า NAS_AGENT_URL ใน Streamlit secrets")
    if not NAS_AGENT_TOKEN:
        raise Exception("ยังไม่ได้ตั้งค่า NAS_AGENT_TOKEN ใน Streamlit secrets")

    headers = {"X-API-Token": NAS_AGENT_TOKEN}
    resp = requests.get(
        f"{NAS_AGENT_URL}/shares",
        headers=headers,
        timeout=NAS_TIMEOUT,
    )

    if resp.status_code == 404:
        if NAS_SHARES_SECRET:
            return sorted([x.strip() for x in NAS_SHARES_SECRET.split(",") if x.strip()])
        raise Exception(
            "NAS Agent ยังไม่มี endpoint /shares — กรุณาอัปเดต nas_agent.py หรือเพิ่ม NAS_SHARES ใน Secrets"
        )

    data = _safe_json_response(resp, "NAS Agent /shares")
    if resp.status_code != 200:
        raise Exception(f"NAS Agent /shares HTTP {resp.status_code}: {data}")

    shares = data.get("shares") or data.get("data") or []
    if not isinstance(shares, list):
        raise Exception(f"NAS Agent /shares รูปแบบข้อมูลไม่ถูกต้อง: {data}")

    return sorted(set(str(x).strip() for x in shares if str(x).strip() and not str(x).strip().startswith("@")))


def nas_agent_get_acl_payload(share):
    """ดึง Permission จาก NAS Agent

    รองรับ Agent เวอร์ชันใหม่ที่มี /share-permissions ก่อน
    ถ้า Agent ยังไม่มี endpoint นี้ จะ fallback เป็น /acl แบบเดิม
    """
    if not NAS_AGENT_URL:
        raise Exception("ยังไม่ได้ตั้งค่า NAS_AGENT_URL ใน Streamlit secrets")
    if not NAS_AGENT_TOKEN:
        raise Exception("ยังไม่ได้ตั้งค่า NAS_AGENT_TOKEN ใน Streamlit secrets")

    headers = {"X-API-Token": NAS_AGENT_TOKEN}

    # Agent ใหม่: คืน permissions ที่แยก Read/Write จาก Synology share privilege แล้ว
    for endpoint in ("share-permissions", "permissions"):
        try:
            resp = requests.get(
                f"{NAS_AGENT_URL}/{endpoint}",
                params={"share": share},
                headers=headers,
                timeout=NAS_TIMEOUT,
            )
            if resp.status_code == 404:
                continue
            data = _safe_json_response(resp, f"NAS Agent /{endpoint} share={share}")
            if resp.status_code != 200:
                raise Exception(f"NAS Agent /{endpoint} HTTP {resp.status_code}: {data}")
            return data
        except requests.exceptions.HTTPError:
            continue
        except Exception:
            # ถ้า endpoint ใหม่ยังไม่พร้อม ให้ลอง /acl ต่อ
            pass

    # Agent เดิม: คืน raw synoacltool
    resp = requests.get(
        f"{NAS_AGENT_URL}/acl",
        params={"share": share},
        headers=headers,
        timeout=NAS_TIMEOUT,
    )
    data = _safe_json_response(resp, f"NAS Agent /acl share={share}")

    if resp.status_code != 200:
        raise Exception(f"NAS Agent /acl HTTP {resp.status_code}: {data}")

    if data.get("returncode", 0) not in (0, None):
        raise Exception(f"synoacltool failed for {share}: {data.get('stderr', '')}")

    return data


def nas_agent_get_acl_raw(share):
    """ดึง Raw ACL จาก NAS Agent ซึ่งรัน synoacltool บน NAS"""
    data = nas_agent_get_acl_payload(share)
    return data.get("stdout") or data.get("acl") or ""



def _clean_nas_principal(name: str) -> str:
    """Normalize Synology ACL principal names.

    Examples:
    - OPTIMALGROUP\\Sasithorn.Su -> Sasithorn.Su
    - user:OPTIMALGROUP\\IT_Network -> IT_Network
    - group:administrators -> administrators
    """
    if name is None:
        return ""

    cleaned = str(name).strip()
    if not cleaned or cleaned.lower() in ("nan", "none", "null"):
        return ""

    # Remove optional prefix that may already be included in parsed text.
    cleaned = re.sub(r"^(user|group)\s*:\s*", "", cleaned, flags=re.I).strip()

    # Synology / AD principals are often DOMAIN\\name.
    if "\\" in cleaned:
        cleaned = cleaned.split("\\")[-1].strip()

    # Remove accidental surrounding quotes/spaces.
    cleaned = cleaned.strip().strip('"').strip("'").strip()

    return cleaned

def parse_nas_agent_permissions_payload(payload, employee_list=None):
    """แปลง payload จาก Agent เวอร์ชันใหม่เป็น ACL Tags และ Matched Employees

    payload ที่รองรับ:
    {
      "permissions": [
        {"entity": "User", "type": "user", "permission": "Read"},
        {"entity": "Group", "type": "group", "permission": "Read/Write"}
      ]
    }
    """
    employee_list = employee_list or []
    rows = payload.get("permissions") or payload.get("data") or []
    if not isinstance(rows, list) or not rows:
        return [], []

    acl_entries = []
    matched_employees = []

    for item in rows:
        if not isinstance(item, dict):
            continue
        entity = _clean_nas_principal(item.get("entity") or item.get("name") or item.get("principal") or "")
        if not entity:
            continue
        ptype = str(item.get("type") or item.get("kind") or "user").lower()
        perm_raw = str(item.get("permission") or item.get("access") or item.get("perm") or "").strip().lower()
        raw_blob = str(item.get("raw_permission") or item.get("permission_blob") or "").strip()

        # ถ้า Agent ส่ง raw_permission มาด้วย ให้แยก Read/Write จาก permission blob จริง
        # Read-only ของ Synology มักเป็น r-x---a-R-c-- ซึ่งมี c แต่ไม่ควรนับเป็น Write
        if raw_blob:
            if any(ch in raw_blob for ch in set("wpdDWo")):
                permission = "Read/Write"
            else:
                permission = "Read"
        elif perm_raw in ("rw", "readwrite", "read/write", "write", "read-write", "read_write"):
            permission = "Read/Write"
        elif perm_raw in ("ro", "read", "readonly", "read-only", "read_only"):
            permission = "Read"
        elif perm_raw in ("deny", "no", "na", "noaccess", "no access"):
            permission = "Deny"
        else:
            permission = "Read/Write" if "write" in perm_raw or perm_raw == "rw" else "Read"

        acl_entries.append(f"{ptype}:{entity} ({permission})")

        for emp in employee_list:
            emp_text = str(emp).strip()
            if emp_text and (entity.lower() in emp_text.lower() or emp_text.lower() in entity.lower()):
                matched_employees.append(f"{emp_text} ({permission})")

    return sorted(set(acl_entries), key=lambda x: x.lower()), sorted(set(matched_employees))


def parse_synoacl_output(raw, employee_list=None):
    """แปลงผลลัพธ์ synoacltool เป็น ACL Tags และ Matched Employees"""
    employee_list = employee_list or []

    if not raw:
        return [], []

    acl_entries = []

    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue

        # ตัวอย่าง:
        # [0] user:ActiveBackup:allow:rwxpdDaARWc--:fd-- (level:0)
        # [4] group:OPTIMALGROUP\Domain Admins:allow:rwxpdDaARWc--:fd-- (level:0)
        m = re.search(
            r'(?:\[\d+\]\s*)?(user|group):(.+?):(allow|deny):([^:\s]+)',
            line,
            flags=re.I,
        )
        if not m:
            continue

        kind = m.group(1).lower()
        principal_name = _clean_nas_principal(m.group(2))
        action = m.group(3).lower()
        # ห้าม lower() permission blob เพราะตัวพิมพ์เล็ก/ใหญ่ของ synoacltool มีความหมายต่างกัน
        # ตัวอย่าง Read-only มักมี r/x/a/R/c ซึ่งเดิมโดนจับเป็น RW เพราะมีตัว a
        # จึงนับเป็น Read/Write เฉพาะสิทธิ์ที่เกี่ยวกับการเขียนจริง ๆ เท่านั้น
        perm_blob = m.group(4).strip()

        if not principal_name:
            continue

        writable_flags = set("wpdDWo")

        if action == "deny":
            permission = "Deny"
        elif any(ch in perm_blob for ch in writable_flags):
            permission = "Read/Write"
        else:
            permission = "Read"

        acl_entries.append(f"{kind}:{principal_name} ({permission})")

    acl_tags = sorted(set(acl_entries), key=lambda x: x.lower())

    matched_employees = []
    for entry in acl_tags:
        match = re.search(r"^(.*?)\s*\((Read(?:/Write)?|Deny)\)", entry)
        if not match:
            continue

        clean_entity = _clean_nas_principal(match.group(1))
        if not clean_entity:
            continue

        for emp in employee_list:
            emp_text = str(emp).strip()
            if not emp_text:
                continue
            if clean_entity.lower() in emp_text.lower() or emp_text.lower() in clean_entity.lower():
                matched_employees.append(f"{emp_text} ({match.group(2).strip()})")

    return acl_tags, sorted(set(matched_employees), key=lambda x: x.lower())


def fetch_acl_agent(share, employee_list):
    """อ่าน Permission ของ Share ผ่าน NAS Agent"""
    try:
        payload = nas_agent_get_acl_payload(share)

        # ถ้า Agent ใหม่ส่ง permissions มา ให้ใช้ผลนี้ก่อน เพราะแยก Read/Write จาก Synology share privilege ได้ตรงกว่า ACL ดิบ
        if isinstance(payload, dict) and (payload.get("permissions") or payload.get("data")):
            tags, matched = parse_nas_agent_permissions_payload(payload, employee_list)
        else:
            raw = payload.get("stdout") or payload.get("acl") or ""
            tags, matched = parse_synoacl_output(raw, employee_list)

        return share, tags, matched
    except Exception as e:
        # เก็บ error ลง raw เพื่อให้ดูรายละเอียดใน Popup ได้ ไม่ทำให้ทั้งหน้าล่ม
        return share, [f"NAS Agent Error: {e}"], []


def load_nas_data_agent():
    """โหลด NAS Data แบบสมบูรณ์: Shares จาก DSM API + ACL จาก NAS Agent"""
    nas_agent_health()

    # ใช้ NAS Agent ดึงรายชื่อ Share ก่อน เพื่อไม่ต้องพึ่ง DSM route (nas-api)
    try:
        shares = nas_agent_get_shares()
    except Exception as agent_share_e:
        if _nas_api_enabled():
            shares = synology_get_shares_api()
        else:
            raise Exception(str(agent_share_e))

    df_emp = load_sp_data("Employees")
    employees = df_emp['field_3'].dropna().unique().tolist() if not df_emp.empty and 'field_3' in df_emp.columns else []

    data = []
    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = [executor.submit(fetch_acl_agent, s, employees) for s in shares]
        for future in as_completed(futures):
            share, tags, matched_emps = future.result()
            data.append({
                "Share": share,
                "ACL Tags (Raw)": ", ".join(sorted(tags, key=lambda x: x.lower())),
                "Matched Employees": ", ".join(sorted(matched_emps, key=lambda x: x.lower())),
            })

    return pd.DataFrame(data).sort_values("Share")


def create_ssh():
    """
    สร้าง SSH Connection ไปยัง Synology NAS

    ใช้เฉพาะกรณี:
    - NAS เปิด SSH ผ่าน VPN/Port Forward/Cloudflare Access TCP แล้ว
    - ต้องการอ่าน ACL ผ่าน synoacltool แบบละเอียด
    """

    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    try:
        ssh.connect(
            hostname=NAS_IP,
            port=NAS_PORT,
            username=SSH_USER,
            password=SSH_PWD,
            timeout=NAS_TIMEOUT,
            banner_timeout=NAS_TIMEOUT,
            auth_timeout=NAS_TIMEOUT
        )

        return ssh

    except Exception as e:
        raise Exception(
            f"NAS SSH Connection Failed | "
            f"Host={NAS_IP} Port={NAS_PORT} | {str(e)}"
        )


def run_command(ssh, cmd):
    full_cmd = f"sudo -S {cmd}"
    stdin, stdout, stderr = ssh.exec_command(full_cmd)
    stdin.write(SSH_PWD + "\n")
    stdin.flush()
    return stdout.read().decode('utf-8', errors='ignore'), stderr.read().decode('utf-8', errors='ignore')


def check_synoacl():
    try:
        ssh = create_ssh()
        output, _ = run_command(ssh, "ls /usr/syno/bin/synoacltool")
        ssh.close()
        return "/usr/syno/bin/synoacltool" in output
    except Exception as e:
        st.warning(f"⚠️ ไม่สามารถเชื่อมต่อ NAS ผ่าน SSH ได้: {e}")
        return False


def get_shares():
    try:
        ssh = create_ssh()
        output, _ = run_command(ssh, "ls /volume1")
        ssh.close()
        return [l.strip() for l in output.splitlines() if l.strip() and not l.startswith("@")]
    except Exception:
        return []


def fetch_acl(share, employee_list):
    try:
        ssh = create_ssh()
        raw, _ = run_command(ssh, f"/usr/syno/bin/synoacltool -get /volume1/{share}")
        ssh.close()
        if not raw:
            return share, [], []
        tags, matched = parse_synoacl_output(raw, employee_list)
        return share, tags, matched
    except Exception:
        return share, [], []


@st.cache_data(ttl=1800)
def load_nas_data():
    """โหลดข้อมูล NAS โดยเลือก NAS Agent ก่อน แล้ว fallback เป็น DSM API หรือ SSH"""
    if _nas_agent_enabled():
        try:
            return load_nas_data_agent()
        except Exception as e:
            st.warning(f"⚠️ ไม่สามารถเชื่อมต่อ NAS Agent ได้: {e}")
            # ถ้าเลือก NAS_MODE=agent ห้าม fallback ไป DSM API route เดิม เพราะจะทำให้ timeout ซ้ำ
            if AGENT_ONLY_MODE:
                return None
            if _nas_api_enabled():
                try:
                    return load_nas_data_api()
                except Exception as api_e:
                    st.warning(f"⚠️ ไม่สามารถเชื่อมต่อ NAS ผ่าน DSM API ได้: {api_e}")
                    return None
            return None

    if _nas_api_enabled():
        try:
            return load_nas_data_api()
        except Exception as e:
            st.warning(f"⚠️ ไม่สามารถเชื่อมต่อ NAS ผ่าน DSM API ได้: {e}")
            return None

    if not check_synoacl():
        return None
    df_emp = load_sp_data("Employees")
    employees = df_emp['field_3'].dropna().unique().tolist() if not df_emp.empty and 'field_3' in df_emp.columns else []
    shares = get_shares()
    data = []
    with ThreadPoolExecutor(max_workers=MAX_THREADS) as executor:
        futures = [executor.submit(fetch_acl, s, employees) for s in shares]
        for future in as_completed(futures):
            share, tags, matched_emps = future.result()
            data.append({"Share": share, "ACL Tags (Raw)": ", ".join(sorted(tags)), "Matched Employees": ", ".join(sorted(matched_emps))})
    return pd.DataFrame(data).sort_values("Share")


def get_nas_connection_status():
    """ใช้แสดงสถานะ NAS แบบสั้น ๆ ใน Dashboard/หน้า NAS"""
    if _nas_agent_enabled():
        try:
            health = nas_agent_health()
            return True, f"NAS Agent Connected: {NAS_AGENT_URL} ({health.get('service', '-')})"
        except Exception as e:
            return False, f"NAS Agent Failed: {e}"

    if _nas_api_enabled():
        try:
            synology_api_info()
            return True, f"DSM API Connected: {NAS_BASE_URL}"
        except Exception as e:
            return False, f"DSM API Failed: {e}"
    try:
        ssh = create_ssh()
        ssh.close()
        return True, f"SSH Connected: {NAS_IP}:{NAS_PORT}"
    except Exception as e:
        return False, str(e)

# =============================================================================
# SECTION 07 : CARD RENDER
# หน้าตาการ์ด Computers / Monitors / Printers
# =============================================================================
def _hw_badge(status):
    cls = {"Active":"badge-active","Inactive":"badge-inactive","Spare":"badge-spare","Repair":"badge-repair"}.get(status,"badge-default")
    return f'<span class="badge {cls}">{status or "—"}</span>' if status else ''


# =============================================================================
# FUNCTION : render_card_computer
# UI OWNER   : Asset Management > Computers
# PURPOSE    : สร้างการ์ดแสดงข้อมูล Computer 1 รายการ
# DATA FLOW  : SharePoint -> row -> HTML Card -> Streamlit
# CSS OWNER  : HARDWARE_THEME / Card CSS
# =============================================================================
def render_card_computer(row, key, admin_mode):
    status = row.get('Status', '')
    with st.container():
        st.markdown(f"""
        <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:4px;">
            <div>
                <div class="hw-card-title">👤 {row.get('field_3','N/A')}</div>
                <div class="hw-card-sub">🏢 {row.get('field_1','-')}</div>
            </div>
            {_hw_badge(status)}
        </div>
        <div class="hw-field"><strong>💻 Hostname</strong>&nbsp;&nbsp;{row.get('field_6','-')}</div>
        <div class="hw-field"><strong>🏷️ Model</strong>&nbsp;&nbsp;{row.get('field_7','-')}</div>
        <div class="hw-field"><strong>💾 RAM</strong>&nbsp;&nbsp;{row.get('field_13','-')}</div>
        """, unsafe_allow_html=True)
        if admin_mode:
            c1, c2 = st.columns(2)
            with c1:
                if st.button("🔍 ดูข้อมูล", key=f"comp_view_{key}", use_container_width=True):
                    show_pop_computer(row.to_dict(), admin_mode=True)
            with c2:
                if st.button("✏️ แก้ไข", key=f"comp_edit_{key}", use_container_width=True):
                    st.session_state[f"edit_computer_{key}"] = True
                    st.rerun()
        else:
            st.caption("🔒 ข้อมูลเพิ่มเติมและการแก้ไขเฉพาะผู้ดูแลระบบ")


# =============================================================================
# FUNCTION : render_card_monitor
# UI OWNER   : Asset Management > Monitors
# PURPOSE    : สร้างการ์ดแสดงข้อมูล Monitor 1 รายการ
# DATA FLOW  : SharePoint -> row -> HTML Card -> Streamlit
# CSS OWNER  : HARDWARE_THEME / Card CSS
# =============================================================================
def render_card_monitor(row, key, admin_mode):
    status = row.get('Status', '')
    with st.container():
        st.markdown(f"""
        <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:4px;">
            <div>
                <div class="hw-card-title">👤 {row.get('field_3','N/A')}</div>
                <div class="hw-card-sub">🏢 {row.get('field_1','-')}</div>
            </div>
            {_hw_badge(status)}
        </div>
        <div class="hw-field"><strong>🖥️ Model</strong>&nbsp;&nbsp;{row.get('field_2','-')}</div>
        {'<div class="hw-field"><strong>🔢 Serial No.</strong>&nbsp;&nbsp;%s</div>' % row.get('field_4','-') if admin_mode else ''}
        """, unsafe_allow_html=True)
        if admin_mode:
            c1, c2 = st.columns(2)
            with c1:
                if st.button("🔍 ดูข้อมูล", key=f"mon_view_{key}", use_container_width=True):
                    show_pop_monitor(row.to_dict(), admin_mode=True)
            with c2:
                if st.button("✏️ แก้ไข", key=f"mon_edit_{key}", use_container_width=True):
                    st.session_state[f"edit_monitor_{key}"] = True
                    st.rerun()
        else:
            st.caption("🔒 ดูข้อมูลเชิงลึกและแก้ไขเฉพาะผู้ดูแลระบบ")


# =============================================================================
# FUNCTION : render_card_printer
# UI OWNER   : Asset Management > Printers
# PURPOSE    : สร้างการ์ดแสดงข้อมูล Printer 1 รายการ
# DATA FLOW  : SharePoint -> row -> HTML Card -> Streamlit
# CSS OWNER  : HARDWARE_THEME / Card CSS
# =============================================================================
def render_card_printer(row, key, admin_mode):
    with st.container():
        st.markdown(f"""
        <div style="margin-bottom:6px;">
            <div class="hw-card-title">🖨️ {row.get('Brand_x0020__x002f__x0020_Model','Printer')}</div>
            <div class="hw-card-sub">🏢 {row.get('field_1','-')}</div>
        </div>
        <div class="hw-field"><strong>👤 User</strong>&nbsp;&nbsp;{row.get('User','-')}</div>
        {'<div class="hw-field"><strong>🔢 Serial No.</strong>&nbsp;&nbsp;%s</div>' % row.get('S_x002f_N_x0020_No_x002e_','-') if admin_mode else ''}
        {'<div class="hw-field"><strong>🌐 IP</strong>&nbsp;&nbsp;%s</div>' % row.get('field_3','-') if admin_mode else ''}
        """, unsafe_allow_html=True)
        if admin_mode:
            c1, c2 = st.columns(2)
            with c1:
                if st.button("🔍 ดูข้อมูล", key=f"prn_view_{key}", use_container_width=True):
                    show_pop_printer(row.to_dict(), admin_mode=True)
            with c2:
                if st.button("✏️ แก้ไข", key=f"prn_edit_{key}", use_container_width=True):
                    st.session_state[f"edit_printer_{key}"] = True
                    st.rerun()
        else:
            st.caption("🔒 ดูข้อมูลเชิงลึกและแก้ไขเฉพาะผู้ดูแลระบบ")

# =============================================================================
# SECTION 08 : VIEW DIALOGS
# Popup แสดงรายละเอียด Asset
# =============================================================================
@st.dialog("📋 รายละเอียด Computer")
def show_pop_computer(data, admin_mode=False):
    st.markdown(f"### 💻 {data.get('field_7', 'Computer')}")
    c1, c2 = st.columns(2)
    with c1:
        st.write(f"**👤 พนักงาน:** {data.get('field_3', '-')}")
        st.write(f"**🏢 บริษัท:** {data.get('field_1', '-')}")
        st.write(f"**💻 Hostname:** {data.get('field_6', '-')}")
    with c2:
        if admin_mode:
            st.write(f"**🔢 Serial No:** {data.get('field_8', '-')}")
        else:
            st.write("**🔢 Serial No:** 🔒 ซ่อนสำหรับผู้ใช้ทั่วไป")
        st.write(f"**✳️ สถานะ:** {data.get('Status', '-')}")
        st.write(f"**💾 RAM:** {data.get('field_13', '-')}")
        st.write(f"**💿 Storage C:** {data.get('field_15', '-')}  D: {data.get('field_16', '-')}")
    with st.expander("📊 ดูข้อมูลดิบ"):
        st.json(data)

@st.dialog("📋 รายละเอียด Monitor")
def show_pop_monitor(data, admin_mode=False):
    st.markdown(f"### 🖥️ {data.get('field_2', 'Monitor')}")
    st.write(f"**👤 พนักงาน:** {data.get('field_3', '-')}")
    st.write(f"**🏢 บริษัท:** {data.get('field_1', '-')}")
    if admin_mode:
        st.write(f"**🔢 Serial No.:** {data.get('field_4', '-')}")
    else:
        st.write("**🔢 Serial No.:** 🔒 ซ่อนสำหรับผู้ใช้ทั่วไป")
    st.write(f"**✅ สถานะ:** {data.get('Status', '-')}")
    with st.expander("📊 ดูข้อมูลดิบ"):
        st.json(data)

@st.dialog("📋 รายละเอียด Printer")
def show_pop_printer(data, admin_mode=False):
    st.markdown(f"### 🖨️ {data.get('field_2', 'Printer')}")
    st.write(f"**🏢 บริษัท:** {data.get('field_1', '-')}")
    st.write(f"**👤 User:** {data.get('User', '-')}")
    if admin_mode:
        st.write(f"**🔢 Serial No.:** {data.get('S_x002f_N_x0020_No_x002e_', '-')}")
        st.write(f"**🌐 IP Address:** {data.get('field_3', '-')}")
    else:
        st.write("**🔢 Serial No.:** 🔒 ซ่อนสำหรับผู้ใช้ทั่วไป")
        st.write("**🌐 IP Address:** 🔒 ซ่อนสำหรับผู้ใช้ทั่วไป")
    with st.expander("📊 ดูข้อมูลดิบ"):
        st.json(data)

# =============================================================================
# SECTION 09 : EDIT DIALOGS
# Popup แก้ไขข้อมูล Asset
# =============================================================================
@st.dialog("✏️ แก้ไข Computer Asset")
def edit_computer_dialog(row, list_name):
    st.markdown(f"### ✏️ แก้ไข: {row.get('field_3', '')}")
    item_id = row.get('_item_id')
    login_account = st.text_input("LoginAccount", value=row.get('LoginAccount', ''))
    department = st.text_input("Department", value=row.get('field_4', ''))
    operating_system = st.text_input("Operating System", value=row.get('field_10', ''))
    _computer_known_fields = {"field_1","field_3","field_4","field_6","field_7","field_8","field_10","field_13","field_14","field_15","field_16"}
    _computer_extra_fields = {}
    _computer_extra_names = sorted(
        (key for key in row if re.fullmatch(r"field_\d+", str(key)) and key not in _computer_known_fields),
        key=lambda key: int(str(key).split("_")[-1])
    )
    if _computer_extra_names:
        with st.expander("Additional SharePoint fields", expanded=False):
            for _field_name in _computer_extra_names:
                _field_value = row.get(_field_name, '')
                if isinstance(_field_value, (str, int, float, bool)) or _field_value is None:
                    _computer_extra_fields[_field_name] = st.text_input(
                        _field_name,
                        value="" if _field_value is None else str(_field_value),
                        key=f"computer_extra_{item_id}_{_field_name}"
                    )
    company = st.selectbox("🏢 บริษัท", COMPANY_OPTIONS, index=COMPANY_OPTIONS.index(row.get('field_1', 'OPT')) if row.get('field_1') in COMPANY_OPTIONS else 0)
    emp_name = st.text_input("👤 ชื่อพนักงาน", value=row.get('field_3', ''))
    hostname = st.text_input("💻 Hostname", value=row.get('field_6', ''))
    model = st.text_input("🏷️ Model", value=row.get('field_7', ''))
    serial = st.text_input("🔢 Serial No.", value=row.get('field_8', ''))
    ram = st.text_input("💾 RAM", value=row.get('field_13', ''))
    storage_type = st.text_input("💿 Storage Type", value=row.get('field_14', ''))
    storage_c = st.text_input("💿 Storage C:", value=row.get('field_15', ''))
    storage_d = st.text_input("💿 Storage D:", value=row.get('field_16', ''))
    status = st.selectbox("✳️ Status", STATUS_OPTIONS, index=STATUS_OPTIONS.index(row.get('Status', 'Active')) if row.get('Status') in STATUS_OPTIONS else 0)
    col_save, col_del = st.columns(2)
    with col_save:
        if st.button("💾 บันทึก", use_container_width=True, type="primary"):
            fields = {"field_1": company, "field_3": emp_name,
                      "LoginAccount": login_account, "field_4": department,
                      "field_10": operating_system, "field_6": hostname,
                      "field_7": model, "field_8": serial, "field_13": ram,
                      "field_14": storage_type, "field_15": storage_c,
                      "field_16": storage_d, "Status": status}
            fields.update(_computer_extra_fields)
            ok, res_data = sp_update_item(list_name, item_id, fields)
            if ok:
                st.success("✅ บันทึกสำเร็จ")
                clear_sp_cache()
                st.rerun()
            else:
                err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
                st.error(f"❌ บันทึกไม่สำเร็จ: {err_msg}")
    with col_del:
        if st.button("🗑️ ลบรายการนี้", use_container_width=True):
            st.session_state['confirm_delete'] = item_id
    if st.session_state.get('confirm_delete') == item_id:
        st.warning("⚠️ ยืนยันการลบ?")
        if st.button("✅ ยืนยันลบ", type="primary"):
            ok = sp_delete_item(list_name, item_id)
            if ok:
                st.success("✅ ลบสำเร็จ")
                st.session_state.pop('confirm_delete', None)
                clear_sp_cache()
                st.rerun()
            else:
                st.error("❌ ลบไม่สำเร็จ")

@st.dialog("🗑️ ยืนยันการลบ Computer Asset")
def delete_computer_dialog(row, list_name):
    """UI confirmation only; deletion still uses the existing SharePoint CRUD helper."""
    item_id = row.get('_item_id')
    computer_name = row.get('field_6') or row.get('field_3') or '-'
    st.warning(f"ต้องการลบ **{computer_name}** ออกจาก Computer Asset ใช่หรือไม่?")
    st.caption("การลบไม่สามารถย้อนกลับได้")
    cancel_col, delete_col = st.columns(2)
    with cancel_col:
        if st.button("ยกเลิก", use_container_width=True):
            st.rerun()
    with delete_col:
        if st.button("🗑️ ยืนยันลบ", type="primary", use_container_width=True):
            if sp_delete_item(list_name, item_id):
                clear_sp_cache()
                st.success("ลบรายการสำเร็จ")
                st.rerun()
            else:
                st.error("ลบรายการไม่สำเร็จ")

@st.dialog("✏️ แก้ไข Monitor")
def edit_monitor_dialog(row, list_name):
    st.markdown(f"### ✏️ แก้ไข: {row.get('field_2', '')}")
    item_id = row.get('_item_id')
    company = st.selectbox("🏢 บริษัท", COMPANY_OPTIONS, index=COMPANY_OPTIONS.index(row.get('field_1', 'OPT')) if row.get('field_1') in COMPANY_OPTIONS else 0)
    emp_name = st.text_input("👤 ชื่อพนักงาน", value=row.get('field_3', ''))
    model = st.text_input("🏷️ Brand/Model", value=row.get('field_2', ''))
    serial = st.text_input("🔢 Serial No.", value=row.get('field_4', ''))
    status = st.selectbox("✅ Status", STATUS_OPTIONS, index=STATUS_OPTIONS.index(row.get('Status', 'Active')) if row.get('Status') in STATUS_OPTIONS else 0)
    col_save, col_del = st.columns(2)
    with col_save:
        if st.button("💾 บันทึก", use_container_width=True, type="primary"):
            fields = {"field_1": company, "field_3": emp_name, "field_2": model, "field_4": serial, "Status": status}
            ok, res_data = sp_update_item(list_name, item_id, fields)
            if ok:
                st.success("✅ บันทึกสำเร็จ")
                clear_sp_cache()
                st.rerun()
            else:
                err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
            st.error(f"❌ บันทึกไม่สำเร็จ: {err_msg}")
    with col_del:
        if st.button("🗑️ ลบรายการนี้", use_container_width=True):
            st.session_state['confirm_delete'] = item_id
    if st.session_state.get('confirm_delete') == item_id:
        st.warning("⚠️ ยืนยันการลบ?")
        if st.button("✅ ยืนยันลบ", type="primary"):
            ok = sp_delete_item(list_name, item_id)
            if ok:
                st.success("✅ ลบสำเร็จ")
                st.session_state.pop('confirm_delete', None)
                clear_sp_cache()
                st.rerun()
            else:
                st.error("❌ ลบไม่สำเร็จ")

@st.dialog("✏️ แก้ไข Printer")
def edit_printer_dialog(row, list_name):
    st.markdown(f"### ✏️ แก้ไข Printer")
    item_id = row.get('_item_id')
    company = st.selectbox("🏢 บริษัท", COMPANY_OPTIONS, index=COMPANY_OPTIONS.index(row.get('field_1', 'OPT')) if row.get('field_1') in COMPANY_OPTIONS else 0)
    user = st.text_input("👤 User", value=row.get('User', ''))
    model = st.text_input("🏷️ Brand/Model", value=row.get('Brand_x0020__x002f__x0020_Model', ''))
    serial = st.text_input("🔢 Serial No.", value=row.get('S_x002f_N_x0020_No_x002e_', ''))
    ip = st.text_input("🌐 IP Address", value=row.get('field_3', ''))
    col_save, col_del = st.columns(2)
    with col_save:
        if st.button("💾 บันทึก", use_container_width=True, type="primary"):
            fields = {"field_1": company, "User": user,
                      "Brand_x0020__x002f__x0020_Model": model,
                      "S_x002f_N_x0020_No_x002e_": serial, "field_3": ip}
            ok, res_data = sp_update_item(list_name, item_id, fields)
            if ok:
                st.success("✅ บันทึกสำเร็จ")
                clear_sp_cache()
                st.rerun()
            else:
                err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
            st.error(f"❌ บันทึกไม่สำเร็จ: {err_msg}")
    with col_del:
        if st.button("🗑️ ลบรายการนี้", use_container_width=True):
            st.session_state['confirm_delete'] = item_id
    if st.session_state.get('confirm_delete') == item_id:
        st.warning("⚠️ ยืนยันการลบ?")
        if st.button("✅ ยืนยันลบ", type="primary"):
            ok = sp_delete_item(list_name, item_id)
            if ok:
                st.success("✅ ลบสำเร็จ")
                st.session_state.pop('confirm_delete', None)
                clear_sp_cache()
                st.rerun()
            else:
                st.error("❌ ลบไม่สำเร็จ")

# =============================================================================
# SECTION 10 : ADD DIALOGS
# Popup เพิ่มข้อมูล Asset
# =============================================================================
@st.dialog("➕ เพิ่ม Computer Asset")
def add_computer_dialog(list_name):
    login_account = st.text_input("LoginAccount")
    department = st.text_input("Department")
    operating_system = st.text_input("Operating System")
    st.markdown("### ➕ เพิ่มอุปกรณ์ใหม่")
    company = st.selectbox("🏢 บริษัท", COMPANY_OPTIONS)
    emp_name = st.text_input("👤 ชื่อพนักงาน")
    hostname = st.text_input("💻 Hostname")
    model = st.text_input("🏷️ Model")
    serial = st.text_input("🔢 Serial No.")
    ram = st.text_input("💾 RAM")
    storage_type = st.text_input("💿 Storage Type")
    storage_c = st.text_input("💿 Storage C:")
    storage_d = st.text_input("💿 Storage D:")
    status = st.selectbox("✳️ Status", STATUS_OPTIONS)
    if st.button("💾 บันทึก", use_container_width=True, type="primary"):
        fields = {"field_1": company, "field_3": emp_name,
                  "LoginAccount": login_account, "field_4": department,
                  "field_10": operating_system, "field_6": hostname,
                  "field_7": model, "field_8": serial, "field_13": ram,
                  "field_14": storage_type, "field_15": storage_c,
                  "field_16": storage_d, "Status": status}
        ok, res_data = sp_create_item(list_name, fields)
        if ok:
            st.success("✅ เพิ่มสำเร็จ")
            clear_sp_cache()
            st.rerun()
        else:
            err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
            st.error(f"❌ เพิ่มไม่สำเร็จ: {err_msg}")

@st.dialog("➕ เพิ่ม Monitor")
def add_monitor_dialog(list_name):
    st.markdown("### ➕ เพิ่ม Monitor ใหม่")
    company = st.selectbox("🏢 บริษัท", COMPANY_OPTIONS)
    emp_name = st.text_input("👤 ชื่อพนักงาน")
    model = st.text_input("🏷️ Brand/Model")
    serial = st.text_input("🔢 Serial No.")
    status = st.selectbox("✅ Status", STATUS_OPTIONS)
    if st.button("💾 บันทึก", use_container_width=True, type="primary"):
        fields = {"field_1": company, "field_3": emp_name, "field_2": model, "field_4": serial, "Status": status}
        ok, res_data = sp_create_item(list_name, fields)
        if ok:
            st.success("✅ เพิ่มสำเร็จ")
            clear_sp_cache()
            st.rerun()
        else:
            err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
            st.error(f"❌ เพิ่มไม่สำเร็จ: {err_msg}")

@st.dialog("➕ เพิ่ม Printer")
def add_printer_dialog(list_name):
    st.markdown("### ➕ เพิ่ม Printer ใหม่")
    company = st.selectbox("🏢 บริษัท", COMPANY_OPTIONS)
    user = st.text_input("👤 User")
    model = st.text_input("🏷️ Brand/Model")
    serial = st.text_input("🔢 Serial No.")
    ip = st.text_input("🌐 IP Address")
    if st.button("💾 บันทึก", use_container_width=True, type="primary"):
        fields = {"field_1": company, "User": user,
                  "Brand_x0020__x002f__x0020_Model": model,
                  "S_x002f_N_x0020_No_x002e_": serial, "field_3": ip}
        ok, res_data = sp_create_item(list_name, fields)
        if ok:
            st.success("✅ เพิ่มสำเร็จ")
            clear_sp_cache()
            st.rerun()
        else:
            err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
            st.error(f"❌ เพิ่มไม่สำเร็จ: {err_msg}")

# =============================================================================
# SECTION 11 : PASSWORD MANAGER
# แสดง Password Card / เพิ่ม / แก้ไข / ลบ
# =============================================================================
def get_sheet_icon(sheet_name):
    icon_map = {"server": "🖥️", "network": "🌐", "sql": "🗄️",
                "software": "📦", "license": "📦", "domain": "🌍",
                "email": "📧", "mail": "📧", "internet": "📡",
                "wifi": "📶", "vpn": "🔒", "firewall": "🔥"}
    s = sheet_name.lower()
    return next((v for k, v in icon_map.items() if k in s), "🔑")

def is_secret_field(col_name):
    return any(k in str(col_name).lower() for k in ['pass', 'pwd', 'secret', 'key', 'รหัส', 'token'])

def render_password_card(row, sheet_name, card_key, admin_mode, df_pw, drive_id, pw_sheets):
    cols_list = list(row.index)
    icon = get_sheet_icon(sheet_name)
    title_col = cols_list[1] if len(cols_list) > 1 else cols_list[0]
    title_val = str(row.get(title_col, 'N/A'))
    if title_val in ('None', '', 'nan'):
        title_val = str(row.get(cols_list[0], 'N/A'))

    normal_fields = [(c, row[c]) for c in cols_list if not is_secret_field(c) and pd.notna(row[c]) and str(row[c]).strip() not in ('', 'None', 'nan')]
    secret_fields = [(c, row[c]) for c in cols_list if is_secret_field(c) and pd.notna(row[c]) and str(row[c]).strip() not in ('', 'None', 'nan')]

    with st.container():
        # header
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:10px;">
            <div style="width:36px;height:36px;border-radius:10px;background:linear-gradient(135deg,#ede9fe,#ddd6fe);
                        display:flex;align-items:center;justify-content:center;font-size:1.1rem;flex-shrink:0;">
                {icon}
            </div>
            <div>
                <div style="font-size:0.95rem;font-weight:700;color:#1e293b;">{title_val}</div>
                <div style="font-size:0.72rem;color:#94a3b8;">{sheet_name}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # normal fields
        if normal_fields:
            fields_html = ""
            for col_name, val in normal_fields:
                fields_html += f'<div class="pw-field-row"><span class="pw-field-label">{col_name}</span><span class="pw-field-value">{val}</span></div>'
            st.markdown(fields_html, unsafe_allow_html=True)

        # secret fields
        if secret_fields:
            
            for col_name, val in secret_fields:
                with st.expander(f"🔐 {col_name}"):
                    st.code(str(val), language=None)

        if admin_mode:
            
            c1, c2 = st.columns(2)
            with c1:
                if st.button("✏️ แก้ไข", key=f"pw_edit_{sheet_name}_{card_key}", use_container_width=True):
                    st.session_state[f"pw_edit_row_{sheet_name}_{card_key}"] = True
                    st.rerun()
            with c2:
                if st.button("🗑️ ลบ", key=f"pw_del_{sheet_name}_{card_key}", use_container_width=True):
                    st.session_state[f"pw_del_confirm_{sheet_name}_{card_key}"] = True

            if st.session_state.get(f"pw_del_confirm_{sheet_name}_{card_key}"):
                st.warning("⚠️ ยืนยันการลบ?")
                if st.button("✅ ยืนยันลบ", key=f"pw_del_ok_{sheet_name}_{card_key}", type="primary"):
                    new_df = df_pw.drop(index=card_key).reset_index(drop=True)
                    pw_sheets[sheet_name] = new_df
                    ok = upload_password_excel(drive_id, pw_sheets)
                    if ok:
                        st.success("✅ ลบสำเร็จ")
                        load_password_excel.clear()
                        st.session_state.pop(f"pw_del_confirm_{sheet_name}_{card_key}", None)
                        st.rerun()
                    else:
                        st.error("❌ ลบไม่สำเร็จ")

@st.dialog("✏️ แก้ไข Password Entry")
def edit_password_dialog(row, row_idx, sheet_name, df_pw, drive_id, pw_sheets):
    st.markdown(f"### ✏️ แก้ไขรายการ")
    new_vals = {}
    for col in df_pw.columns:
        cur_val = str(row.get(col, '')) if pd.notna(row.get(col)) else ''
        if is_secret_field(col):
            new_vals[col] = st.text_input(f"🔐 {col}", value=cur_val, type="password")
        else:
            new_vals[col] = st.text_input(col, value=cur_val)
    if st.button("💾 บันทึก", use_container_width=True, type="primary"):
        for col, val in new_vals.items():
            df_pw.at[row_idx, col] = val if val.strip() != '' else None
        pw_sheets[sheet_name] = df_pw
        ok = upload_password_excel(drive_id, pw_sheets)
        if ok:
            st.success("✅ บันทึกสำเร็จ")
            load_password_excel.clear()
            st.rerun()
        else:
            err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
            st.error(f"❌ บันทึกไม่สำเร็จ: {err_msg}")

@st.dialog("➕ เพิ่ม Password Entry")
def add_password_dialog(sheet_name, df_pw, drive_id, pw_sheets):
    st.markdown(f"### ➕ เพิ่มรายการใหม่ — {sheet_name}")
    new_vals = {}
    for col in df_pw.columns:
        if is_secret_field(col):
            new_vals[col] = st.text_input(f"🔐 {col}", type="password")
        else:
            new_vals[col] = st.text_input(col)
    if st.button("💾 บันทึก", use_container_width=True, type="primary"):
        new_row = {col: (val if val.strip() != '' else None) for col, val in new_vals.items()}
        new_df = pd.concat([df_pw, pd.DataFrame([new_row])], ignore_index=True)
        pw_sheets[sheet_name] = new_df
        ok = upload_password_excel(drive_id, pw_sheets)
        if ok:
            st.success("✅ เพิ่มสำเร็จ")
            load_password_excel.clear()
            st.rerun()
        else:
            err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
            st.error(f"❌ เพิ่มไม่สำเร็จ: {err_msg}")

# =============================================================================
# SECTION 12 : INK STOCK UI
# Card และ Dialog ของระบบหมึกพิมพ์
# =============================================================================
def ink_stock_color_badge(color):
    color_map = {
        "Black":      ("#222", "#fff"),
        "Cyan":       ("#00b4d8", "#fff"),
        "Magenta":    ("#c77dff", "#fff"),
        "Yellow":     ("#f9c74f", "#333"),
        "Color (Tri)":("#43aa8b", "#fff"),
        "Other":      ("#adb5bd", "#333"),
}
    bg, fg = color_map.get(color, ("#adb5bd", "#333"))
    return f"<span style='background:{bg};color:{fg};padding:3px 12px;border-radius:12px;font-size:0.82em;font-weight:bold;'>{color}</span>"

def ink_qty_badge(qty, min_qty):
    qty = int(qty) if str(qty).isdigit() else 0
    min_qty = int(min_qty) if str(min_qty).isdigit() else INK_LOW_THRESHOLD
    if qty == 0:
        return f"<span style='background:#dc3545;color:#fff;padding:3px 14px;border-radius:12px;font-weight:bold;font-size:0.95em;'>หมด ❌</span>"
    elif qty <= min_qty:
        return f"<span style='background:#fd7e14;color:#fff;padding:3px 14px;border-radius:12px;font-weight:bold;font-size:0.95em;'>⚠️ เหลือ {qty}</span>"
    return f"<span style='background:#198754;color:#fff;padding:3px 14px;border-radius:12px;font-weight:bold;font-size:0.95em;'>✅ {qty}</span>"

def render_ink_card(row, key, admin_mode, requester_name):
    qty     = int(row.get("Quantity", 0)) if str(row.get("Quantity", 0)).lstrip("-").isdigit() else 0
    min_qty = int(row.get("Min_Qty", INK_LOW_THRESHOLD)) if str(row.get("Min_Qty", INK_LOW_THRESHOLD)).isdigit() else INK_LOW_THRESHOLD
    title   = row.get("Title", "N/A")
    color   = row.get("Color", "Other")
    item_id = row.get("_item_id")

    with st.container():
        c_left, c_right = st.columns([0.65, 0.35])
        with c_left:
            st.markdown(f"#### 🖨️ {title}")
            st.markdown(
                ink_stock_color_badge(color) + "&nbsp;&nbsp;" + ink_qty_badge(qty, min_qty),
                unsafe_allow_html=True,
            )
            st.caption(f"🏢 {row.get('Company', '-')}  |  📠 {row.get('Printer_Model', '-')}")
            price = row.get("Unit_Price", "")
            if price and str(price) not in ("None", "nan", ""):
                st.caption(f"💰 {price} บาท/ชิ้น")
        with c_right:
            # ปุ่ม เบิก / เพิ่ม / แก้ไข
            if st.button("📤 เบิก", key=f"ink_withdraw_{key}", use_container_width=True):
                st.session_state[f"ink_withdraw_{key}"] = True
                st.rerun()
            if admin_mode:
                if st.button("📥 เพิ่ม", key=f"ink_add_qty_{key}", use_container_width=True):
                    st.session_state[f"ink_add_qty_{key}"] = True
                    st.rerun()
                if st.button("✏️ แก้ไข", key=f"ink_edit_{key}", use_container_width=True):
                    st.session_state[f"ink_edit_{key}"] = True
                    st.rerun()

    # ---- dialogs triggered by session state ----
    if st.session_state.get(f"ink_withdraw_{key}"):
        st.session_state.pop(f"ink_withdraw_{key}")
        withdraw_ink_dialog(row.to_dict(), key, requester_name)

    if admin_mode and st.session_state.get(f"ink_add_qty_{key}"):
        st.session_state.pop(f"ink_add_qty_{key}")
        add_ink_qty_dialog(row.to_dict(), key, requester_name)

    if admin_mode and st.session_state.get(f"ink_edit_{key}"):
        st.session_state.pop(f"ink_edit_{key}")
        edit_ink_dialog(row.to_dict(), key)

@st.dialog("📤 เบิกหมึกพิมพ์")
def withdraw_ink_dialog(row, key, requester_name):
    qty   = int(row.get("Quantity", 0)) if str(row.get("Quantity", 0)).lstrip("-").isdigit() else 0
    title = row.get("Title", "N/A")
    color = row.get("Color", "-")
    st.markdown(f"### 📤 เบิก: {title}")
    st.markdown(ink_stock_color_badge(color), unsafe_allow_html=True)
    if qty <= 0:
        st.error("❌ หมึกหมดสต็อก ไม่สามารถเบิกได้")
        return
    st.info(f"สต็อกปัจจุบัน: **{qty}** ชิ้น")
    withdraw_qty = st.number_input("จำนวนที่ต้องการเบิก", min_value=1, max_value=qty, value=1, step=1)
    note = st.text_input("หมายเหตุ (เช่น เครื่อง/ชั้น/ห้อง)")
    if st.button("✅ ยืนยันเบิก", use_container_width=True, type="primary"):
        ok, new_qty = ink_adjust_quantity(
            row["_item_id"], qty, -withdraw_qty,
            title, color, requester_name, note, "เบิก"
        )
        if ok:
            st.success(f"✅ เบิกสำเร็จ — สต็อกคงเหลือ: {new_qty} ชิ้น")
            st.rerun()
        else:
            st.error("❌ เบิกไม่สำเร็จ")

@st.dialog("📥 เพิ่มจำนวนสต็อกหมึก")
def add_ink_qty_dialog(row, key, requester_name):
    qty   = int(row.get("Quantity", 0)) if str(row.get("Quantity", 0)).lstrip("-").isdigit() else 0
    title = row.get("Title", "N/A")
    color = row.get("Color", "-")
    st.markdown(f"### 📥 เพิ่มสต็อก: {title}")
    st.markdown(ink_stock_color_badge(color), unsafe_allow_html=True)
    st.info(f"สต็อกปัจจุบัน: **{qty}** ชิ้น")
    add_qty = st.number_input("จำนวนที่ต้องการเพิ่ม", min_value=1, value=1, step=1)
    note = st.text_input("หมายเหตุ (เช่น เลข PO / ผู้ขาย)")
    if st.button("✅ ยืนยันเพิ่มสต็อก", use_container_width=True, type="primary"):
        ok, new_qty = ink_adjust_quantity(
            row["_item_id"], qty, add_qty,
            title, color, requester_name, note, "เพิ่มสต็อก"
        )
        if ok:
            st.success(f"✅ เพิ่มสำเร็จ — สต็อกคงเหลือ: {new_qty} ชิ้น")
            st.rerun()
        else:
            st.error("❌ เพิ่มไม่สำเร็จ")

@st.dialog("✏️ แก้ไขข้อมูลหมึกพิมพ์")
def edit_ink_dialog(row, key):
    item_id = row.get("_item_id")
    st.markdown(f"### ✏️ แก้ไข: {row.get('Title', '')}")
    title      = st.text_input("รุ่นหมึก", value=row.get("Title", ""))
    color      = st.selectbox("สี", INK_COLOR_OPTIONS,
                              index=INK_COLOR_OPTIONS.index(row.get("Color", "Black"))
                              if row.get("Color") in INK_COLOR_OPTIONS else 0)
    printer_m  = st.text_input("รุ่นเครื่องพิมพ์", value=row.get("Printer_Model", ""))
    company    = st.selectbox("บริษัท", ["ALL"] + COMPANY_OPTIONS,
                              index=(["ALL"] + COMPANY_OPTIONS).index(row.get("Company", "ALL"))
                              if row.get("Company") in ["ALL"] + COMPANY_OPTIONS else 0)
    qty        = st.number_input("จำนวนคงเหลือ", min_value=0, value=int(row.get("Quantity", 0)) if str(row.get("Quantity", 0)).isdigit() else 0)
    min_qty    = st.number_input("จุดแจ้งเตือน (Low Stock)", min_value=1, value=int(row.get("Min_Qty", INK_LOW_THRESHOLD)) if str(row.get("Min_Qty", INK_LOW_THRESHOLD)).isdigit() else INK_LOW_THRESHOLD)
    unit_price = st.text_input("ราคา/ชิ้น (บาท)", value=str(row.get("Unit_Price", "")).replace("None","").replace("nan",""))
    notes      = st.text_input("หมายเหตุ", value=str(row.get("Notes", "")).replace("None","").replace("nan",""))
    col_save, col_del = st.columns(2)
    with col_save:
        if st.button("💾 บันทึก", use_container_width=True, type="primary"):
            fields = {"Title": title, "Color": color, "Printer_Model": printer_m,
                      "Company": company, "Quantity": qty, "Min_Qty": min_qty,
                      "Unit_Price": unit_price or None, "Notes": notes or None}
            ok, res_data = ink_update(item_id, fields)
            if ok:
                st.success("✅ บันทึกสำเร็จ")
                clear_sp_cache()
                st.rerun()
            else:
                err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
                st.error(f"❌ บันทึกไม่สำเร็จ: {err_msg}")
    with col_del:
        if st.button("🗑️ ลบรายการนี้", use_container_width=True):
            st.session_state["ink_confirm_delete"] = item_id
    if st.session_state.get("ink_confirm_delete") == item_id:
        st.warning("⚠️ ยืนยันการลบ?")
        if st.button("✅ ยืนยันลบ", type="primary"):
            ok = ink_delete(item_id)
            if ok:
                st.success("✅ ลบสำเร็จ")
                st.session_state.pop("ink_confirm_delete", None)
                clear_sp_cache()
                st.rerun()
            else:
                st.error("❌ ลบไม่สำเร็จ")

@st.dialog("➕ เพิ่มหมึกพิมพ์ใหม่")
def add_ink_dialog():
    st.markdown("### ➕ เพิ่มหมึกพิมพ์ใหม่")
    title      = st.text_input("รุ่นหมึก *", placeholder="เช่น HP 680 Black")
    color      = st.selectbox("สี *", INK_COLOR_OPTIONS)
    printer_m  = st.text_input("รุ่นเครื่องพิมพ์", placeholder="เช่น HP DeskJet 2135")
    company    = st.selectbox("บริษัท", ["ALL"] + COMPANY_OPTIONS)
    qty        = st.number_input("จำนวนเริ่มต้น", min_value=0, value=0)
    min_qty    = st.number_input("จุดแจ้งเตือน (Low Stock)", min_value=1, value=INK_LOW_THRESHOLD)
    unit_price = st.text_input("ราคา/ชิ้น (บาท)", placeholder="เช่น 350")
    notes      = st.text_input("หมายเหตุ")
    if st.button("💾 บันทึก", use_container_width=True, type="primary"):
        if not title.strip():
            st.warning("⚠️ กรุณาระบุรุ่นหมึก")
            return
        fields = {"Title": title.strip(), "Color": color, "Printer_Model": printer_m,
                  "Company": company, "Quantity": qty, "Min_Qty": min_qty,
                  "Unit_Price": unit_price or None, "Notes": notes or None}
        ok, res_data = ink_create(fields)
        if ok:
            st.success("✅ เพิ่มสำเร็จ")
            clear_sp_cache()
            st.rerun()
        else:
            err_msg = res_data.get("error", {}).get("message", str(res_data)) if isinstance(res_data, dict) else str(res_data)
            st.error(f"❌ เพิ่มไม่สำเร็จ: {err_msg}")

# =============================================================================
# SECTION 13 : SIDEBAR BADGES
# ตัวเลขนับรายการใน Sidebar
# =============================================================================
@st.cache_data(ttl=300, show_spinner=False)
def get_sidebar_nav_badges():
    """Cached counts for sidebar badges — display only."""
    badges = {
        "computers": 0, "monitors": 0, "printers": 0, "projector": 0,
        "ups": 0, "misc": 0, "password": 0, "user_perm": 0,
        "ink_stock": 0, "consumables": 0,
    }
    try:
        badges["computers"] = len(load_sp_data("Computer Asset"))
        badges["monitors"] = len(load_sp_data("Asset Monitor"))
        badges["printers"] = len(load_sp_data("Asset Printer"))
        badges["projector"] = len(load_sp_data("Asset Projector"))
        badges["ups"] = len(load_sp_data("Asset UPS"))
        badges["misc"] = len(load_sp_data("Asset Misc"))
        badges["ink_stock"] = len(load_sp_data(INK_STOCK_LIST))
        badges["consumables"] = len(load_sp_data(INK_HISTORY_LIST))
        pw_sheets, _ = load_password_excel()
        if isinstance(pw_sheets, dict) and "_error" not in pw_sheets:
            badges["password"] = sum(
                len(df) for df in pw_sheets.values() if isinstance(df, pd.DataFrame)
            )
        nas_df = load_nas_data()
        badges["user_perm"] = len(nas_df) if nas_df is not None else 0
    except Exception:
        pass
    return badges


@st.cache_data(ttl=900, show_spinner=False)
def get_sidebar_profile_title(user_identity: str):
    """Return the user's current AD job title without changing auth logic."""
    identity = str(user_identity or "").strip()
    if not identity:
        return ""
    if _ad_agent_enabled():
        try:
            summary = get_ad_agent_policy_summary(identity)
            user_obj = summary.get("user", {}) if isinstance(summary, dict) else {}
            title = str(user_obj.get("title") or user_obj.get("jobTitle") or "").strip()
            if title:
                return title
        except Exception:
            pass
    if _ldap_enabled():
        try:
            user_obj = ldap_find_user(identity) or {}
            title = str(user_obj.get("title") or "").strip()
            if title:
                return title
        except Exception:
            pass
    try:
        user_obj = graph_find_user(identity) or {}
        return str(user_obj.get("jobTitle") or user_obj.get("title") or "").strip()
    except Exception:
        return ""


# =============================================================================
# THEME LOADER (SAFE REFACTOR)
# ใช้โหลด Theme หลายชุดพร้อมกัน
# =============================================================================
# =============================================================================
# FUNCTION : load_theme
#
# PURPOSE
#   โหลด CSS Theme หลายชุดพร้อมกัน
#
# EXAMPLE
#   load_theme(MODERN_THEME, SIDEBAR_V32_THEME)
# =============================================================================
def load_theme(*themes):
    st.markdown("\n".join(themes), unsafe_allow_html=True)


# =============================================================================

# =============================================================================
# SIDEBAR NAVIGATION TREE
#
# Dashboard
#
# Reports & Analytics
#
# Administration
# ├─ Asset Management
# │  ├─ Computers
# │  ├─ Monitors
# │  ├─ Printers
# │  ├─ Projectors
# │  ├─ UPS
# │  └─ Miscellaneous
# │
# ├─ Security
# │  ├─ Password Manager
# │  └─ NAS Permission Analyzer
# │
# └─ Inventory
#    ├─ Ink Stock
#    └─ Consumables
#
# =============================================================================

# SECTION 14 : MAIN UI
# Theme / Sidebar / Dashboard / Navigation
# =============================================================================
st.set_page_config(layout="wide", page_title="DocumentReportUnified", page_icon="🛡️")
cookie_manager = get_manager()

# --- AUTH ---
saved_user  = cookie_manager.get(cookie="user_name")
saved_email = cookie_manager.get(cookie="user_email")

if 'is_auth' not in st.session_state:
    st.session_state.is_auth  = False
    st.session_state.user_name  = ""
    st.session_state.user_email = ""
    st.session_state.skip_cookie_login = False

if saved_user and not st.session_state.get('is_auth') and not st.session_state.get('skip_cookie_login'):
    st.session_state.is_auth  = True
    st.session_state.user_name  = saved_user
    st.session_state.user_email = saved_email or ""

# =============================================================================
# THEME : GLOBAL
# ใช้กับทั้งระบบ
# =============================================================================
# ===== MODERN UI THEME =====
# =============================================================================
# THEME : GLOBAL
# ใช้กับทั้งระบบ
# =============================================================================
MODERN_THEME = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=IBM+Plex+Sans+Thai:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"]{font-family:'Inter','IBM Plex Sans Thai',sans-serif !important;}
[data-testid="stAppViewContainer"]{background:radial-gradient(circle at top right, rgba(99,102,241,.10), transparent 22%),radial-gradient(circle at bottom left, rgba(14,165,233,.08), transparent 18%),#eef2ff !important;}
[data-testid="stMetric"]{background:rgba(255,255,255,.84) !important;border:none !important;border-radius:24px !important;padding:1.3rem !important;box-shadow:0 10px 30px rgba(15,23,42,.06) !important;}
[data-testid="stMain"] div[data-testid="stVerticalBlockBorderWrapper"] > div{border:none !important;border-radius:24px !important;background:rgba(255,255,255,.86) !important;box-shadow:0 10px 30px rgba(15,23,42,.05) !important;transition:all .2s ease !important;}
[data-testid="stMain"] div[data-testid="stVerticalBlockBorderWrapper"] > div:hover{transform:translateY(-2px);box-shadow:0 18px 40px rgba(99,102,241,.12) !important;}
[data-testid="stSidebar"] div[data-testid="stVerticalBlockBorderWrapper"] > div{background:transparent !important;border:none !important;box-shadow:none !important;}
.stButton button{border-radius:14px !important;font-weight:700 !important;}
.stTextInput input,.stSelectbox div[data-baseweb="select"] > div{border-radius:14px !important;}
[data-testid="stDataFrame"]{border-radius:22px !important;overflow:hidden !important;}
</style>
"""

# ── BASE CSS (always) ──────────────────────────────────────────────────────────
load_theme(MODERN_THEME,SIDEBAR_V32_THEME)

st.markdown("""
<style>
#MainMenu { visibility: hidden; }
footer { visibility: hidden; }
[data-testid="stHeader"] { background: transparent !important; box-shadow: none !important; }
</style>
""", unsafe_allow_html=True)

# =============================================================================
# THEME : LOGIN
# CSS สำหรับหน้า Login เท่านั้น
# =============================================================================

# ══════════════════════════════════════════════════════════════════════════════
# # THEME BLOCK : LOGIN
# แก้ UI Login ทั้งหมดในช่วงด้านล่างนี้

# LOGIN PAGE
# ══════════════════════════════════════════════════════════════════════════════
if not st.session_state.is_auth:
    # -----------------------------------------------------------------------------
# CSS THEME BLOCK
# ถ้าต้องการแยกเป็น LOGIN_THEME/SIDEBAR_THEME ในอนาคต
# ให้เริ่มจาก Style Block นี้
# -----------------------------------------------------------------------------
    st.markdown("""
    <style>
    /* ===== AURORA GLASS LOGIN ===== */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

    html, body, [class*="css"]{
        font-family:'Inter',sans-serif !important;
}

    [data-testid="stAppViewContainer"]{
        background:
            radial-gradient(circle at top left, rgba(56,189,248,.35), transparent 28%),
            radial-gradient(circle at bottom right, rgba(139,92,246,.30), transparent 30%),
            linear-gradient(135deg,#e0f2fe 0%,#dbeafe 35%,#ede9fe 100%) !important;
        overflow:hidden;
}

    section[data-testid="stMain"],
    [data-testid="stMainBlockContainer"]{
        background:transparent !important;
}

    [data-testid="stMainBlockContainer"]{
        padding-top:6vh !important;
}

    /* floating blur orbs */
    [data-testid="stAppViewContainer"]::before{
        content:'';
        position:fixed;
        width:420px;
        height:420px;
        top:-120px;
        left:-100px;
        border-radius:50%;
        background:radial-gradient(circle,#67e8f9 0%, rgba(103,232,249,0) 70%);
        filter:blur(40px);
        animation:float1 12s ease-in-out infinite;
        z-index:0;
}

    [data-testid="stAppViewContainer"]::after{
        content:'';
        position:fixed;
        width:520px;
        height:520px;
        bottom:-180px;
        right:-120px;
        border-radius:50%;
        background:radial-gradient(circle,#a78bfa 0%, rgba(167,139,250,0) 70%);
        filter:blur(60px);
        animation:float2 14s ease-in-out infinite;
        z-index:0;
}

    @keyframes float1{
        0%,100%{transform:translateY(0px) translateX(0px);}
        50%{transform:translateY(25px) translateX(20px);}
}

    @keyframes float2{
        0%,100%{transform:translateY(0px);}
        50%{transform:translateY(-20px);}
}

    div[data-testid="column"] > div[data-testid="stVerticalBlock"]{
        position:relative;
        z-index:2;
        background:rgba(255,255,255,0.18) !important;
        backdrop-filter:blur(24px) saturate(180%) !important;
        -webkit-backdrop-filter:blur(24px) saturate(180%) !important;
        border:1px solid rgba(255,255,255,0.35) !important;
        border-radius:34px !important;
        padding:3rem 2.6rem 2.2rem !important;
        box-shadow:
            0 20px 60px rgba(99,102,241,.18),
            inset 0 1px 0 rgba(255,255,255,.45) !important;
}

    div[data-testid="column"] label{
        color:#475569 !important;
        font-weight:600 !important;
}

    div[data-testid="column"] .stTextInput > div > div > input{
        background:rgba(255,255,255,0.40) !important;
        border:1px solid rgba(255,255,255,0.50) !important;
        border-radius:18px !important;
        color:#0f172a !important;
        padding:14px 18px !important;
        font-size:.95rem !important;
        transition:all .2s ease !important;
}

    div[data-testid="column"] .stTextInput > div > div > input:focus{
        border-color:#60a5fa !important;
        box-shadow:0 0 0 4px rgba(96,165,250,.20) !important;
        background:rgba(255,255,255,0.58) !important;
}

    div[data-testid="column"] [data-testid="stFormSubmitButton"] > button{
        background:linear-gradient(135deg,#38bdf8 0%, #6366f1 55%, #8b5cf6 100%) !important;
        color:white !important;
        border:none !important;
        border-radius:18px !important;
        font-weight:700 !important;
        font-size:1rem !important;
        padding:14px !important;
        box-shadow:0 12px 30px rgba(99,102,241,.28) !important;
        transition:all .25s ease !important;
}

    div[data-testid="column"] [data-testid="stFormSubmitButton"] > button:hover{
        transform:translateY(-2px) scale(1.01) !important;
        box-shadow:0 18px 40px rgba(99,102,241,.35) !important;
}

    div[data-testid="column"] [data-testid="stAlert"]{
        border-radius:14px !important;
}
</style>
    """, unsafe_allow_html=True)

    # V5 login polish — same indigo / purple / sky palette as the application.
    st.markdown("""
    <style>
    [data-testid="stMainBlockContainer"]{max-width:1160px!important;padding:clamp(2rem,7vh,5rem) 1rem 3rem!important}
    [data-testid="stAppViewContainer"]{background:radial-gradient(circle at 8% 8%,rgba(56,189,248,.26),transparent 28rem),radial-gradient(circle at 92% 88%,rgba(139,92,246,.23),transparent 31rem),linear-gradient(145deg,#EFF8FF,#EEF2FF 48%,#F5F3FF)!important}
    div[data-testid="column"]>div[data-testid="stVerticalBlock"]{background:rgba(255,255,255,.82)!important;border:1px solid rgba(255,255,255,.9)!important;border-radius:28px!important;padding:clamp(1.6rem,4vw,2.8rem)!important;box-shadow:0 28px 80px rgba(79,70,229,.16),0 2px 10px rgba(15,23,42,.05)!important;backdrop-filter:blur(22px) saturate(150%)!important}
    div[data-testid="column"] .stTextInput input{min-height:52px!important;background:#F8FAFC!important;border:1px solid #DDE4F0!important;border-radius:14px!important;color:#0F172A!important;padding:0 16px!important}
    div[data-testid="column"] .stTextInput input:focus{background:#FFF!important;border-color:#6366F1!important;box-shadow:0 0 0 4px rgba(99,102,241,.12)!important}
    div[data-testid="column"] [data-testid="stFormSubmitButton"] button{min-height:52px!important;border-radius:14px!important;background:linear-gradient(135deg,#38BDF8,#6366F1 54%,#8B5CF6)!important;box-shadow:0 12px 28px rgba(99,102,241,.27)!important}
    @media(max-width:720px){[data-testid="stMainBlockContainer"]{padding:1rem .75rem!important}div[data-testid="column"]>div[data-testid="stVerticalBlock"]{padding:1.35rem!important;border-radius:22px!important}}
    </style>
    """, unsafe_allow_html=True)

    _, center_col, _ = st.columns([1, 1.4, 1])
    with center_col:
        st.markdown("""
        <div style="text-align:center; margin-bottom:28px;">
            <div style="width:72px;height:72px;background:linear-gradient(135deg,#6366f1,#8b5cf6);
                        border-radius:20px;display:inline-flex;align-items:center;justify-content:center;
                        font-size:2rem;box-shadow:0 8px 28px rgba(99,102,241,.55);margin-bottom:16px;">🛡️</div>
            <h1 style="color:#0f172a;font-size:2rem;font-weight:800;margin:0 0 8px;letter-spacing:-1px;">
                DocumentReportUnified</h1>
            <p style="color:#475569;font-size:.92rem;margin:0;font-weight:500;">
                Enterprise IT Management Platform</p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            u = st.text_input("📧 Microsoft 365 Email", placeholder="yourname@optimal.co.th")
            p = st.text_input("🔒 Password", type="password", placeholder="••••••••••••")
            submitted = st.form_submit_button("Sign In →", use_container_width=True)
            if submitted:
                if not u.strip() or not p.strip():
                    st.warning("⚠️ กรุณากรอก Email และ Password")
                else:
                    with st.spinner("กำลังตรวจสอบ..."):
                        success, name, email = check_ms_login(u, p)
                    if success:
                        st.session_state.is_auth  = True
                        st.session_state.skip_cookie_login = False
                        st.session_state.user_name  = name
                        st.session_state.user_email = email
                        cookie_manager.set("user_name",  name,  key="auth_token")
                        cookie_manager.set("user_email", email, key="auth_email")
                        st.rerun()
                    else:
                        st.error(f"❌ {name}")

        st.markdown("""
        <p style="color:rgba(255,255,255,.25);font-size:.75rem;text-align:center;margin-top:20px;">
            ☁️ Secure Microsoft 365 Access &nbsp;·&nbsp; Optimal Group © 2025</p>
        """, unsafe_allow_html=True)

# =============================================================================
# THEME : SIDEBAR / # THEME BLOCK : SIDEBAR / DASHBOARD
# แก้ UI หลัง Login ทั้งหมดในช่วงด้านล่างนี้

# MAIN APP
# CSS หลัง Login สำเร็จ
# =============================================================================

# ══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════════════
else:
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Thai:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');

    html, body, [class*="css"] {
        font-family: 'IBM Plex Sans Thai', sans-serif;
}

    /* ── APP BG ── */
    [data-testid="stAppViewContainer"] { background: #F8FAFC !important; }
    section[data-testid="stMain"]      { background: #F8FAFC !important; }
    [data-testid="stMainBlockContainer"] {
        padding: 1.8rem 2rem 3rem !important;
        max-width: 1400px !important;
}
    [data-testid="stHeader"] { background: transparent !important; box-shadow: none !important; }

    /* ── SIDEBAR — Reference mockup (M365 / Entra / Azure) ── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    [data-testid="stSidebar"]{
    background:
        radial-gradient(circle at top left,
            rgba(56,189,248,.25),
            transparent 30%),
        radial-gradient(circle at bottom right,
            rgba(139,92,246,.20),
            transparent 35%),
        linear-gradient(
            135deg,
            #e0f2fe 0%,
            #dbeafe 35%,
            #ede9fe 100%
        ) !important;
    }
    [data-testid="stSidebar"] {
    background:
        linear-gradient(
            180deg,
            #e0f2fe 0%,
            #dbeafe 40%,
            #ede9fe 100%
        ) !important;
    }
    [data-testid="stSidebarContent"] {
        background: #FFFFFF !important;
        margin: 0 !important;
        padding: 20px 16px 24px 16px !important;
        border: none !important;
        border-radius: 0 !important;
        box-shadow: none !important;
        font-family: 'Inter', 'Segoe UI', 'IBM Plex Sans Thai', sans-serif !important;
    }
    [data-testid="stSidebar"] * {
        font-family: 'Inter', 'Segoe UI', 'IBM Plex Sans Thai', sans-serif !important;
    }
    [data-testid="stSidebar"] label {
        color: #64748B !important;
        font-size: 0.65rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.1em !important;
        text-transform: uppercase !important;
    }
    [data-testid="stSidebar"] hr {
        border: none !important;
        height: 1px !important;
        background: #E2E8F0 !important;
        margin: 12px 0 !important;
    }
    [data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
        background: #FFFFFF !important;
        border: 1px solid #E2E8F0 !important;
        border-radius: 12px !important;
        color: #0F172A !important;
    }
    /* Brand header */
    [data-testid="stSidebar"] .ref-brand {
        padding: 0 2px 18px 2px;
        margin-bottom: 0;
        border-bottom: 1px solid #E2E8F0;
    }
    [data-testid="stSidebar"] .ref-logo {
        width: 36px; height: 36px; border-radius: 12px;
        background: #2563EB; color: #FFFFFF;
        display: flex; align-items: center; justify-content: center;
        font-size: 0.7rem; font-weight: 700; flex-shrink: 0;
    }
    [data-testid="stSidebar"] .ref-brand-title {
        font-size: 0.875rem; font-weight: 700; color: #0F172A;
        line-height: 1.3; letter-spacing: -0.02em;
    }
    [data-testid="stSidebar"] .ref-brand-sub {
        font-size: 0.7rem; color: #64748B; margin-top: 1px; font-weight: 500;
    }
    /* Profile card — reference layout */
    [data-testid="stSidebar"] .ref-profile {
        padding: 0 2px 14px 2px;
        border-bottom: 1px solid #E2E8F0;
        margin-bottom: 0;
    }
    [data-testid="stSidebar"] .ref-avatar {
        width: 44px; height: 44px; border-radius: 12px;
        background: #2563EB; color: #FFFFFF;
        display: flex; align-items: center; justify-content: center;
        font-size: 0.8rem; font-weight: 700; flex-shrink: 0;
    }
    [data-testid="stSidebar"] .ref-profile-dept {
        font-size: 0.875rem; font-weight: 700; color: #0F172A;
        line-height: 1.3;
        white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    }
    [data-testid="stSidebar"] .ref-profile-name {
        font-size: 0.8125rem; font-weight: 400; color: #64748B;
        margin-top: 2px;
        white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    }
    [data-testid="stSidebar"] .ref-status-row {
        display: flex; align-items: center; gap: 8px;
        margin-top: 14px; padding-top: 14px;
        border-top: 1px solid #E2E8F0;
        font-size: 0.8125rem; font-weight: 500; color: #64748B;
    }
    [data-testid="stSidebar"] .ref-status-dot {
        width: 8px; height: 8px; border-radius: 50%;
        background: #22C55E; flex-shrink: 0;
    }
    [data-testid="stSidebar"] .ref-toolbar {
        display: flex; justify-content: flex-end; gap: 4px;
        margin: -4px 0 8px 0;
    }
    [data-testid="stSidebar"] .ref-toolbar button {
        min-height: 28px !important;
        padding: 2px 8px !important;
        font-size: 0.75rem !important;
        color: #94A3B8 !important;
        border: 1px solid #E2E8F0 !important;
        background: #FFFFFF !important;
        border-radius: 8px !important;
    }
    [data-testid="stSidebar"] .nav-section-label {
        font-size: 0.625rem; font-weight: 600; letter-spacing: 0.12em;
        text-transform: uppercase; color: #94A3B8;
        padding: 10px 4px 6px 4px; margin: 0;
    }
    [data-testid="stSidebar"] .nav-badge-pill {
        display: inline-flex; align-items: center; justify-content: center;
        min-width: 32px; height: 22px; padding: 0 8px;
        border-radius: 999px; font-size: 0.72rem; font-weight: 600;
        line-height: 1; flex-shrink: 0;
    }
    [data-testid="stSidebar"] .nav-badge-blue {
        background: #EFF6FF; color: #2563EB; border: none;
    }
    [data-testid="stSidebar"] .nav-badge-green {
        background: #F0FDF4; color: #16A34A; border: none;
    }
    [data-testid="stSidebar"] .nav-badge-red {
        background: #FEF2F2; color: #DC2626; border: none;
    }
    /* ── Tight vertical rhythm (fix Streamlit default gaps) ── */
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
        gap: 0.2rem !important;
    }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div {
        gap: 0.2rem !important;
    }
    [data-testid="stSidebar"] .element-container {
        margin-bottom: 0 !important;
        padding-bottom: 0 !important;
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] {
        margin-bottom: 0 !important;
        padding: 0 !important;
    }
    [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] {
        gap: 6px !important;
        margin-bottom: 0 !important;
        padding-bottom: 0 !important;
        align-items: center !important;
        min-height: 0 !important;
        flex-wrap: nowrap !important;
    }
    [data-testid="stSidebar"] [data-testid="column"] {
        gap: 0 !important;
        padding-top: 0 !important;
        padding-bottom: 0 !important;
        min-height: 0 !important;
    }
    [data-testid="stSidebar"] div[data-testid="column"] .stButton {
        margin-bottom: 0 !important;
    }
    [data-testid="stSidebar"] .nav-row-h {
        margin-bottom: 0 !important;
    }
    [data-testid="stSidebar"] .stButton {
        margin: 0 !important;
        padding: 0 !important;
    }
    [data-testid="stSidebar"] .stButton > button {
        width: 100% !important;
        min-height: 34px !important;
        height: 34px !important;
        padding: 0 10px !important;
        margin: 0 !important;
        border-radius: 10px !important;
        border: 1px solid transparent !important;
        background: transparent !important;
        color: #0F172A !important;
        font-size: 0.8125rem !important;
        font-weight: 500 !important;
        text-align: left !important;
        justify-content: flex-start !important;
        box-shadow: none !important;
        cursor: pointer !important;
        transition: background 0.18s ease, color 0.18s ease, box-shadow 0.18s ease !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: #F8FAFC !important;
        color: #0F172A !important;
        border-color: transparent !important;
    }
    [data-testid="stSidebar"] .stButton > button[kind="primary"] {
        background: transparent !important;
        color: #2563EB !important;
        border: none !important;
        border-left: 4px solid #2563EB !important;
        padding-left: 9px !important;
        font-weight: 600 !important;
        box-shadow: 0 1px 3px rgba(37, 99, 235, 0.12) !important;
    }
    [data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
        background: #DBEAFE !important;
        color: #1D4ED8 !important;
        box-shadow: 0 2px 6px rgba(37, 99, 235, 0.16) !important;
    }
    [data-testid="stSidebar"] .stButton > .stButton > button {
        background: transparent !important;
        color: #0F172A !important;
    }
    [data-testid="stSidebar"] .nav-group-btn > button,
    [data-testid="stSidebar"] button.nav-group-toggle-el {
        font-weight: 600 !important;
        color: #0F172A !important;
        min-height: 36px !important;
        height: 36px !important;
    }
    [data-testid="stSidebar"] .nav-toolbar-row .stButton > button {
        min-height: 30px !important;
        height: 30px !important;
        padding: 0 8px !important;
        font-size: 0.75rem !important;
        border: 1px solid #E2E8F0 !important;
        background: #FFFFFF !important;
        color: #64748B !important;
        border-radius: 8px !important;
    }
    [data-testid="stSidebar"] .nav-toolbar-row {
        margin-bottom: 4px !important;
    }
    [data-testid="stSidebar"] .nav-toolbar-row [data-testid="stHorizontalBlock"] {
        margin-bottom: 0 !important;
    }
    [data-testid="stSidebar"] .nav-group-btn > button:hover {
        background: #F8FAFC !important;
    }
    [data-testid="stSidebar"] .nav-signout .stButton > button {
        color: #EF4444 !important;
        font-weight: 500 !important;
        border: none !important;
        background: transparent !important;
        min-height: 44px !important;
        margin-top: 4px !important;
    }
    [data-testid="stSidebar"] .nav-signout .stButton > button:hover {
        background: #FEF2F2 !important;
        color: #DC2626 !important;
    }
    [data-testid="stSidebar"].sidebar-compact {
        min-width: 72px !important;
    }
    [data-testid="stSidebar"] .hide-when-compact { display: block; }
    [data-testid="stSidebar"].sidebar-compact .hide-when-compact { display: none !important; }
    @media (max-width: 768px) {
        [data-testid="stSidebar"] { min-width: 72px !important; }
        [data-testid="stSidebar"] .hide-when-compact { display: none !important; }
    }

    /* ── METRIC CARDS ── */
    [data-testid="stMetric"] {
        background: #ffffff !important;
        border-radius: 14px !important;
        padding: 1.1rem 1.4rem 1rem !important;
        border: 1px solid #e2e8f0 !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.07), 0 4px 16px rgba(0,0,0,0.04) !important;
        position: relative; overflow: hidden;
}
    [data-testid="stMetric"]::before {
        content: ''; position: absolute; top: 0; left: 0;
        width: 4px; height: 100%;
        background: linear-gradient(180deg, #6366f1, #8b5cf6);
}
    [data-testid="stMetricLabel"] { font-size: 0.76rem !important; color: #94a3b8 !important; font-weight: 600 !important; letter-spacing: 0.5px !important; text-transform: uppercase !important; }
    [data-testid="stMetricValue"] { font-size: 2.1rem !important; font-weight: 700 !important; color: #0f172a !important; letter-spacing: -1px !important; font-family: 'IBM Plex Mono', monospace !important; }

    /* ── CONTENT CARDS ── */
    div[data-testid="stVerticalBlockBorderWrapper"]:has(div[data-testid="stElementContainer"]) > div {
        background: #ffffff !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 14px !important;
        box-shadow: 0 1px 3px rgba(0,0,0,0.06), 0 4px 16px rgba(0,0,0,0.03) !important;
        transition: box-shadow .2s, transform .2s !important;
}
    div[data-testid="stVerticalBlockBorderWrapper"]:has(div[data-testid="stElementContainer"]) > div:hover {
        box-shadow: 0 4px 20px rgba(99,102,241,0.12), 0 1px 3px rgba(0,0,0,0.06) !important;
        transform: translateY(-1px) !important;
}

    /* ── TYPOGRAPHY ── */
    h1 { color: #0f172a !important; font-weight: 700 !important; letter-spacing: -0.5px !important; font-size: 1.6rem !important; }
    h2 { color: #1e293b !important; font-weight: 600 !important; }
    h3 { color: #334155 !important; font-weight: 600 !important; font-size: 1rem !important; }
    p, .stMarkdown p { color: #475569 !important; }

    /* ── PRIMARY BUTTON — main content only ── */
    [data-testid="stMain"] [data-testid="stButton"] > button[kind="primary"],
    [data-testid="stFormSubmitButton"] > button {
        background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
        border: none !important; border-radius: 10px !important;
        color: #fff !important; font-weight: 600 !important;
        box-shadow: 0 2px 8px rgba(99,102,241,0.35) !important;
        transition: all .15s !important;
        font-family: 'IBM Plex Sans Thai', sans-serif !important;
}
    [data-testid="stMain"] [data-testid="stButton"] > button[kind="primary"]:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 16px rgba(99,102,241,0.45) !important;
}

    /* ── SECONDARY BUTTON — main content only (NOT sidebar) ── */
    [data-testid="stMain"] [data-testid="stButton"] > button:not([kind="primary"]) {
        background: #fff !important; border: 1px solid #e2e8f0 !important;
        border-radius: 10px !important; color: #475569 !important;
        font-family: 'IBM Plex Sans Thai', sans-serif !important;
        transition: all .15s !important;
}
    [data-testid="stMain"] [data-testid="stButton"] > button:not([kind="primary"]):hover {
        border-color: #6366f1 !important; color: #6366f1 !important;
        background: #f5f3ff !important;
}

    /* ── INPUTS ── */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input {
        border-radius: 10px !important; border: 1.5px solid #e2e8f0 !important;
        font-family: 'IBM Plex Sans Thai', sans-serif !important;
        background: #fff !important; color: #1e293b !important;
        padding: 10px 14px !important; font-size: 0.88rem !important;
        transition: border .15s, box-shadow .15s !important;
}
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {
        border-color: #6366f1 !important;
        box-shadow: 0 0 0 3px rgba(99,102,241,0.12) !important;
}

    /* ── SELECTBOX ── */
    [data-baseweb="select"] > div {
        border-radius: 10px !important; border-color: #e2e8f0 !important;
        font-family: 'IBM Plex Sans Thai', sans-serif !important;
        background: #fff !important;
}
    [data-baseweb="select"] > div:hover { border-color: #a5b4fc !important; }

    /* ── DATAFRAME ── */
    [data-testid="stDataFrame"] {
        border-radius: 12px !important; overflow: hidden !important;
        border: 1px solid #e2e8f0 !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.05) !important;
}

    /* ── DIVIDER ── */
    hr { border-color: #e2e8f0 !important; }

    /* ── EXPANDER ── */
    details > summary {
        border-radius: 10px !important; background: #f8fafc !important;
        border: 1px solid #e2e8f0 !important; padding: 10px 14px !important;
        font-size: 0.85rem !important; font-weight: 500 !important; color: #475569 !important;
}
    details[open] > summary { border-bottom-left-radius: 0 !important; border-bottom-right-radius: 0 !important; }

    /* ── ALERTS ── */
    [data-testid="stAlert"][data-baseweb="notification"] {
        border-radius: 12px !important; border-left: 4px solid !important;
        font-size: 0.88rem !important;
}

    /* ── SPINNER ── */
    [data-testid="stSpinner"] { opacity: 0.7; }

    /* ── GAP — main content only ── */
    [data-testid="stMain"] [data-testid="stVerticalBlock"] { gap: 0.6rem !important; }

    

    /* REMOVE EMPTY STREAMLIT CONTAINERS */
    div[data-testid="stHorizontalBlock"]:empty,
    div[data-testid="stVerticalBlock"]:empty,
    div[data-testid="stVerticalBlockBorderWrapper"]:empty{
        display:none !important;
        margin:0 !important;
        padding:0 !important;
        height:0 !important;
        min-height:0 !important;
}

    /* FIX EMPTY COLUMN GAPS */
    div[data-testid="column"]:empty{
        display:none !important;
}

    /* REMOVE AUTO EMPTY PLACEHOLDERS */
    .element-container:empty{
        display:none !important;
}


    /* ── CHECKBOX ── */
    [data-testid="stCheckbox"] label { color: #475569 !important; font-size: 0.85rem !important; }

    /* ── CAPTION ── */
    [data-testid="stCaptionContainer"] p { color: #94a3b8 !important; font-size: 0.78rem !important; }

    /* ── SIDEBAR STAT CARDS ── */
    .ov-card {
        background: #fff;
        border-radius: 14px;
        padding: 16px 18px 12px;
        border: 1px solid #e8eaf0;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
        transition: box-shadow .2s, transform .2s;
        height: 100%;
}
    .ov-card:hover { box-shadow: 0 4px 18px rgba(99,102,241,0.13); transform: translateY(-2px); }
    .ov-card-top { display: flex; align-items: center; gap: 10px; margin-bottom: 10px; }
    .ov-card-icon { width: 38px; height: 38px; border-radius: 10px; display: flex; align-items: center; justify-content: center; font-size: 1.15rem; flex-shrink: 0; }
    .ov-card-label { font-size: 0.72rem; font-weight: 700; letter-spacing: 0.6px; text-transform: uppercase; color: #94a3b8; }
    .ov-card-value { font-size: 2.1rem; font-weight: 800; line-height: 1.1; letter-spacing: -1px; margin-bottom: 10px; }
    .ov-card-link { display: flex; align-items: center; gap: 4px; font-size: 0.78rem; font-weight: 500; padding-top: 8px; border-top: 1px solid #f1f5f9; cursor: pointer; }
    

    /* ── STATUS BADGE (custom html) ── */
    .badge {
        display: inline-block; padding: 3px 11px; border-radius: 20px;
        font-size: 0.73rem; font-weight: 700; letter-spacing: 0.4px;
}
    .badge-active   { background: #dcfce7; color: #15803d; }
    .badge-inactive { background: #fee2e2; color: #b91c1c; }
    .badge-spare    { background: #dbeafe; color: #1d4ed8; }
    .badge-repair   { background: #fef3c7; color: #b45309; }
    .badge-default  { background: #f1f5f9; color: #64748b; }

    /* ── PAGE HEADER (custom html) ── */
    .page-header {
        background: linear-gradient(135deg, #0f172a 0%, #1e1b4b 60%, #312e81 100%);
        border-radius: 16px; padding: 24px 28px; margin-bottom: 1.5rem;
        display: flex; align-items: center; gap: 16px;
        box-shadow: 0 4px 20px rgba(99,102,241,0.20);
        position: relative; overflow: hidden;
}
    .page-header::before {
        content: ''; position: absolute; right: -40px; top: -40px;
        width: 200px; height: 200px; border-radius: 50%;
        background: rgba(99,102,241,0.18); filter: blur(40px);
}
    .page-header-icon {
        font-size: 2.2rem; line-height: 1;
        background: rgba(255,255,255,0.10); border-radius: 14px;
        padding: 12px; border: 1px solid rgba(255,255,255,0.14);
}
    .page-header-text h1 {
        color: #fff !important; font-size: 1.4rem !important;
        font-weight: 700 !important; margin: 0 0 4px !important;
        letter-spacing: -0.3px !important;
}
    .page-header-text p { color: rgba(255,255,255,0.5) !important; font-size: 0.82rem !important; margin: 0 !important; }

    /* ── SECTION HEADER ── */
    .section-title {
        font-size: 0.72rem; font-weight: 700; letter-spacing: 1px;
        text-transform: uppercase; color: #94a3b8;
        margin: 1.2rem 0 0.5rem; padding-bottom: 6px;
        border-bottom: 1px solid #e2e8f0;
}

    /* ── HARDWARE CARD ── */
    .hw-card-title { font-size: 0.98rem; font-weight: 600; color: #1e293b; margin: 0; }
    .hw-card-sub   { font-size: 0.76rem; color: #94a3b8; margin: 3px 0 8px; }
    .hw-field      { font-size: 0.82rem; color: #475569; padding: 5px 0; border-bottom: 1px solid #f1f5f9; }
    .hw-field:last-of-type { border-bottom: none; }
    .hw-field strong { color: #334155; font-weight: 600; }

    /* ── NAS CARD ── */
    .nas-folder-badge {
        background: linear-gradient(135deg, #ede9fe, #ddd6fe);
        border: 1px solid #c4b5fd; border-radius: 10px;
        padding: 10px 14px; display: inline-block;
}
    .nas-folder-name { font-size: 1.05rem; font-weight: 700; color: #4c1d95; }
    .nas-user-tag {
        display: inline-block; margin: 2px 3px; padding: 3px 10px;
        border-radius: 16px; font-size: 0.75rem; font-weight: 600;
}
    .nas-rw  { background: #fee2e2; color: #991b1b; border: 1px solid #fca5a5; }
    .nas-ro  { background: #dcfce7; color: #166534; border: 1px solid #86efac; }

    /* ── PASSWORD CARD ── */
    .pw-field-row {
        display: flex; gap: 8px; align-items: baseline;
        padding: 6px 0; border-bottom: 1px solid #f8fafc; font-size: 0.83rem;
}
    .pw-field-row:last-of-type { border-bottom: none; }
    .pw-field-label { color: #94a3b8; font-size: 0.72rem; font-weight: 600; min-width: 90px; flex-shrink: 0; }
    .pw-field-value { color: #334155; font-weight: 500; word-break: break-all; }

    </style>
    """, unsafe_allow_html=True)

    # ── helpers ──────────────────────────────────────────────
    # V5 UNIFIED UI — final visual layer for all authenticated screens.
    st.markdown("""
    <style>
    :root{--brand:#6366F1;--brand-dark:#4F46E5;--violet:#8B5CF6;--sky:#38BDF8;--success:#10B981;--warning:#F59E0B;--danger:#EF4444;--ink:#0F172A;--muted:#64748B;--line:#E2E8F0;--canvas:#F6F8FC;--card:#FFF;--r:18px;--rl:24px;--shadow:0 10px 30px rgba(15,23,42,.055);--shadow-up:0 18px 48px rgba(79,70,229,.12)}
    html,body,[class*="css"]{font-family:'Inter','IBM Plex Sans Thai','Segoe UI',sans-serif!important;color:var(--ink)}
    [data-testid="stAppViewContainer"],section[data-testid="stMain"]{background:radial-gradient(circle at 90% 0,rgba(99,102,241,.075),transparent 24rem),var(--canvas)!important}
    [data-testid="stMainBlockContainer"]{max-width:1480px!important;padding:1.65rem clamp(1rem,3vw,2.75rem) 4rem!important}
    [data-testid="stHeader"]{height:2.2rem!important;background:rgba(246,248,252,.78)!important;backdrop-filter:blur(14px)}
    [data-testid="stSidebar"]{background:linear-gradient(180deg,#EEF8FF 0%,#EEF2FF 52%,#F3EFFF 100%)!important}
    [data-testid="stSidebar"] [data-testid="stSidebarContent"]{background:rgba(255,255,255,.76)!important;border-right:1px solid rgba(148,163,184,.22)!important;padding:22px 14px 18px!important;box-shadow:12px 0 40px rgba(79,70,229,.06)!important;backdrop-filter:blur(20px) saturate(145%)!important}
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"]{gap:.35rem!important}
    [data-testid="stSidebar"] .ref-brand{padding:2px 4px 18px!important;margin:0 0 12px!important;border-bottom:1px solid rgba(148,163,184,.22)!important}
    [data-testid="stSidebar"] .ref-logo{width:46px!important;height:46px!important;border-radius:15px!important;background:linear-gradient(135deg,var(--sky),var(--brand) 56%,var(--violet))!important;box-shadow:0 10px 24px rgba(99,102,241,.24)!important}
    [data-testid="stSidebar"] .ref-profile{padding:14px!important;margin:0 2px 15px!important;border:1px solid rgba(148,163,184,.22)!important;border-radius:18px!important;background:rgba(255,255,255,.72)!important;box-shadow:0 8px 24px rgba(15,23,42,.05)!important}
    [data-testid="stSidebar"] .ref-avatar{width:48px!important;height:48px!important;border-radius:15px!important;background:linear-gradient(135deg,var(--brand),var(--violet))!important}
    [data-testid="stSidebar"] .nav-section-label{padding:10px 10px 7px!important;color:#94A3B8!important;font-size:.65rem!important;font-weight:800!important;letter-spacing:.14em!important}
    [data-testid="stSidebar"] .stButton>button{height:46px!important;min-height:46px!important;border:1px solid transparent!important;border-radius:13px!important;padding:0 13px!important;background:transparent!important;color:#334155!important;font-size:.87rem!important;font-weight:650!important;box-shadow:none!important;justify-content:flex-start!important;transition:.18s ease!important}
    [data-testid="stSidebar"] .stButton>button:hover{background:rgba(255,255,255,.72)!important;color:var(--brand-dark)!important;border-color:rgba(99,102,241,.14)!important;transform:translateX(2px)!important}
    [data-testid="stSidebar"] .stButton>button[kind="primary"]{background:linear-gradient(135deg,rgba(219,234,254,.96),rgba(237,233,254,.94))!important;color:#4338CA!important;border:1px solid rgba(99,102,241,.18)!important;border-left:3px solid var(--brand)!important;box-shadow:0 8px 20px rgba(79,70,229,.10)!important}
    [data-testid="stSidebar"] .nav-signout .stButton>button{margin-top:1.25rem!important;height:44px!important;min-height:44px!important;justify-content:center!important;border:1px solid rgba(239,68,68,.14)!important;background:rgba(255,255,255,.58)!important;color:#64748B!important}
    .page-header,.asset-hero{min-height:132px!important;padding:26px 30px!important;border-radius:var(--rl)!important;background:linear-gradient(125deg,#172554 0%,#312E81 50%,#6D28D9 100%)!important;box-shadow:0 18px 42px rgba(49,46,129,.20)!important;margin-bottom:1.35rem!important;position:relative;overflow:hidden}
    .page-header:after,.asset-hero:after{content:'';position:absolute;width:260px;height:260px;right:-65px;bottom:-165px;border-radius:50%;background:rgba(56,189,248,.16)}
    .page-header-icon{width:52px;height:52px;display:flex;align-items:center;justify-content:center;padding:0!important;border-radius:16px!important;background:rgba(255,255,255,.12)!important}
    .page-header-text h1,.asset-title{font-size:clamp(1.45rem,2vw,1.9rem)!important;font-weight:800!important;letter-spacing:-.035em!important}
    .page-header-text p,.asset-sub{color:rgba(255,255,255,.72)!important;font-size:.9rem!important}
    [data-testid="stMetric"],.ov-card,.asset-stat,.asset-card,.nas-folder-badge{background:rgba(255,255,255,.96)!important;border:1px solid var(--line)!important;border-radius:var(--r)!important;box-shadow:var(--shadow)!important;transition:transform .18s ease,box-shadow .18s ease,border-color .18s ease!important}
    [data-testid="stMetric"]{padding:1.2rem 1.25rem!important;min-height:116px}
    [data-testid="stMetric"]:hover,.ov-card:hover,.asset-card:hover{transform:translateY(-3px)!important;box-shadow:var(--shadow-up)!important;border-color:#C7D2FE!important}
    [data-testid="stMetricLabel"]{color:var(--muted)!important;font-size:.72rem!important;font-weight:800!important;letter-spacing:.08em!important;text-transform:uppercase}
    [data-testid="stMetricValue"]{color:var(--ink)!important;font-size:1.85rem!important;font-weight:800!important;letter-spacing:-.04em}
    [data-testid="stMain"] div[data-testid="stVerticalBlockBorderWrapper"]>div{background:var(--card)!important;border:1px solid var(--line)!important;border-radius:var(--r)!important;box-shadow:var(--shadow)!important}
    .stTextInput input,.stNumberInput input,.stTextArea textarea,.stDateInput input,.stSelectbox div[data-baseweb="select"]>div,.stMultiSelect div[data-baseweb="select"]>div{min-height:46px!important;background:#FFF!important;border:1px solid #D9E1EC!important;border-radius:12px!important;color:var(--ink)!important;box-shadow:0 1px 2px rgba(15,23,42,.02)!important}
    .stTextInput input:focus,.stNumberInput input:focus,.stTextArea textarea:focus{border-color:var(--brand)!important;box-shadow:0 0 0 4px rgba(99,102,241,.10)!important}
    label[data-testid="stWidgetLabel"] p{color:#475569!important;font-size:.8rem!important;font-weight:700!important}
    .stButton>button,.stDownloadButton>button,[data-testid="stFormSubmitButton"]>button{min-height:44px!important;border-radius:12px!important;border:1px solid #D9E1EC!important;background:#FFF!important;color:#334155!important;font-weight:700!important;box-shadow:0 2px 6px rgba(15,23,42,.035)!important;transition:.18s ease!important}
    .stButton>button:hover,.stDownloadButton>button:hover{border-color:#A5B4FC!important;color:var(--brand-dark)!important;transform:translateY(-1px)!important;box-shadow:0 8px 18px rgba(79,70,229,.10)!important}
    .stButton>button[kind="primary"],[data-testid="stFormSubmitButton"]>button{background:linear-gradient(135deg,var(--brand-dark),var(--brand) 55%,var(--violet))!important;border-color:transparent!important;color:#FFF!important;box-shadow:0 8px 18px rgba(99,102,241,.22)!important}
    .stTabs [data-baseweb="tab-list"]{gap:6px;background:#EEF2FF;border-radius:14px;padding:5px;width:max-content;max-width:100%;overflow-x:auto}
    .stTabs [data-baseweb="tab"]{height:40px;border-radius:10px;padding:0 16px;color:#64748B;font-weight:700}
    .stTabs [aria-selected="true"]{background:#FFF!important;color:var(--brand-dark)!important;box-shadow:0 3px 10px rgba(15,23,42,.08)!important}
    .stTabs [data-baseweb="tab-highlight"],.stTabs [data-baseweb="tab-border"]{display:none!important}
    [data-testid="stAlert"]{border-radius:14px!important;border:1px solid rgba(99,102,241,.14)!important;box-shadow:none!important}
    [data-testid="stDataFrame"],div[data-testid="stTable"]{border:1px solid var(--line)!important;border-radius:16px!important;overflow:hidden!important;background:#FFF!important;box-shadow:var(--shadow)!important}
    [data-testid="stPlotlyChart"]{background:#FFF;border:1px solid var(--line);border-radius:18px;padding:10px;box-shadow:var(--shadow);overflow:hidden}
    [data-testid="stDialog"]>div{border-radius:22px!important;border:1px solid var(--line)!important;box-shadow:0 28px 80px rgba(15,23,42,.20)!important}
    @media(max-width:900px){[data-testid="stMainBlockContainer"]{padding:1rem .85rem 3rem!important}.page-header,.asset-hero{min-height:110px!important;padding:21px!important;border-radius:20px!important}.page-header-icon{width:44px;height:44px}.asset-stat-wrap{grid-template-columns:repeat(2,1fr)!important}[data-testid="stHorizontalBlock"]{gap:.65rem!important}}
    @media(max-width:640px){.page-header{align-items:flex-start!important}.asset-stat-wrap{grid-template-columns:1fr!important}[data-testid="stMetric"]{min-height:96px;padding:1rem!important}.stTabs [data-baseweb="tab"]{padding:0 11px;font-size:.78rem}}
    </style>
    """, unsafe_allow_html=True)

    def status_badge(status: str) -> str:
        cls = {"Active": "badge-active", "Inactive": "badge-inactive",
               "Spare": "badge-spare", "Repair": "badge-repair"}.get(status, "badge-default")
        return f'<span class="badge {cls}">{status or "—"}</span>'

    def page_header(icon: str, title: str, subtitle: str = ""):
        st.markdown(f"""
        <div class="page-header">
            <div class="page-header-icon">{icon}</div>
            <div class="page-header-text">
                <h1>{title}</h1>
                {'<p>' + subtitle + '</p>' if subtitle else ''}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── SIDEBAR — Reference-aligned navigation (UI only) ──────
    admin_mode = is_admin(st.session_state.user_email)
    name = st.session_state.user_name or "User"
    email = st.session_state.user_email or ""
    initials = "".join([w[0].upper() for w in name.split()[:2]]) or "U"
    nav_badges = get_sidebar_nav_badges()
    # Clean Enterprise uses a full-size navigation panel only.
    compact = False
    st.session_state.sidebar_compact = False

    _profile_fallback = "OPG Human Resource" if admin_mode else (
        "Information Technology" if email and "optimal" in email.lower()
        else (email.split("@")[-1].replace(".co.th", "").replace(".com", "").title() if email and "@" in email
              else "Corporate Operations")
    )
    profile_dept = get_sidebar_profile_title(email or name) or _profile_fallback

    # Inject sidebar CSS inside sidebar (wins over global MODERN_THEME)
    st.sidebar.markdown("""
    <style>
    [data-testid="stSidebar"],
    [data-testid="stSidebar"] > div:first-child {
        background:
            radial-gradient(circle at 0% 0%, rgba(56, 189, 248, 0.18), transparent 30%),
            radial-gradient(circle at 100% 100%, rgba(139, 92, 246, 0.14), transparent 34%),
            linear-gradient(180deg, #f8fbff 0%, #eff6ff 46%, #f3efff 100%) !important;
    }
    [data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        background: rgba(255, 255, 255, 0.84) !important;
        border-right: 1px solid rgba(148, 163, 184, 0.26) !important;
        border-radius: 0 !important;
        margin: 0 !important;
        padding: 24px 16px 18px !important;
        box-shadow: 10px 0 34px rgba(37, 99, 235, 0.07) !important;
        backdrop-filter: blur(20px) saturate(150%) !important;
        -webkit-backdrop-filter: blur(20px) saturate(150%) !important;
    }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: 0.42rem !important; }
    [data-testid="stSidebar"] .element-container { margin-bottom: 0 !important; }
    [data-testid="stSidebar"] [data-testid="stHorizontalBlock"] {
        gap: 6px !important; margin-bottom: 0 !important; align-items: center !important;
    }
    [data-testid="stSidebar"] .ref-brand {
        background: transparent !important;
        border: none !important;
        padding: 4px 4px 18px !important;
        margin-bottom: 8px !important;
        box-shadow: none !important;
    }
    [data-testid="stSidebar"] .ref-profile {
        background: rgba(255, 255, 255, 0.78) !important;
        border: 1px solid rgba(148, 163, 184, 0.22) !important;
        border-radius: 14px !important;
        padding: 18px 14px !important;
        margin: 0 2px 22px !important;
        box-shadow: 0 12px 28px rgba(15, 23, 42, 0.07) !important;
    }
    [data-testid="stSidebar"] .ref-brand-title {
        font-size: 0.98rem !important;
        font-weight: 800 !important;
        color: #0F172A !important;
    }
    [data-testid="stSidebar"] .ref-brand-sub,
    [data-testid="stSidebar"] .ref-profile-name {
        color: #475569 !important;
    }
    [data-testid="stSidebar"] .ref-avatar,
    [data-testid="stSidebar"] .ref-logo {
        background: linear-gradient(135deg, #2563EB 0%, #6366F1 58%, #8B5CF6 100%) !important;
        box-shadow: 0 8px 20px rgba(37, 99, 235, 0.22) !important;
    }
    [data-testid="stSidebar"] .ref-logo {
        width: 50px !important;
        height: 50px !important;
        border-radius: 12px !important;
        font-size: 1rem !important;
    }
    [data-testid="stSidebar"] .ref-avatar {
        width: 58px !important;
        height: 58px !important;
        border-radius: 999px !important;
        font-size: 0.98rem !important;
    }
    [data-testid="stSidebar"] .ref-status-row {
        border-top: 1px solid rgba(148, 163, 184, 0.20) !important;
        margin-top: 12px !important;
        padding-top: 10px !important;
    }
    [data-testid="stSidebar"] .nav-section-label {
        color: #334155 !important;
        font-size: 0.70rem !important;
        letter-spacing: 0.14em !important;
        padding: 8px 10px 10px !important;
    }
    [data-testid="stSidebar"] .stButton > button {
        min-height: 64px !important;
        height: 64px !important;
        border-radius: 12px !important;
        border: 1px solid transparent !important;
        background: transparent !important;
        color: #1E293B !important;
        font-size: 0.94rem !important;
        font-weight: 600 !important;
        padding: 0 18px !important;
        text-align: left !important;
        justify-content: flex-start !important;
        box-shadow: none !important;
        transition: background .18s ease, border-color .18s ease, color .18s ease, transform .18s ease, box-shadow .18s ease !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(219, 234, 254, 0.68) !important;
        border-color: rgba(96, 165, 250, 0.22) !important;
        color: #1D4ED8 !important;
        transform: translateX(1px) !important;
        box-shadow: 0 10px 24px rgba(37, 99, 235, 0.09) !important;
    }
    [data-testid="stSidebar"] .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, rgba(219, 234, 254, 0.96), rgba(237, 233, 254, 0.92)) !important;
        color: #0B57D0 !important;
        border: 1px solid rgba(147, 197, 253, 0.42) !important;
        border-left: 4px solid #2563EB !important;
        padding-left: 14px !important;
        font-weight: 700 !important;
        box-shadow: 0 12px 26px rgba(37, 99, 235, 0.12) !important;
    }
    [data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, rgba(191, 219, 254, 0.94), rgba(221, 214, 254, 0.92)) !important;
        color: #1E40AF !important;
    }
    [data-testid="stSidebar"] .nav-toolbar-row {
        display: none !important;
    }
    [data-testid="stSidebar"] .nav-toolbar-row .stButton > button {
        min-height: 32px !important;
        height: 32px !important;
        justify-content: center !important;
        border-radius: 12px !important;
        background: rgba(255, 255, 255, 0.56) !important;
        border: 1px solid rgba(148, 163, 184, 0.24) !important;
        color: #334155 !important;
    }
    [data-testid="stSidebar"] hr {
        background: rgba(148, 163, 184, 0.26) !important;
        margin: 12px 6px !important;
    }
    [data-testid="stSidebar"] .nav-signout .stButton > button {
        justify-content: center !important;
        background: rgba(255, 255, 255, 0.64) !important;
        border: 1px solid rgba(148, 163, 184, 0.24) !important;
        color: #1E293B !important;
        min-height: 56px !important;
        height: 56px !important;
        margin-top: 34vh !important;
    }
    [data-testid="stSidebar"] .nav-signout .stButton > button:hover {
        background: rgba(254, 242, 242, 0.86) !important;
        color: #DC2626 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # V5.1 ICON RAIL — compact by default, expands as an overlay on hover.
    # Keeping the icon as the first character of every button means the rail
    # remains usable even when the labels are clipped.
    st.sidebar.markdown("""
    <style>
    section[data-testid="stSidebar"]{
        width:88px!important;min-width:88px!important;max-width:88px!important;
        overflow:visible!important;z-index:999!important;
        transition:width .24s cubic-bezier(.2,.8,.2,1),max-width .24s cubic-bezier(.2,.8,.2,1)!important;
    }
    section[data-testid="stSidebar"]:hover,
    section[data-testid="stSidebar"]:focus-within{
        width:292px!important;min-width:292px!important;max-width:292px!important;
    }
    section[data-testid="stSidebar"]>div:first-child{
        width:88px!important;min-width:88px!important;overflow:hidden!important;
        transition:width .24s cubic-bezier(.2,.8,.2,1)!important;
        box-shadow:14px 0 36px rgba(49,46,129,.10)!important;
    }
    section[data-testid="stSidebar"]:hover>div:first-child,
    section[data-testid="stSidebar"]:focus-within>div:first-child{
        width:292px!important;min-width:292px!important;
    }
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"]{
        width:292px!important;min-width:292px!important;
        padding:20px 14px 18px!important;overflow-x:hidden!important;
    }
    [data-testid="stSidebarCollapseButton"]{display:none!important}
    section[data-testid="stSidebar"] .ref-brand,
    section[data-testid="stSidebar"] .ref-profile{
        width:260px!important;white-space:nowrap!important;overflow:hidden!important;
    }
    section[data-testid="stSidebar"] .ref-brand{padding:3px 7px 17px!important}
    section[data-testid="stSidebar"] .ref-profile{padding:10px 7px!important;background:transparent!important;border-color:transparent!important;box-shadow:none!important}
    section[data-testid="stSidebar"] .ref-logo,
    section[data-testid="stSidebar"] .ref-avatar{
        width:46px!important;height:46px!important;min-width:46px!important;border-radius:15px!important;
        transition:transform .2s ease!important;
    }
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-brand-title,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-brand-sub,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-profile-dept,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-profile-name,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-status-dot,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-status-dot + span,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .nav-section-label{
        opacity:0!important;pointer-events:none!important;
    }
    section[data-testid="stSidebar"] .ref-brand-title,
    section[data-testid="stSidebar"] .ref-brand-sub,
    section[data-testid="stSidebar"] .ref-profile-dept,
    section[data-testid="stSidebar"] .ref-profile-name,
    section[data-testid="stSidebar"] .nav-section-label{
        transition:opacity .13s ease .08s!important;
    }
    section[data-testid="stSidebar"] .nav-section-label{height:24px!important;margin:0!important;white-space:nowrap!important}
    section[data-testid="stSidebar"] .stButton{width:260px!important;overflow:hidden!important}
    section[data-testid="stSidebar"] .stButton>button{
        width:260px!important;height:50px!important;min-height:50px!important;
        padding:0 18px!important;border-radius:14px!important;white-space:nowrap!important;
        overflow:hidden!important;text-overflow:clip!important;justify-content:flex-start!important;
        font-size:.90rem!important;letter-spacing:.005em!important;
    }
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button{
        width:58px!important;color:#475569!important;font-size:0!important;padding:0!important;
        justify-content:center!important;
    }
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button p{
        width:30px!important;max-width:30px!important;overflow:hidden!important;
        white-space:nowrap!important;font-size:1.15rem!important;line-height:1!important;
    }
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button[kind="primary"]{
        border-left:3px solid #6366F1!important;background:linear-gradient(135deg,#DBEAFE,#EDE9FE)!important;
    }
    section[data-testid="stSidebar"] .nav-signout .stButton>button{margin-top:1rem!important}
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) hr{width:54px!important}
    @media(max-width:768px){
        section[data-testid="stSidebar"]{width:76px!important;min-width:76px!important;max-width:76px!important}
        section[data-testid="stSidebar"]>div:first-child{width:76px!important;min-width:76px!important}
        section[data-testid="stSidebar"]:hover,
        section[data-testid="stSidebar"]:focus-within{width:270px!important;min-width:270px!important;max-width:270px!important}
        section[data-testid="stSidebar"]:hover>div:first-child,
        section[data-testid="stSidebar"]:focus-within>div:first-child{width:270px!important;min-width:270px!important}
    }
    </style>
    """, unsafe_allow_html=True)

    # V5.2 ICON RAIL FIX — real icons, overlay expansion, no content reflow.
    st.sidebar.markdown("""
    <style>
    /* Keep the layout rail narrow; only its inner panel expands above content. */
    section[data-testid="stSidebar"],
    section[data-testid="stSidebar"]:hover,
    section[data-testid="stSidebar"]:focus-within{
        width:80px!important;min-width:80px!important;max-width:80px!important;
        overflow:visible!important;z-index:999!important;
    }
    section[data-testid="stSidebar"]>div:first-child{
        width:80px!important;min-width:80px!important;max-width:none!important;
        overflow:hidden!important;border-right:1px solid rgba(99,102,241,.14)!important;
        background:linear-gradient(180deg,rgba(235,248,255,.97),rgba(238,242,255,.97) 52%,rgba(245,243,255,.97))!important;
        box-shadow:8px 0 26px rgba(49,46,129,.07)!important;
        transition:width .22s cubic-bezier(.2,.8,.2,1),box-shadow .22s ease!important;
    }
    section[data-testid="stSidebar"]:hover>div:first-child,
    section[data-testid="stSidebar"]:focus-within>div:first-child{
        width:286px!important;min-width:286px!important;
        box-shadow:20px 0 48px rgba(49,46,129,.16)!important;
    }
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"]{
        width:286px!important;min-width:286px!important;padding:18px 12px!important;
        overflow-x:hidden!important;background:transparent!important;
    }
    [data-testid="stSidebarCollapseButton"]{display:none!important}

    /* Brand and profile become two clean square marks in rail mode. */
    section[data-testid="stSidebar"] .ref-brand,
    section[data-testid="stSidebar"] .ref-profile{
        width:262px!important;margin:0!important;padding:4px 5px 15px!important;
        border:0!important;border-radius:0!important;background:transparent!important;
        box-shadow:none!important;overflow:hidden!important;white-space:nowrap!important;
    }
    section[data-testid="stSidebar"] .ref-profile{padding-top:12px!important;padding-bottom:12px!important;border-top:1px solid rgba(148,163,184,.18)!important}
    section[data-testid="stSidebar"] .ref-logo,
    section[data-testid="stSidebar"] .ref-avatar{
        width:46px!important;height:46px!important;min-width:46px!important;border-radius:14px!important;
        font-size:.82rem!important;box-shadow:0 8px 18px rgba(99,102,241,.20)!important;
    }
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-brand-title,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-brand-sub,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-profile-dept,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-profile-name,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-status-dot,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ref-status-dot+span,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .nav-section-label{opacity:0!important}
    section[data-testid="stSidebar"] .nav-section-label{width:250px!important;height:25px!important;margin:0!important;padding:7px 8px 4px!important;transition:opacity .12s ease!important}

    /* Neutralise legacy sizes and red/native primary treatments. */
    section[data-testid="stSidebar"] .stButton{width:262px!important;overflow:hidden!important}
    section[data-testid="stSidebar"] .stButton>button{
        width:262px!important;height:48px!important;min-height:48px!important;margin:0!important;
        padding:0 14px!important;border:1px solid transparent!important;border-radius:13px!important;
        background:transparent!important;color:#475569!important;box-shadow:none!important;
        justify-content:flex-start!important;overflow:hidden!important;white-space:nowrap!important;
        transform:none!important;transition:background .16s ease,border-color .16s ease,color .16s ease!important;
    }
    section[data-testid="stSidebar"] .stButton>button:hover{
        background:rgba(255,255,255,.80)!important;border-color:rgba(99,102,241,.14)!important;
        color:#4338CA!important;box-shadow:0 7px 18px rgba(79,70,229,.08)!important;transform:none!important;
    }
    section[data-testid="stSidebar"] .stButton>button[kind="primary"],
    section[data-testid="stSidebar"] .stButton>button[data-testid="stBaseButton-primary"]{
        background:linear-gradient(135deg,#E0EAFF,#EDE9FE)!important;color:#4338CA!important;
        border:1px solid rgba(99,102,241,.18)!important;border-left:3px solid #6366F1!important;
        box-shadow:0 7px 18px rgba(79,70,229,.09)!important;
    }

    /* Collapsed rail: hide labels and render deterministic icons. */
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button{
        width:56px!important;padding:0!important;justify-content:center!important;
    }
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button p{display:none!important}
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button:before{
        content:none!important;display:none!important;width:0;text-align:center;font-size:1.12rem!important;
        line-height:1;color:inherit;font-family:'Segoe UI Symbol','Segoe UI Emoji',sans-serif!important;
    }
    section[data-testid="stSidebar"]:hover .stButton>button:before,
    section[data-testid="stSidebar"]:focus-within .stButton>button:before{display:none!important}

    section[data-testid="stSidebar"] .nav-signout .stButton>button{margin-top:12px!important}
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) hr{width:54px!important;margin:8px 1px!important}
    @media(max-width:768px){
        section[data-testid="stSidebar"],section[data-testid="stSidebar"]:hover,section[data-testid="stSidebar"]:focus-within{width:72px!important;min-width:72px!important;max-width:72px!important}
        section[data-testid="stSidebar"]>div:first-child{width:72px!important;min-width:72px!important}
        section[data-testid="stSidebar"]:hover>div:first-child,section[data-testid="stSidebar"]:focus-within>div:first-child{width:272px!important;min-width:272px!important}
    }
    </style>
    """, unsafe_allow_html=True)

    if compact:
        st.sidebar.markdown(
            '<style>[data-testid="stSidebar"]{min-width:72px!important;}'
            '.hide-when-compact{display:none!important;}</style>',
            unsafe_allow_html=True,
        )

    # =============================================================================
# FUNCTION : _nav_item
#
# PURPOSE
#   สร้างเมนู Sidebar 1 รายการ
#
# USED BY
#   Asset Management / Security / Inventory / Dashboard
#
# VISUAL
#   💻 Computers        [154]
#
# NOTES
#   - รองรับ Badge
#   - รองรับ Active Menu
#   - รองรับ Sub Menu
#   - เปลี่ยนหน้าโดยแก้ st.session_state.active_nav
# =============================================================================
    # Streamlit Cloud may accept Material icon syntax but fail to load the
    # Material Symbols font, which exposes icon names as raw text.  Keep the
    # navigation portable by rendering icons inside the button label instead.
    _ST_BUTTON_HAS_ICON = False
    _NAV_MATERIAL_ICONS = {
        "overview": "space_dashboard", "ad_policy": "policy",
        "hardware_dashboard": "devices", "software_dashboard": "apps",
        "permission_dashboard": "verified_user", "vendor_list": "storefront",
        "computers": "computer", "monitors": "desktop_windows",
        "printers": "print", "projector": "videocam",
        "ups": "battery_charging_full", "misc": "devices_other",
        "password": "key", "user_perm": "folder_shared",
        "ink_stock": "water_drop", "admin_users": "group",
        "admin_settings": "settings", "admin_logs": "history",
    }
    _GROUP_MATERIAL_ICONS = {
        "open_grp_assets": "inventory_2", "open_grp_security": "shield_lock",
        "open_grp_inventory": "inventory", "open_grp_admin": "admin_panel_settings",
    }
    _NAV_FALLBACK_ICONS = {
        "overview": "🏠", "ad_policy": "🛡️", "computers": "💻", "monitors": "🖥️",
        "hardware_dashboard": "🖥️", "software_dashboard": "💿",
        "permission_dashboard": "🔐", "vendor_list": "🏢",
        "printers": "🖨️", "projector": "📽️", "ups": "🔋", "misc": "📦",
        "password": "🔐", "user_perm": "📂", "ink_stock": "💧",
        "admin_users": "👥", "admin_settings": "⚙️", "admin_logs": "🕘",
    }
    _GROUP_FALLBACK_ICONS = {
        "open_grp_assets": "🗃️", "open_grp_security": "🔐",
        "open_grp_inventory": "📚", "open_grp_admin": "🛠️",
    }
    _SAFE_EMOJI_ICONS = {
        "space_dashboard": "🏠", "policy": "🛡️", "computer": "💻",
        "desktop_windows": "🖥️", "print": "🖨️", "videocam": "📽️",
        "battery_charging_full": "🔋", "devices_other": "📦", "key": "🔑",
        "folder_shared": "📂", "water_drop": "💧", "group": "👥",
        "settings": "⚙️", "history": "🕘", "inventory_2": "🗃️",
        "shield_lock": "🔐", "inventory": "📚", "admin_panel_settings": "🛠️",
        "logout": "↪️", "arrow_forward": "→",
    }

    def _button_icon(material_name: str):
        if not _ST_BUTTON_HAS_ICON:
            return {}
        safe_material_name = material_name if re.fullmatch(r"[a-z0-9_]+", material_name or "") else "circle"
        return {"icon": f":material/{safe_material_name}:"}

    def _nav_item(nav_key: str, icon: str, text: str, badge_key: str = None, badge_tone: str = "blue", *, sub: bool = False):
        
       
        
        active = st.session_state.active_nav == nav_key
        val = nav_badges.get(badge_key, 0) if badge_key else None
        label = text if val is None else f"{text}   {val}"
        if not _ST_BUTTON_HAS_ICON:
            label = f"{_NAV_FALLBACK_ICONS.get(nav_key, '·')}  {label}"

        if st.sidebar.button(
            label,
            use_container_width=True,
            type="secondary",
            key=f"nav_{nav_key}_{'active' if active else 'idle'}",
            **_button_icon(_NAV_MATERIAL_ICONS.get(nav_key, "circle")),
        ):
            st.session_state.active_nav = nav_key
            
            st.rerun()

    # =============================================================================
# FUNCTION : _group_toggle
#
# PURPOSE
#   สร้างหัวข้อเมนูแบบพับ/ขยาย
#
# EXAMPLE
#   Asset Management
#   Security
#   Inventory
#
# STATE
#   เก็บสถานะเปิด/ปิดใน st.session_state
# =============================================================================
    def _group_toggle(state_key: str, icon: str, text: str):
        open_ = st.session_state.get(state_key, False)
        label = f"{text}   {'⌄' if open_ else '›'}"
        if not _ST_BUTTON_HAS_ICON:
            label = f"{_GROUP_FALLBACK_ICONS.get(state_key, '◇')}  {label}"
        if st.sidebar.button(label, use_container_width=True, type="secondary",
                             key=f"tog_{state_key}",
                             **_button_icon(_GROUP_MATERIAL_ICONS.get(state_key, "folder"))):
            st.session_state[state_key] = not open_
            st.rerun()

    def _nav_leaf(nav_key: str, icon: str, text: str, badge_text: str = ""):
        _parent_nav = {
            "computers": "hardware_dashboard", "monitors": "hardware_dashboard",
            "printers": "hardware_dashboard", "projector": "hardware_dashboard",
            "ups": "hardware_dashboard", "misc": "hardware_dashboard",
            "cctv": "hardware_dashboard", "access_control": "hardware_dashboard",
            "software_group_email": "software_dashboard", "software_office365": "software_dashboard",
            "software_pdf": "software_dashboard", "software_windows": "software_dashboard",
            "software_offboarded": "software_dashboard",
            "user_perm": "permission_dashboard", "ad_policy": "permission_dashboard",
        }
        active = st.session_state.active_nav == nav_key or _parent_nav.get(st.session_state.active_nav) == nav_key
        label = text
        if not _ST_BUTTON_HAS_ICON:
            label = f"{_NAV_FALLBACK_ICONS.get(nav_key, '·')}  {label}"
        if st.sidebar.button(label, use_container_width=True, type="secondary",
                             key=f"nav_{nav_key}_{'active' if active else 'idle'}",
                             **_button_icon(_NAV_MATERIAL_ICONS.get(nav_key, "circle"))):
            st.session_state.active_nav = nav_key
            st.rerun()

    # ── User profile first (reference image) ────────────────
    st.sidebar.markdown(f"""
    <div class="ce-brand hd-sidebar">
        <div style="display:flex;align-items:center;gap:14px;">
            <div class="ce-logo">DR</div>
            <div style="min-width:0;flex:1;">
                <div class="ce-brand-title">DocumentReportUnified</div>
                <div class="ce-brand-sub">Enterprise IT Platform</div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.sidebar.markdown(f"""
    <div class="ce-profile hd-user-card">
        <div style="display:flex;align-items:center;gap:14px;">
            <div class="ce-avatar">{initials}</div>
            <div style="min-width:0;flex:1;">
                <div class="ce-user-name">{name}</div>
                <div class="ce-user-role">{profile_dept}</div>
                <div class="ce-user-status">
                    <span class="ce-status-dot"></span><span>Online</span>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if compact:
        st.sidebar.markdown(
            f'<div style="text-align:center;padding:8px 0;border-bottom:1px solid #E2E8F0;">'
            f'<div class="ref-avatar" style="margin:0 auto;">{initials}</div></div>',
            unsafe_allow_html=True,
        )

    if False:
        st.sidebar.markdown('<div class="nav-toolbar-row">', unsafe_allow_html=True)
        rt1, rt2 = st.sidebar.columns(2, gap="small")
        with rt1:
            if st.button("◀", use_container_width=True, key="nav_toggle_compact", help="Collapse"):
                st.session_state.sidebar_compact = not compact
                st.rerun()
        with rt2:
            if st.button("↻", use_container_width=True, key="nav_refresh_badges", help="Refresh"):
                get_sidebar_nav_badges.clear()
                st.rerun()
        st.sidebar.markdown("</div>", unsafe_allow_html=True)

    # ── Navigation ───────────────────────────────────────────
    if "active_nav" not in st.session_state:
        st.session_state.active_nav = "overview"
    if st.session_state.active_nav == "reports":
        st.session_state.active_nav = "overview"

    for _gk in ("open_grp_assets", "open_grp_security", "open_grp_inventory", "open_grp_admin"):
        if _gk not in st.session_state:
            st.session_state[_gk] = False

    # Keep the group containing the current page visible after reruns.
    _active_for_group = st.session_state.active_nav
    if _active_for_group in ("computers", "monitors", "printers", "projector", "ups", "misc"):
        st.session_state.open_grp_assets = True
    elif _active_for_group in ("password", "user_perm"):
        st.session_state.open_grp_security = True
    elif _active_for_group in ("ink_stock", "ink_history", "consumables"):
        st.session_state.open_grp_inventory = True
    elif _active_for_group in ("admin_users", "admin_settings", "admin_logs"):
        st.session_state.open_grp_admin = True

    # removed force redirect to computers

    st.sidebar.markdown('<p class="nav-section-label hide-when-compact">WORKSPACE</p>', unsafe_allow_html=True)

    # Flat enterprise navigation. Module details live inside each dashboard.
    _nav_leaf("overview", "⌂", "Dashboard")
    _nav_leaf("hardware_dashboard", "🖥", "Hardware")
    _nav_leaf("software_dashboard", "💿", "Software")
    _nav_leaf("permission_dashboard", "🔐", "Permission")
    _nav_leaf("password", "🔑", "Password")
    _nav_leaf("vendor_list", "🏢", "Vendor List")
    _nav_leaf("ink_stock", "💧", "Ink Stock")

    st.sidebar.markdown("---")

    st.sidebar.markdown('<div class="nav-signout">', unsafe_allow_html=True)
    if st.sidebar.button("Sign out", use_container_width=True, key="logout_btn",
                         **_button_icon("logout")):
        expire_time = datetime.datetime.now() - datetime.timedelta(days=1)
        for cookie_key in ("user_name", "user_email"):
            try:
                cookie_manager.delete(cookie_key, key=f"logout_delete_{cookie_key}")
            except Exception:
                pass
            cookie_manager.set(cookie_key, "", expires_at=expire_time, max_age=0, key=f"logout_set_{cookie_key}")

        cookie_manager.get_all(key="logout_get_all")

        if "cookie_manager" in st.session_state:
            del st.session_state["cookie_manager"]

        st.session_state.is_auth = False
        st.session_state.skip_cookie_login = True
        st.session_state.user_name = ""
        st.session_state.user_email = ""

        st.rerun()
    st.sidebar.markdown("</div>", unsafe_allow_html=True)

    # -----------------------------------------------------------------------
    # V6 TOP NAVIGATION
    # Replaces the icon rail completely.  The legacy sidebar widgets remain
    # non-rendered for backward-compatible state handling only.
    # -----------------------------------------------------------------------
    st.markdown("""
    <style>
    section[data-testid="stSidebar"],
    [data-testid="stSidebarCollapsedControl"],
    [data-testid="stSidebarCollapseButton"]{display:none!important}
    [data-testid="stAppViewContainer"]>.main{margin-left:0!important}
    [data-testid="stMainBlockContainer"]{padding-top:1rem!important}

    .top-appbar{
        min-height:82px;display:flex;align-items:center;justify-content:space-between;gap:24px;
        padding:15px 20px;margin:0 0 10px;border:1px solid rgba(148,163,184,.20);
        border-radius:20px;background:rgba(255,255,255,.90);
        box-shadow:0 10px 32px rgba(15,23,42,.055);backdrop-filter:blur(18px);
    }
    .top-brand{display:flex;align-items:center;gap:13px;min-width:0}
    .top-brand-mark{width:48px;height:48px;display:flex;align-items:center;justify-content:center;
        flex:0 0 48px;border-radius:15px;color:#fff;font-size:.8rem;font-weight:800;letter-spacing:.04em;
        background:linear-gradient(135deg,#38BDF8,#6366F1 55%,#8B5CF6);
        box-shadow:0 9px 22px rgba(99,102,241,.24)}
    .top-brand-name{font-size:1rem;font-weight:800;letter-spacing:-.025em;color:#0F172A;line-height:1.25}
    .top-brand-sub{font-size:.74rem;font-weight:550;color:#64748B;margin-top:3px}
    .top-user{display:flex;align-items:center;justify-content:flex-end;gap:11px;min-width:0}
    .top-user-copy{text-align:right;min-width:0}
    .top-user-name{font-size:.86rem;font-weight:750;color:#1E293B;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
    .top-user-role{font-size:.72rem;color:#64748B;margin-top:2px}
    .top-user-avatar{width:40px;height:40px;display:flex;align-items:center;justify-content:center;
        flex:0 0 40px;border-radius:13px;background:#EEF2FF;color:#4F46E5;font-size:.75rem;font-weight:800;
        border:1px solid #DDE3FF}
    .top-online{display:inline-block;width:7px;height:7px;border-radius:50%;background:#10B981;margin-right:5px}

    [class*="st-key-topnav_"] .stButton>button,
    [class*="st-key-topsub_"] .stButton>button,
    .st-key-top_logout .stButton>button{
        width:100%!important;height:44px!important;min-height:44px!important;padding:0 13px!important;
        border-radius:12px!important;border:1px solid transparent!important;background:transparent!important;
        color:#475569!important;box-shadow:none!important;font-size:.82rem!important;font-weight:700!important;
        white-space:nowrap!important;transform:none!important;transition:.16s ease!important;
    }
    [class*="st-key-topnav_"] .stButton>button:hover,
    [class*="st-key-topsub_"] .stButton>button:hover{
        background:#F1F5FF!important;color:#4338CA!important;border-color:#DDE3FF!important;transform:none!important;
    }
    [class*="st-key-topnav_"] .stButton>button[kind="primary"],
    [class*="st-key-topnav_"] .stButton>button[data-testid="stBaseButton-primary"]{
        background:linear-gradient(135deg,#4F46E5,#6366F1 58%,#7C3AED)!important;
        color:#FFF!important;border-color:transparent!important;box-shadow:0 8px 18px rgba(99,102,241,.22)!important;
    }
    .st-key-top_logout .stButton>button{color:#64748B!important;border-color:#E2E8F0!important;background:#FFF!important}
    .st-key-top_logout .stButton>button:hover{color:#DC2626!important;background:#FFF5F5!important;border-color:#FECACA!important}
    [class*="st-key-topsub_"] .stButton>button{height:38px!important;min-height:38px!important;font-size:.78rem!important;background:#FFF!important;border-color:#E2E8F0!important}
    [class*="st-key-topsub_"] .stButton>button[kind="primary"],
    [class*="st-key-topsub_"] .stButton>button[data-testid="stBaseButton-primary"]{
        background:#EEF2FF!important;color:#4338CA!important;border-color:#C7D2FE!important;box-shadow:none!important;
    }
    .top-nav-rule{height:1px;background:linear-gradient(90deg,transparent,#DDE3EE 8%,#DDE3EE 92%,transparent);margin:7px 0 12px}
    .top-sub-label{font-size:.65rem;font-weight:800;letter-spacing:.13em;color:#94A3B8;text-transform:uppercase;margin:2px 0 7px}
    @media(max-width:900px){
        .top-appbar{min-height:70px;padding:12px 14px;border-radius:16px}.top-brand-mark{width:42px;height:42px;flex-basis:42px}
        .top-brand-sub,.top-user-role{display:none}.top-user-copy{max-width:150px}
        [class*="st-key-topnav_"] .stButton>button{font-size:.75rem!important;padding:0 8px!important}
    }
    /* V6.3 Clean Enterprise final layer */
    [data-testid="stAppViewContainer"],section[data-testid="stMain"]{
        background:radial-gradient(circle at 92% 2%,rgba(99,102,241,.08),transparent 26rem),radial-gradient(circle at 2% 96%,rgba(56,189,248,.07),transparent 24rem),#F6F8FC!important;
    }
    [data-testid="stMainBlockContainer"]{max-width:1500px!important;padding-top:.85rem!important}
    .top-appbar{
        min-height:76px!important;border-radius:24px!important;padding:14px 18px!important;margin-bottom:14px!important;
        background:linear-gradient(135deg,rgba(255,255,255,.96),rgba(248,250,252,.92))!important;
        border:1px solid rgba(203,213,225,.78)!important;box-shadow:0 14px 40px rgba(15,23,42,.065)!important;
    }
    .top-brand-mark{border-radius:16px!important}
    .top-brand-name{font-size:1.08rem!important}
    .top-user-avatar{border-radius:999px!important;background:linear-gradient(135deg,#EEF2FF,#E0EAFF)!important}
    [class*="st-key-topnav_"] .stButton>button{height:46px!important;min-height:46px!important;border-radius:999px!important;background:#FFFFFF!important;border:1px solid #E2E8F0!important;color:#334155!important}
    [class*="st-key-topnav_"] .stButton>button:hover{background:#EEF2FF!important;color:#4338CA!important;border-color:#C7D2FE!important;box-shadow:0 8px 18px rgba(79,70,229,.10)!important}
    [class*="st-key-topnav_"] .stButton>button[kind="primary"],
    [class*="st-key-topnav_"] .stButton>button[data-testid="stBaseButton-primary"]{background:linear-gradient(135deg,#2563EB,#6366F1 56%,#8B5CF6)!important;color:#FFF!important;border-color:transparent!important}
    [class*="st-key-topsub_"] .stButton>button{border-radius:999px!important;background:#FFF!important}
    .top-nav-rule{margin:12px 0 18px!important}
    .page-header{border-radius:26px!important;background:linear-gradient(125deg,#1E3A8A,#4338CA 58%,#7C3AED)!important;box-shadow:0 20px 48px rgba(67,56,202,.20)!important}
    .ov-card{border-radius:22px!important;padding:20px!important;background:linear-gradient(180deg,#FFFFFF,#FBFCFF)!important;border:1px solid #E2E8F0!important;box-shadow:0 12px 32px rgba(15,23,42,.055)!important}
    .ov-card:hover{transform:translateY(-4px)!important;border-color:#C7D2FE!important;box-shadow:0 20px 44px rgba(79,70,229,.13)!important}
    .ov-card-icon{width:46px!important;height:46px!important;border-radius:16px!important;font-size:1.28rem!important}
    .ov-card-label{font-size:.76rem!important;color:#64748B!important}
    .ov-card-value{font-size:2.25rem!important;color:#0F172A!important}
    .ov-card-link{border-top:1px solid #EEF2F7!important;color:#4F46E5!important;font-weight:750!important}
    [data-testid="stMain"] div[data-testid="stVerticalBlockBorderWrapper"]>div{border-radius:22px!important;border-color:#E2E8F0!important;box-shadow:0 10px 28px rgba(15,23,42,.045)!important}
    @media(max-width:900px){
        .top-appbar{min-height:70px;padding:12px 14px;border-radius:18px}.top-brand-mark{width:42px;height:42px;flex-basis:42px}
        .top-brand-sub,.top-user-role{display:none}.top-user-copy{max-width:150px}
        [class*="st-key-topnav_"] .stButton>button{font-size:.75rem!important;padding:0 8px!important}
    }
    @media(max-width:640px){
        .top-user-copy{display:none}.top-appbar{gap:10px}.top-brand-name{font-size:.9rem}
        [class*="st-key-topnav_"] .stButton>button{font-size:.7rem!important;padding:0 5px!important}
    }
    </style>
    """, unsafe_allow_html=True)

    # V7 CLEAN ENTERPRISE SIDEBAR — final navigation layer.
    st.markdown("""
    <style>
    /* Remove the discarded top navigation completely. */
    .top-appbar,.top-nav-rule,.top-sub-label,
    [class*="st-key-topnav_"],[class*="st-key-topsub_"],.st-key-top_logout{display:none!important}
    div[data-testid="stElementContainer"]:has(.top-appbar),
    div[data-testid="stElementContainer"]:has(.top-nav-rule),
    div[data-testid="stElementContainer"]:has(.top-sub-label){display:none!important;height:0!important;min-height:0!important;margin:0!important;padding:0!important}
    [data-testid="stHorizontalBlock"]:has([class*="st-key-topnav_"]),
    [data-testid="stHorizontalBlock"]:has([class*="st-key-topsub_"]){display:none!important;height:0!important;min-height:0!important;margin:0!important;padding:0!important}

    /* Stable full-width sidebar; no hover expansion and no icon rail. */
    section[data-testid="stSidebar"],
    section[data-testid="stSidebar"]:hover,
    section[data-testid="stSidebar"]:focus-within{
        display:flex!important;width:304px!important;min-width:304px!important;max-width:304px!important;
        overflow:hidden!important;z-index:100!important;background:linear-gradient(180deg,#F4FAFF 0%,#F2F5FF 55%,#F7F4FF 100%)!important;
    }
    section[data-testid="stSidebar"]>div:first-child,
    section[data-testid="stSidebar"]:hover>div:first-child,
    section[data-testid="stSidebar"]:focus-within>div:first-child{
        width:304px!important;min-width:304px!important;max-width:304px!important;overflow:hidden!important;
        border-right:1px solid rgba(148,163,184,.24)!important;background:rgba(255,255,255,.76)!important;
        box-shadow:10px 0 34px rgba(15,23,42,.055)!important;backdrop-filter:blur(18px)!important;
    }
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"]{
        width:304px!important;min-width:304px!important;padding:20px 16px 18px!important;overflow-x:hidden!important;
    }
    [data-testid="stSidebarCollapseButton"]{display:none!important}
    section[data-testid="stSidebar"] .ref-brand,
    section[data-testid="stSidebar"] .ref-profile{
        width:auto!important;overflow:visible!important;white-space:normal!important;opacity:1!important;
        background:transparent!important;border-radius:0!important;box-shadow:none!important;
    }
    section[data-testid="stSidebar"] .ref-brand{padding:3px 3px 18px!important;margin:0 0 12px!important;border-bottom:1px solid #E2E8F0!important}
    section[data-testid="stSidebar"] .ref-profile{padding:13px!important;margin:0 0 15px!important;border:1px solid #E2E8F0!important;border-radius:16px!important;background:rgba(255,255,255,.82)!important}
    section[data-testid="stSidebar"] .ref-logo{width:44px!important;height:44px!important;min-width:44px!important;border-radius:13px!important}
    section[data-testid="stSidebar"] .ref-avatar{width:44px!important;height:44px!important;min-width:44px!important;border-radius:999px!important}
    section[data-testid="stSidebar"] .ref-brand-title,
    section[data-testid="stSidebar"] .ref-brand-sub,
    section[data-testid="stSidebar"] .ref-profile-dept,
    section[data-testid="stSidebar"] .ref-profile-name,
    section[data-testid="stSidebar"] .ref-status-dot,
    section[data-testid="stSidebar"] .ref-status-dot+span,
    section[data-testid="stSidebar"] .nav-section-label{opacity:1!important;display:block!important}
    section[data-testid="stSidebar"] .nav-section-label{width:auto!important;height:auto!important;padding:9px 8px 6px!important;color:#94A3B8!important;font-size:.64rem!important;letter-spacing:.14em!important}
    section[data-testid="stSidebar"] [data-testid="stVerticalBlock"]{gap:.28rem!important}
    section[data-testid="stSidebar"] .stButton{width:100%!important;overflow:visible!important}
    section[data-testid="stSidebar"] .stButton>button,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button{
        width:100%!important;height:44px!important;min-height:44px!important;padding:0 13px!important;
        justify-content:flex-start!important;overflow:visible!important;border-radius:11px!important;
        border:1px solid transparent!important;background:transparent!important;color:#475569!important;
        font-size:.84rem!important;font-weight:650!important;box-shadow:none!important;transform:none!important;
    }
    section[data-testid="stSidebar"] .stButton>button p,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button p{
        display:block!important;width:auto!important;max-width:none!important;overflow:visible!important;
        white-space:pre!important;font-size:.84rem!important;line-height:1.25!important;
    }
    section[data-testid="stSidebar"] .stButton>button:before{display:none!important;content:none!important}
    section[data-testid="stSidebar"] .stButton>button:hover{background:#F1F5FF!important;color:#4338CA!important;border-color:#DDE3FF!important;box-shadow:none!important}
    section[data-testid="stSidebar"] .stButton>button[kind="primary"],
    section[data-testid="stSidebar"] .stButton>button[data-testid="stBaseButton-primary"]{
        background:linear-gradient(135deg,#E0EAFF,#EDE9FE)!important;color:#4338CA!important;
        border:1px solid #D5D9FF!important;border-left:3px solid #6366F1!important;
        padding-left:11px!important;box-shadow:0 6px 16px rgba(79,70,229,.08)!important;
    }
    section[data-testid="stSidebar"] [class*="st-key-tog_"] .stButton>button{margin-top:3px!important;font-weight:750!important;color:#334155!important}
    section[data-testid="stSidebar"] .nav-signout .stButton>button{margin-top:10px!important;justify-content:center!important;border-color:#E2E8F0!important;background:#FFF!important;color:#64748B!important}
    section[data-testid="stSidebar"] .nav-signout .stButton>button:hover{background:#FFF5F5!important;color:#DC2626!important;border-color:#FECACA!important}
    section[data-testid="stSidebar"] .ce-brand{padding:2px 2px 18px;margin:0 0 14px;border-bottom:1px solid #E2E8F0}
    section[data-testid="stSidebar"] .ce-logo{width:44px;height:44px;display:flex;align-items:center;justify-content:center;flex:0 0 44px;border-radius:13px;background:linear-gradient(135deg,#38BDF8,#6366F1 55%,#8B5CF6);color:#FFF;font-size:.76rem;font-weight:850;box-shadow:0 8px 20px rgba(99,102,241,.22)}
    section[data-testid="stSidebar"] .ce-brand-title{font-size:.88rem;font-weight:850;letter-spacing:-.025em;color:#0F172A;line-height:1.25}
    section[data-testid="stSidebar"] .ce-brand-sub{margin-top:3px;font-size:.69rem;font-weight:550;color:#64748B}
    section[data-testid="stSidebar"] .ce-profile{padding:13px;margin:0 0 16px;border:1px solid #E2E8F0;border-radius:16px;background:linear-gradient(145deg,#FFF,#F8FAFF);box-shadow:0 7px 20px rgba(15,23,42,.045)}
    section[data-testid="stSidebar"] .ce-avatar{width:42px;height:42px;display:flex;align-items:center;justify-content:center;flex:0 0 42px;border-radius:12px;background:#EEF2FF;color:#4F46E5;border:1px solid #DDE3FF;font-size:.74rem;font-weight:850}
    section[data-testid="stSidebar"] .ce-user-name{font-size:.82rem;font-weight:800;color:#1E293B;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
    section[data-testid="stSidebar"] .ce-user-role{margin-top:2px;font-size:.68rem;color:#64748B;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
    section[data-testid="stSidebar"] .ce-user-status{display:flex;align-items:center;gap:5px;margin-top:7px;font-size:.66rem;font-weight:650;color:#64748B}
    section[data-testid="stSidebar"] .ce-status-dot{width:7px;height:7px;border-radius:50%;background:#10B981;box-shadow:0 0 0 3px rgba(16,185,129,.12)}
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ce-brand,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ce-profile,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ce-brand-title,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ce-brand-sub,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ce-user-name,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ce-user-role,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ce-user-status{display:block!important;opacity:1!important;visibility:visible!important}
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .ce-user-status{display:flex!important}
    section[data-testid="stSidebar"] .stButton>button[data-testid*="primary"],
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button[data-testid*="primary"]{background:linear-gradient(135deg,#4F46E5,#6366F1 58%,#7C3AED)!important;color:#FFF!important;border:1px solid transparent!important;border-left:3px solid #C4B5FD!important;box-shadow:0 8px 18px rgba(79,70,229,.20)!important}

    /* V7.2 SaaS navigation: calm surfaces, consistent Material icons. */
    section[data-testid="stSidebar"] [class*="material-symbols"],
    section[data-testid="stSidebar"] span[data-testid="stIconMaterial"]{
        font-size:19px!important;font-weight:400!important;color:currentColor!important;
    }
    section[data-testid="stSidebar"] .stButton>button,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button{
        height:42px!important;min-height:42px!important;padding:0 12px!important;border-radius:10px!important;
        background:transparent!important;border:1px solid transparent!important;color:#475569!important;
        font-size:.81rem!important;font-weight:650!important;box-shadow:none!important;gap:10px!important;
    }
    section[data-testid="stSidebar"] .stButton>button:hover{
        background:#F6F7FF!important;border-color:#E7E9FF!important;color:#4F46E5!important;box-shadow:none!important;
    }
    section[data-testid="stSidebar"] .stButton>button[data-testid*="primary"],
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) .stButton>button[data-testid*="primary"]{
        background:#EEF2FF!important;color:#4338CA!important;border:1px solid #DDE3FF!important;
        border-left:1px solid #DDE3FF!important;padding-left:12px!important;box-shadow:none!important;font-weight:750!important;
    }
    section[data-testid="stSidebar"] [class*="st-key-nav_"][class*="_active"] .stButton>button,
    section[data-testid="stSidebar"]:not(:hover):not(:focus-within) [class*="st-key-nav_"][class*="_active"] .stButton>button{
        background:linear-gradient(135deg,#2563EB 0%,#4F46E5 55%,#7C3AED 100%)!important;
        color:#FFF!important;border:1px solid transparent!important;box-shadow:0 9px 22px rgba(79,70,229,.24)!important;
    }
    section[data-testid="stSidebar"] [class*="st-key-nav_"][class*="_active"] .stButton>button:hover{
        background:linear-gradient(135deg,#1D4ED8 0%,#4338CA 55%,#6D28D9 100%)!important;color:#FFF!important;
    }
    section[data-testid="stSidebar"] [class*="st-key-tog_"] .stButton>button{
        margin-top:5px!important;background:transparent!important;border-color:transparent!important;
        color:#334155!important;font-weight:750!important;
    }
    section[data-testid="stSidebar"] [class*="st-key-tog_"] .stButton>button:hover{
        background:#F8FAFC!important;border-color:#EEF2F7!important;color:#4338CA!important;
    }
    section[data-testid="stSidebar"] [class*="st-key-tog_"] .stButton>button:before,
    section[data-testid="stSidebar"] [class*="st-key-nav_"] .stButton>button:before,
    section[data-testid="stSidebar"] .st-key-logout_btn .stButton>button:before{
        display:none!important;content:none!important;
    }
    section[data-testid="stSidebar"] .stButton[class*="st-key-tog_"]>button:before,
    section[data-testid="stSidebar"] .stButton[class*="st-key-nav_"]>button:before,
    section[data-testid="stSidebar"] .stButton.st-key-logout_btn>button:before,
    section[data-testid="stSidebar"] [class*="st-key-tog_"]>button:before,
    section[data-testid="stSidebar"] [class*="st-key-nav_"]>button:before{
        display:none!important;content:none!important;width:0!important;margin:0!important;
    }
    section[data-testid="stSidebar"] .stButton[class*="st-key-nav_"][class*="_active"]>button,
    section[data-testid="stSidebar"] [class*="st-key-nav_"][class*="_active"]>button{
        background:linear-gradient(135deg,#2563EB 0%,#4F46E5 55%,#7C3AED 100%)!important;
        color:#FFF!important;border:1px solid transparent!important;box-shadow:0 9px 22px rgba(79,70,229,.24)!important;
    }
    section[data-testid="stSidebar"] .st-key-nav_computers .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_monitors .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_printers .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_projector .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_ups .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_misc .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_password .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_user_perm .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_ink_stock .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_admin_users .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_admin_settings .stButton>button,
    section[data-testid="stSidebar"] .st-key-nav_admin_logs .stButton>button{
        width:calc(100% - 12px)!important;margin-left:12px!important;height:39px!important;min-height:39px!important;
        color:#64748B!important;font-size:.77rem!important;
    }
    section[data-testid="stSidebar"] .nav-signout .stButton>button{
        height:42px!important;min-height:42px!important;margin-top:8px!important;background:transparent!important;
        border-color:transparent!important;color:#64748B!important;justify-content:flex-start!important;
    }
    section[data-testid="stSidebar"] .nav-signout .stButton>button:hover{
        background:#FFF5F5!important;border-color:#FEE2E2!important;color:#DC2626!important;
    }
    /* Authenticated pages: remove Streamlit's reserved cloud header space. */
    header[data-testid="stHeader"],[data-testid="stToolbar"],[data-testid="stDecoration"],
    [data-testid="stStatusWidget"],.stAppHeader{display:none!important;height:0!important;min-height:0!important}
    [data-testid="stAppViewContainer"]{padding-top:0!important;background:radial-gradient(circle at 8% 8%,rgba(56,189,248,.26),transparent 28rem),radial-gradient(circle at 92% 88%,rgba(139,92,246,.23),transparent 31rem),linear-gradient(145deg,#EFF8FF,#EEF2FF 48%,#F5F3FF)!important}
    section[data-testid="stMain"]{padding-top:0!important;margin-top:0!important;background:transparent!important}
    section[data-testid="stMain"] [data-testid="stMainBlockContainer"],
    section[data-testid="stMain"] .block-container{max-width:1500px!important;padding:6px clamp(1rem,2.6vw,2.5rem) 4rem!important;margin-top:0!important}
    section[data-testid="stSidebar"] [data-testid="stSidebarHeader"]{display:none!important;height:0!important;min-height:0!important;padding:0!important;margin:0!important}
    section[data-testid="stSidebar"] [data-testid="stSidebarUserContent"],
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"]{padding-top:6px!important;margin-top:0!important;background:transparent!important}
    @media(max-width:900px){
        section[data-testid="stSidebar"],section[data-testid="stSidebar"]:hover,section[data-testid="stSidebar"]:focus-within{width:274px!important;min-width:274px!important;max-width:274px!important}
        section[data-testid="stSidebar"]>div:first-child,section[data-testid="stSidebar"]:hover>div:first-child{width:274px!important;min-width:274px!important;max-width:274px!important}
        section[data-testid="stSidebar"] [data-testid="stSidebarContent"]{width:274px!important;min-width:274px!important}
    }
    </style>
    """, unsafe_allow_html=True)

    def _top_category(nav_key: str) -> str:
        if nav_key in ("computers", "monitors", "printers", "projector", "ups", "misc"):
            return "assets"
        if nav_key in ("password", "user_perm"):
            return "security"
        if nav_key in ("ink_stock", "ink_history", "consumables"):
            return "inventory"
        if nav_key in ("admin_users", "admin_settings", "admin_logs"):
            return "admin"
        return nav_key

    _current_category = _top_category(st.session_state.active_nav)
    _top_items = [("overview", "ภาพรวม", "overview"), ("assets", "สินทรัพย์ IT", "computers")]
    if admin_mode:
        _top_items = [
            ("overview", "ภาพรวม", "overview"),
            ("ad_policy", "AD / Firewall", "ad_policy"),
            ("assets", "สินทรัพย์ IT", "computers"),
            ("security", "ความปลอดภัย", "password"),
            ("inventory", "คลังสินค้า", "ink_stock"),
            ("admin", "ผู้ดูแลระบบ", "admin_users"),
        ]

    _subnav = {
        "assets": [("computers", "คอมพิวเตอร์"), ("monitors", "จอภาพ"), ("printers", "เครื่องพิมพ์"),
                   ("projector", "โปรเจกเตอร์"), ("ups", "UPS"), ("misc", "อุปกรณ์อื่น")],
        "security": [("password", "Password Manager"), ("user_perm", "NAS Permission")],
        "inventory": [("ink_stock", "สต็อกหมึก")],
        "admin": [("admin_users", "ผู้ใช้งาน"), ("admin_settings", "ตั้งค่า"), ("admin_logs", "Activity Logs")],
    }.get(_current_category, [])

    # The legacy top navigation is no longer rendered; the sidebar owns routing.

    _nav = st.session_state.active_nav

    # ── ROUTE ────────────────────────────────────────────────────────────────
    _ROUTE = {
        "overview":   ("📊 Overview Dashboard",  None),
        "hardware_dashboard": ("🖥 Hardware Dashboard", None),
        "software_dashboard": ("💿 Software Dashboard", None),
        "permission_dashboard": ("🔐 Permission Dashboard", None),
        "computers":  ("💻 Hardware Asset",       "Computer Asset"),
        "monitors":   ("💻 Hardware Asset",       "Asset Monitor"),
        "projector":  ("💻 Hardware Asset",       "Asset Projector"),
        "printers":   ("💻 Hardware Asset",       "Asset Printer"),
        "ups":        ("💻 Hardware Asset",       "Asset UPS"),
        "misc":       ("💻 Hardware Asset",       "Asset Misc"),
        "cctv":       ("💻 Hardware Asset",       "Asset CCTV"),
        "access_control": ("💻 Hardware Asset",    "Asset Access Control"),
        "software_group_email": ("💿 Software Module", "Group E-mail"),
        "software_office365": ("💿 Software Module", "Office 365"),
        "software_pdf": ("💿 Software Module", "PDF"),
        "software_windows": ("💿 Software Module", "Windows"),
        "software_offboarded": ("💿 Software Module", "พนักงานลาออก"),
        "email":      ("🔑 Password Information", None),
        "domain":     ("🔑 Password Information", None),
        "vendor":     ("🔑 Password Information", None),
        "vendor_list": ("🏢 Vendor List", None),
        "user_perm":  ("📂 NAS Drive Check",      None),
        "password":   ("🔑 Password Information", None),
        "ink_stock":  ("🖨️ Stock หมึกพิมพ์",       None),
        "ink_history":("🖨️ Stock หมึกพิมพ์",       None),
        "consumables":("🖨️ Stock หมึกพิมพ์",       None),
        "ad_policy":      ("🌐 AD / Firewall Policy", None),
        "admin_users":    ("⚙ Administration", None),
        "admin_settings": ("⚙ Administration", None),
        "admin_logs":     ("⚙ Administration", None),
}
    main_menu, _hw_sub_override = _ROUTE.get(_nav, ("📊 Overview Dashboard", None))
    show_ink_history_only = (_nav in ("ink_history", "consumables"))

    if not admin_mode and main_menu not in ("📊 Overview Dashboard", "🖥 Hardware Dashboard", "💻 Hardware Asset"):
        st.session_state.active_nav = "overview"
        _nav = "overview"
        main_menu, _hw_sub_override = _ROUTE["overview"]
        show_ink_history_only = False
        st.warning("🔒 สิทธิ์การใช้งานถูกจำกัด: ผู้ใช้ทั่วไปสามารถเข้าถึงได้เฉพาะ Dashboard และ Hardware Asset เท่านั้น")

    def _render_module_hub(title, subtitle, icon, modules):
        """Render a flat module dashboard without adding sidebar submenus."""
        st.markdown("""
        <style>
        .hub-hero{position:relative;overflow:hidden;padding:25px 28px;margin-bottom:16px;border-radius:24px;color:#FFF;background:linear-gradient(125deg,#2563EB,#6366F1 56%,#8B5CF6);box-shadow:0 16px 38px rgba(79,70,229,.20)}
        .hub-hero:after{content:'';position:absolute;width:260px;height:260px;right:-75px;top:-135px;border-radius:50%;background:rgba(255,255,255,.10)}.hub-title{font-size:28px;font-weight:850;letter-spacing:-.035em}.hub-sub{margin-top:5px;font-size:13px;color:rgba(255,255,255,.82)}
        [class*="st-key-hub_card_"] .stButton>button{position:relative!important;align-items:flex-start!important;justify-content:flex-start!important;width:100%!important;min-height:132px!important;padding:18px 48px 18px 18px!important;border:1px solid #E2E8F0!important;border-radius:18px!important;background:#FFF!important;color:#172554!important;text-align:left!important;white-space:pre-line!important;box-shadow:0 7px 20px rgba(15,23,42,.045)!important}
        [class*="st-key-hub_card_"] .stButton>button:after{content:'›';position:absolute;right:14px;bottom:14px;display:grid;place-items:center;width:27px;height:27px;border:1px solid #C7D2FE;border-radius:50%;color:#4F46E5;font-size:16px;background:#FFF}
        [class*="st-key-hub_card_"] .stButton>button:hover{transform:translateY(-3px)!important;border-color:#A5B4FC!important;box-shadow:0 14px 30px rgba(79,70,229,.11)!important}.hub-note{margin:14px 0 8px;color:#64748B;font-size:12px}
        [class*="st-key-hub_card_"] .stButton>button p{width:100%;white-space:pre-line!important;text-align:left!important;color:#64748B!important;font-size:12px!important;line-height:1.55!important;font-weight:500!important}[class*="st-key-hub_card_"] .stButton>button p:first-line{color:#172554!important;font-size:15px!important;font-weight:850!important}
        </style>
        """, unsafe_allow_html=True)
        st.markdown(f'<div class="hub-hero"><div class="hub-title">{icon} {title}</div><div class="hub-sub">{subtitle}</div></div>', unsafe_allow_html=True)
        _hub_cols = st.columns(3, gap="medium")
        for _hub_index, (_target, _item_icon, _item_title, _item_desc) in enumerate(modules):
            with _hub_cols[_hub_index % 3]:
                if st.button(f"{_item_icon}  {_item_title}\n{_item_desc}", key=f"hub_card_{title}_{_hub_index}_{_target}", use_container_width=True):
                    st.session_state.active_nav = _target
                    st.rerun()

    def _render_password_sheet_module(title, subtitle, icon, keywords):
        """Reuse the existing password workbook safely for software/vendor modules."""
        page_header(icon, title, subtitle)
        with st.spinner("กำลังโหลดข้อมูล..."):
            _module_result = load_password_excel()
            _module_sheets, _module_drive_id = _module_result if isinstance(_module_result, tuple) else ({}, None)
        if not _module_sheets or "_error" in _module_sheets:
            st.error("ไม่สามารถโหลดข้อมูลจาก SharePoint ได้")
            return
        _module_sheet_name = None
        for _candidate in _module_sheets:
            _candidate_key = str(_candidate).lower().replace("_", " ").replace("-", " ")
            if any(str(_keyword).lower() in _candidate_key for _keyword in keywords):
                _module_sheet_name = _candidate
                break
        if not _module_sheet_name:
            st.info(f"ยังไม่พบหมวดข้อมูลสำหรับ {title} ในไฟล์ปัจจุบัน")
            return
        _module_df = _module_sheets[_module_sheet_name].copy()
        _module_header, _module_add = st.columns([0.8, 0.2])
        with _module_header:
            st.subheader(f"{get_sheet_icon(_module_sheet_name)} {_module_sheet_name}")
            st.caption(f"พบข้อมูลทั้งหมด {len(_module_df)} รายการ")
        with _module_add:
            if admin_mode and st.button("➕ เพิ่มรายการ", key=f"module_add_{_nav}", use_container_width=True, type="primary"):
                add_password_dialog(_module_sheet_name, _module_df, _module_drive_id, _module_sheets)
        if _module_df.empty:
            st.info("ยังไม่มีข้อมูลในหมวดนี้")
            return
        _module_cols = st.columns(2)
        for _module_pos, (_module_idx, _module_row) in enumerate(_module_df.iterrows()):
            with _module_cols[_module_pos % 2]:
                render_password_card(_module_row, _module_sheet_name, _module_idx, admin_mode, _module_df, _module_drive_id, _module_sheets)
                if admin_mode and st.session_state.get(f"pw_edit_row_{_module_sheet_name}_{_module_idx}"):
                    st.session_state.pop(f"pw_edit_row_{_module_sheet_name}_{_module_idx}")
                    edit_password_dialog(_module_row, _module_idx, _module_sheet_name, _module_df, _module_drive_id, _module_sheets)

    def _render_hardware_command_center():
        """Enterprise hardware command center built only from existing asset data."""
        def _hd_svg(kind):
            _paths = {
                "dashboard": '<rect x="3" y="4" width="18" height="13" rx="2"/><path d="M8 21h8M12 17v4"/><path d="m7 12 3-3 2 2 5-5"/>',
                "assets": '<rect x="4" y="4" width="16" height="12" rx="2"/><path d="M8 20h8M12 16v4"/>',
                "online": '<rect x="3" y="4" width="18" height="13" rx="2"/><path d="M8 21h8M12 17v4M8 10l3 3 5-6"/>',
                "maintenance": '<path d="m14.7 6.3 3-3a4 4 0 0 1-5 5L5 16l-2 5 5-2 7.7-7.7a4 4 0 0 1 5-5l-3 3"/>',
                "offline": '<rect x="3" y="4" width="18" height="13" rx="2"/><path d="M8 21h8M12 17v4M9 8l6 6M15 8l-6 6"/>',
                "computer": '<rect x="3" y="4" width="18" height="13" rx="2"/><path d="M8 21h8M12 17v4"/>',
                "monitor": '<rect x="3" y="3" width="18" height="15" rx="2"/><path d="M7 21h10"/>',
                "printer": '<path d="M6 9V3h12v6"/><rect x="6" y="14" width="12" height="7" rx="1"/><path d="M6 17H4a2 2 0 0 1-2-2v-4a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v4a2 2 0 0 1-2 2h-2"/>',
                "projector": '<rect x="3" y="6" width="18" height="11" rx="2"/><circle cx="15.5" cy="11.5" r="3"/><path d="M7 17v3M17 17v3M6 10h3"/>',
                "ups": '<rect x="6" y="3" width="12" height="18" rx="2"/><path d="M10 7h4M9 16h6M12 12v4"/>',
                "cctv": '<path d="m4 8 12-4 2 6-12 4-2-6Z"/><path d="m15 11 3 4M18 15h3M9 13v5M6 21h8"/>',
                "access": '<rect x="5" y="3" width="14" height="18" rx="2"/><path d="M9 3v18M15 12h.01"/>',
            }
            return f'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">{_paths[kind]}</svg>'

        _hd_categories = [
            ("computers", "Computers", "Computer Asset", _hd_svg("computer"), "#3B82F6", "#EAF3FF"),
            ("monitors", "Monitors", "Asset Monitor", _hd_svg("monitor"), "#7C3AED", "#F2ECFF"),
            ("printers", "Printers", "Asset Printer", _hd_svg("printer"), "#C026D3", "#FAE8FF"),
            ("projector", "Projector", "Asset Projector", _hd_svg("projector"), "#DB2777", "#FCE7F3"),
            ("ups", "UPS", "Asset UPS", _hd_svg("ups"), "#10B981", "#E7F8EE"),
            ("cctv", "CCTV", "Asset CCTV", _hd_svg("cctv"), "#F59E0B", "#FFF4E5"),
            ("access_control", "Access Control", "Asset Access Control", _hd_svg("access"), "#4F46E5", "#EEF2FF"),
        ]

        def _hd_load(list_name):
            try:
                frame = load_sp_data(list_name)
                return frame.copy() if isinstance(frame, pd.DataFrame) else pd.DataFrame()
            except Exception:
                return pd.DataFrame()

        with st.spinner("กำลังรวบรวมข้อมูล Hardware..."):
            _hd_frames = {key: _hd_load(list_name) for key, _, list_name, _, _, _ in _hd_categories}
            _hd_misc = _hd_load("Asset Misc")

        # Reuse Miscellaneous records when dedicated CCTV / Access Control lists do not exist.
        if not _hd_misc.empty:
            _misc_text = _hd_misc.fillna("").astype(str).agg(" ".join, axis=1).str.lower()
            if _hd_frames["cctv"].empty:
                _hd_frames["cctv"] = _hd_misc[_misc_text.str.contains("cctv|camera|กล้อง", regex=True)].copy()
            if _hd_frames["access_control"].empty:
                _hd_frames["access_control"] = _hd_misc[_misc_text.str.contains("access control|door|ประตู|สแกน", regex=True)].copy()

        st.markdown(f'<div class="hd-dashboard"><div class="hd-header"><div class="hd-header-icon">{_hd_svg("dashboard")}</div><div><div class="hd-header-title">Hardware Dashboard</div><div class="hd-header-sub">ศูนย์รวมสินทรัพย์และอุปกรณ์ทั้งหมดขององค์กร</div></div></div></div>', unsafe_allow_html=True)
        _hd_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7)))

        def _hd_text(value):
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return ""
            return str(value).strip()

        def _hd_row_value(row, candidates=(), contains=()):
            _columns = {str(column).lower(): column for column in row.index}
            for candidate in candidates:
                if candidate.lower() in _columns:
                    value = _hd_text(row.get(_columns[candidate.lower()]))
                    if value:
                        return value
            for column in row.index:
                _column_key = str(column).lower().replace("_x0020_", " ")
                if any(token in _column_key for token in contains):
                    value = _hd_text(row.get(column))
                    if value:
                        return value
            return ""

        def _hd_status(row):
            raw = _hd_row_value(row, ("Status", "AssetStatus", "ComputerStatus", "state"), ("status", "สถานะ")).lower()
            if not raw:
                return "unknown"
            if any(token in raw for token in ("repair", "maintenance", "ซ่อม", "ปรับปรุง")):
                return "maintenance"
            if any(token in raw for token in ("inactive", "offline", "down", "เสีย", "ยกเลิก")):
                return "offline"
            if any(token in raw for token in ("active", "online", "normal", "ใช้งาน", "ปกติ", "พร้อม")):
                return "online"
            return "unknown"

        _hd_category_stats = []
        _hd_records = []
        for _key, _label, _list_name, _icon, _tone, _soft in _hd_categories:
            _frame = _hd_frames[_key]
            _counts = {"online": 0, "maintenance": 0, "offline": 0, "unknown": 0}
            for _row_index, _row in _frame.iterrows():
                _state = _hd_status(_row)
                _counts[_state] += 1
                _hd_records.append((_key, _label, _row_index, _row, _state))
            _hd_category_stats.append({"key": _key, "label": _label, "icon": _icon, "tone": _tone, "soft": _soft, "total": len(_frame), **_counts})

        _hd_total = len(_hd_records)
        _hd_online = sum(1 for *_, state in _hd_records if state == "online")
        _hd_maintenance = sum(1 for *_, state in _hd_records if state == "maintenance")
        _hd_offline = sum(1 for *_, state in _hd_records if state == "offline")

        _hd_locations = {}
        _hd_activity = []
        _hd_month_start = pd.Timestamp(_hd_now.replace(day=1, hour=0, minute=0, second=0, microsecond=0, tzinfo=None))
        _hd_months = [_hd_month_start - pd.DateOffset(months=offset) for offset in range(5, -1, -1)]
        _hd_trend = {"online": [0] * 6, "maintenance": [0] * 6, "offline": [0] * 6, "unknown": [0] * 6}

        for _key, _label, _row_index, _row, _state in _hd_records:
            _location = _hd_row_value(_row, ("Location", "Site", "Branch", "Company", "field_1"), ("location", "site", "branch", "company", "สถานที่", "บริษัท")) or "ไม่ระบุ"
            _hd_locations[_location] = _hd_locations.get(_location, 0) + 1

            _modified_raw = _hd_row_value(_row, ("Modified", "LastModified", "Updated", "LastSeen"), ("modified", "updated", "last seen"))
            _created_raw = _hd_row_value(_row, ("Created", "CreatedDate"), ("created",))
            _modified = pd.to_datetime(_modified_raw, errors="coerce")
            _created = pd.to_datetime(_created_raw, errors="coerce")
            if not pd.isna(_modified):
                if getattr(_modified, "tzinfo", None) is not None:
                    _modified = _modified.tz_localize(None)
                for _month_index, _month in enumerate(_hd_months):
                    if _modified.year == _month.year and _modified.month == _month.month:
                        _hd_trend[_state][_month_index] += 1
                        break
                _asset_name = _hd_row_value(_row, ("Title", "ComputerName", "Hostname", "field_6", "AssetName", "Name", "Model", "field_2"), ("hostname", "asset name", "title", "model")) or _label
                _editor = _hd_row_value(_row, ("Editor", "Modified By", "ModifiedBy", "Author"), ("editor", "modified by")) or "System"
                _action = "ส่งซ่อม" if _state == "maintenance" else ("อุปกรณ์ Offline" if _state == "offline" else "แก้ไขข้อมูลอุปกรณ์")
                if not pd.isna(_created):
                    try:
                        if abs((_modified - _created).total_seconds()) < 120:
                            _action = "เพิ่มอุปกรณ์ใหม่"
                    except Exception:
                        pass
                _hd_activity.append((_modified, _action, _asset_name, _editor, _state))

        _hd_activity.sort(key=lambda item: item[0], reverse=True)

        _hd_total_trend = [sum(_hd_trend[state][idx] for state in _hd_trend) for idx in range(6)]
        _hd_kpis = [
            ("Total Assets", _hd_total, "รายการจาก SharePoint", _hd_svg("assets"), "#3B82F6", "#EAF3FF"),
            ("Online", _hd_online, "สถานะ Active / Online", _hd_svg("online"), "#10B981", "#E7F8EE"),
            ("Under Maintenance", _hd_maintenance, "สถานะ Repair / Maintenance", _hd_svg("maintenance"), "#F59E0B", "#FFF4E5"),
            ("Offline", _hd_offline, "สถานะ Inactive / Offline", _hd_svg("offline"), "#EF4444", "#FEECEF"),
        ]

        st.markdown("""
        <style>
        .hd-dashboard{color:#0F172A}.hd-header{display:flex;align-items:center;gap:14px;min-height:70px}.hd-header-icon{display:grid;place-items:center;width:58px;height:58px;border-radius:17px;background:linear-gradient(135deg,#E0E7FF,#F3E8FF);color:#4F46E5;box-shadow:0 9px 20px rgba(79,70,229,.10)}.hd-header-icon svg{width:29px;height:29px}.hd-header-title{font-size:30px;font-weight:850;letter-spacing:-.04em}.hd-header-sub{margin-top:4px;color:#64748B;font-size:13px}.hd-search{display:none}
        .hd-kpi-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px;margin:14px 0}.hd-kpi-card{position:relative;height:124px;padding:15px;border:1px solid #E2E8F0;border-radius:22px;background:#FFF;box-shadow:0 8px 22px rgba(15,23,42,.05);overflow:hidden}.hd-kpi-label{font-size:12px;font-weight:800;color:#334155}.hd-kpi-value{margin-top:13px;font-size:36px;line-height:1;font-weight:850;letter-spacing:-.05em;color:#0F172A;white-space:nowrap}.hd-kpi-note{margin-top:9px;color:#64748B;font-size:10px}.hd-kpi-icon{position:absolute;right:14px;top:14px;display:grid;place-items:center;width:42px;height:42px;border-radius:50%}.hd-kpi-icon svg{width:21px;height:21px}
        .hd-section-title{margin:0 0 11px;font-size:15px;font-weight:850;color:#334155}.hd-category-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:9px}.hd-category-card{min-height:126px}.hd-category-head{display:flex;align-items:center;gap:11px}.hd-category-icon{display:grid;place-items:center;flex:0 0 44px;width:44px;height:44px;border-radius:13px}.hd-category-icon svg{width:23px;height:23px}.hd-category-title{font-size:13px;font-weight:850}.hd-category-total{margin-top:3px;color:#64748B;font-size:10px}.hd-category-status{display:grid;grid-template-columns:repeat(3,1fr);gap:5px;margin:11px 0 6px;color:#64748B;font-size:9px}.hd-category-status b{display:block;margin-top:2px;font-size:11px}.hd-category-unknown{margin-bottom:8px;color:#94A3B8;font-size:8.5px}.hd-progress{height:6px;border-radius:999px;background:#EDF2F7;overflow:hidden}.hd-progress>span{display:block;height:100%;border-radius:999px}
        [class*="st-key-hd_category_card_"]{position:relative!important;padding:13px!important;border:1px solid #E2E8F0!important;border-radius:17px!important;background:#FFF!important;box-shadow:0 6px 18px rgba(15,23,42,.04)!important;transition:.18s ease!important}[class*="st-key-hd_category_card_"]:hover{transform:translateY(-2px);border-color:#C7D2FE!important;box-shadow:0 12px 25px rgba(79,70,229,.09)!important}
        [class*="st-key-hd_category_card_"] [class*="st-key-hd_open_"]{position:static!important;width:100%!important;height:32px!important;margin-top:9px!important}[class*="st-key-hd_category_card_"] [class*="st-key-hd_open_"] .stButton,[class*="st-key-hd_category_card_"] [class*="st-key-hd_open_"] button{display:flex!important;justify-content:flex-end!important;width:100%!important;height:32px!important;min-height:32px!important;padding:0 10px!important;margin:0!important;border:1px solid #E0E7FF!important;border-radius:10px!important;color:#4F46E5!important;background:#F8FAFF!important;box-shadow:none!important;font-size:10px!important;font-weight:800!important}[class*="st-key-hd_category_card_"] [class*="st-key-hd_open_"] button p{font-size:10px!important;font-weight:800!important;color:inherit!important}[class*="st-key-hd_category_card_"] [class*="st-key-hd_open_"] button:hover{color:#4338CA!important;background:#EEF2FF!important;border-color:#C7D2FE!important;transform:none!important}
        [class*="st-key-hd_category_card_access_control"] .hd-category-card{display:grid;grid-template-columns:minmax(220px,1.2fr) minmax(260px,1fr);grid-template-rows:auto auto auto;align-items:center;column-gap:24px;min-height:78px}[class*="st-key-hd_category_card_access_control"] .hd-category-head{grid-column:1;grid-row:1/3}[class*="st-key-hd_category_card_access_control"] .hd-category-status{grid-column:2;grid-row:1;margin:0 40px 3px 0}[class*="st-key-hd_category_card_access_control"] .hd-category-unknown{grid-column:2;grid-row:2;margin:0 40px 5px 0}[class*="st-key-hd_category_card_access_control"] .hd-progress{grid-column:1/3;grid-row:3;margin-top:7px}
        .hd-chart-panel,.hd-table-panel,.hd-activity-panel{padding:14px;border:1px solid #E2E8F0;border-radius:18px;background:#FFF;box-shadow:0 7px 20px rgba(15,23,42,.045)}.hd-chart-panel{min-height:210px}.hd-chart-title{margin-bottom:10px;font-size:12px;font-weight:850}.hd-donut-layout{display:flex;align-items:center;gap:14px}.hd-donut{position:relative;width:118px;height:118px;flex:0 0 118px;border-radius:50%}.hd-donut:after{content:'';position:absolute;inset:22px;border-radius:50%;background:#FFF}.hd-donut-center{position:absolute;z-index:2;inset:0;display:grid;place-content:center;text-align:center;font-size:10px;color:#64748B}.hd-donut-center b{font-size:21px;color:#0F172A}.hd-legend{display:grid;gap:8px;flex:1}.hd-legend-row{display:grid;grid-template-columns:8px 1fr auto;align-items:center;gap:7px;color:#64748B;font-size:9px}.hd-legend-row i{width:8px;height:8px;border-radius:50%}.hd-legend-row b{color:#334155}.hd-trend{margin-top:10px}.hd-trend svg{width:100%;height:130px}.hd-trend-labels{display:flex;justify-content:space-between;color:#94A3B8;font-size:8px}.hd-chart-summary{margin-top:8px;color:#94A3B8;font-size:9px;text-align:right}
        .hd-table-panel{min-height:280px}.hd-table{width:100%;border-collapse:collapse;font-size:10px}.hd-table th{padding:9px 8px;text-align:left;color:#64748B;background:#F8FAFC;border-bottom:1px solid #E2E8F0}.hd-table td{padding:10px 8px;border-bottom:1px solid #EDF2F7;color:#334155}.hd-table tr:last-child td{border-bottom:0}.hd-pill{display:inline-flex;padding:3px 7px;border-radius:999px;font-size:9px;font-weight:800}.hd-pill-green{background:#E7F8EE;color:#10B981}.hd-pill-yellow{background:#FFF4E5;color:#F59E0B}.hd-pill-red{background:#FEECEF;color:#EF4444}.hd-empty{display:grid;place-items:center;min-height:150px;color:#94A3B8;font-size:11px}.hd-activity-list{display:grid}.hd-activity-row{display:grid;grid-template-columns:34px 1fr auto;align-items:center;gap:9px;padding:9px 0;border-bottom:1px solid #EDF2F7}.hd-activity-row:last-child{border-bottom:0}.hd-activity-icon{display:grid;place-items:center;width:32px;height:32px;border-radius:50%;font-size:13px}.hd-activity-title{font-size:10px;font-weight:800}.hd-activity-sub,.hd-activity-time{margin-top:2px;color:#64748B;font-size:8.5px}.hd-activity-time{text-align:right;white-space:nowrap}
        @media(max-width:1200px){.hd-kpi-grid{grid-template-columns:repeat(3,1fr)}.hd-category-grid{grid-template-columns:repeat(2,1fr)}}@media(max-width:700px){.hd-kpi-grid{grid-template-columns:repeat(2,1fr)}.hd-category-grid{grid-template-columns:1fr}.hd-header-title{font-size:24px}.hd-kpi-value{font-size:30px}}@media(max-width:430px){.hd-kpi-grid{grid-template-columns:1fr}}
        @media(max-width:700px){[class*="st-key-hd_category_card_access_control"] .hd-category-card{display:block;min-height:126px}[class*="st-key-hd_category_card_access_control"] .hd-category-status{margin:11px 0 6px}[class*="st-key-hd_category_card_access_control"] .hd-category-unknown{margin:0 0 8px}}
        </style>
        """, unsafe_allow_html=True)

        _hd_kpi_html = ['<div class="hd-kpi-grid">']
        for _label, _value, _note, _icon, _tone, _soft in _hd_kpis:
            _hd_kpi_html.append(f'<div class="hd-kpi-card"><div class="hd-kpi-label">{_label}</div><div class="hd-kpi-value">{_value}</div><div class="hd-kpi-note">{_note}</div><div class="hd-kpi-icon" style="color:{_tone};background:{_soft}">{_icon}</div></div>')
        _hd_kpi_html.append('</div>')
        st.markdown("".join(_hd_kpi_html), unsafe_allow_html=True)

        _hd_main_left = st.container()
        with _hd_main_left:
            st.markdown('<div class="hd-section-title">สรุปประเภทอุปกรณ์</div>', unsafe_allow_html=True)
            def _render_hd_category_card(_cat):
                with st.container(key=f"hd_category_card_{_cat['key']}"):
                    _online_pct = (_cat["online"] / max(_cat["total"], 1)) * 100
                    if _cat["total"] == 0:
                        _unknown_note = '<div class="hd-category-unknown">ยังไม่มีข้อมูลในหมวดนี้</div>'
                    elif _cat["unknown"]:
                        _unknown_note = f'<div class="hd-category-unknown">ไม่ระบุ Status: {_cat["unknown"]} รายการ</div>'
                    else:
                        _unknown_note = '<div class="hd-category-unknown">Status จากข้อมูลจริงครบถ้วน</div>'
                    st.markdown(f'<div class="hd-category-card"><div class="hd-category-head"><div class="hd-category-icon" style="color:{_cat["tone"]};background:{_cat["soft"]}">{_cat["icon"]}</div><div><div class="hd-category-title">{_cat["label"]}</div><div class="hd-category-total">{_cat["total"]} รายการ</div></div></div><div class="hd-category-status"><span>Online<b style="color:#10B981">{_cat["online"]}</b></span><span>Offline<b style="color:#EF4444">{_cat["offline"]}</b></span><span>Maintenance<b style="color:#F59E0B">{_cat["maintenance"]}</b></span></div>{_unknown_note}<div class="hd-progress"><span style="width:{_online_pct:.1f}%;background:{_cat["tone"]}"></span></div></div>', unsafe_allow_html=True)
                    if st.button("ดูรายละเอียด  →", key=f"hd_open_{_cat['key']}", help=f"เปิด {_cat['label']}"):
                        st.session_state.active_nav = _cat["key"]
                        st.rerun()

            _hd_category_cols = st.columns(3, gap="small")
            for _cat_index, _cat in enumerate(_hd_category_stats[:6]):
                with _hd_category_cols[_cat_index % 3]:
                    _render_hd_category_card(_cat)
            if len(_hd_category_stats) > 6:
                _render_hd_category_card(_hd_category_stats[6])

        # Hardware Dashboard intentionally ends after the category command grid.
        # Detail pages remain reachable through each card's action button.
        return

        def _hd_donut(title, items, total, center_label):
            _colors = ["#10B981", "#F59E0B", "#EF4444", "#4F46E5", "#A855F7", "#38BDF8", "#F97316"]
            _cursor = 0.0
            _segments = []
            _legend = []
            for _index, (label, value) in enumerate(items):
                _pct = value / max(total, 1) * 100
                _segments.append(f'{_colors[_index % len(_colors)]} {_cursor:.2f}% {_cursor + _pct:.2f}%')
                _cursor += _pct
                _legend.append(f'<div class="hd-legend-row"><i style="background:{_colors[_index % len(_colors)]}"></i><span>{html.escape(str(label))}</span><b>{value}</b></div>')
            _background = ",".join(_segments) if _segments else "#E2E8F0 0 100%"
            return f'<div class="hd-chart-panel"><div class="hd-chart-title">{title}</div><div class="hd-donut-layout"><div class="hd-donut" style="background:conic-gradient({_background})"><div class="hd-donut-center"><b>{total}</b>{center_label}</div></div><div class="hd-legend">{"".join(_legend)}</div></div></div>'

        with _hd_main_right:
            _hd_donut_cols = st.columns(2, gap="small")
            with _hd_donut_cols[0]:
                _hd_known_status_total = _hd_online + _hd_maintenance + _hd_offline
                st.markdown(_hd_donut("Status Overview", [("Online", _hd_online), ("Maintenance", _hd_maintenance), ("Offline", _hd_offline)], _hd_known_status_total, "มี Status"), unsafe_allow_html=True)
            with _hd_donut_cols[1]:
                _top_locations = sorted(_hd_locations.items(), key=lambda item: item[1], reverse=True)[:4]
                st.markdown(_hd_donut("Asset by Location", _top_locations, sum(value for _, value in _top_locations), "มี Location"), unsafe_allow_html=True)

            def _hd_poly(values, color):
                _max_value = max(max(_hd_total_trend or [0]), max(values or [0]), 1)
                _points = " ".join(f"{25 + index * 105},{110 - (value / _max_value * 88):.1f}" for index, value in enumerate(values))
                return f'<polyline points="{_points}" fill="none" stroke="{color}" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/>'
            _month_labels = "".join(f'<span>{month.strftime("%b %y")}</span>' for month in _hd_months)
            st.markdown(f'<div class="hd-chart-panel hd-trend"><div class="hd-chart-title">Asset Updates by Current Status (6 เดือนล่าสุด)</div><svg viewBox="0 0 575 125" preserveAspectRatio="none"><path d="M25 22H550M25 66H550M25 110H550" stroke="#E8EDF4" stroke-width="1"/>{_hd_poly(_hd_trend["online"], "#3B82F6")}{_hd_poly(_hd_trend["maintenance"], "#F59E0B")}{_hd_poly(_hd_trend["offline"], "#EF4444")}</svg><div class="hd-trend-labels">{_month_labels}</div><div class="hd-chart-summary">อิงวันที่ Modified เท่านั้น · ● Online &nbsp; <span style="color:#F59E0B">● Maintenance</span> &nbsp; <span style="color:#EF4444">● Offline</span></div></div>', unsafe_allow_html=True)

        _hd_bottom = st.columns([1.15, 0.85], gap="medium")
        with _hd_bottom[0]:
            _top_categories = sorted(_hd_category_stats, key=lambda item: item["total"], reverse=True)[:5]
            _rows = []
            for _cat in _top_categories:
                _rows.append(f'<tr><td><b>{_cat["label"]}</b></td><td>{_cat["total"]}</td><td><span class="hd-pill hd-pill-green">{_cat["online"]}</span></td><td><span class="hd-pill hd-pill-red">{_cat["offline"]}</span></td><td><span class="hd-pill hd-pill-yellow">{_cat["maintenance"]}</span></td><td>{_cat["unknown"]}</td></tr>')
            st.markdown(f'<div class="hd-table-panel"><div class="hd-chart-title">Top 5 Assets by Category</div><table class="hd-table"><thead><tr><th>Category</th><th>Total</th><th>Online</th><th>Offline</th><th>Maintenance</th><th>ไม่ระบุ</th></tr></thead><tbody>{"".join(_rows)}</tbody></table></div>', unsafe_allow_html=True)
        with _hd_bottom[1]:
            if _hd_activity:
                _activity_rows = []
                for _modified, _action, _asset, _editor, _state in _hd_activity[:5]:
                    _tone = "#EF4444" if _state == "offline" else ("#F59E0B" if _state == "maintenance" else "#3B82F6")
                    _soft = "#FEECEF" if _state == "offline" else ("#FFF4E5" if _state == "maintenance" else "#EAF3FF")
                    _activity_rows.append(f'<div class="hd-activity-row"><div class="hd-activity-icon" style="color:{_tone};background:{_soft}">●</div><div><div class="hd-activity-title">{html.escape(_action)}</div><div class="hd-activity-sub">{html.escape(_asset)} · {html.escape(_editor)}</div></div><div class="hd-activity-time">{_modified.strftime("%d/%m/%Y")}<br>{_modified.strftime("%H:%M")}</div></div>')
                _activity_body = f'<div class="hd-activity-list">{"".join(_activity_rows)}</div>'
            else:
                _activity_body = '<div class="hd-empty">ยังไม่มีข้อมูลกิจกรรมล่าสุด</div>'
            st.markdown(f'<div class="hd-activity-panel"><div class="hd-chart-title">Recent Activity</div>{_activity_body}</div>', unsafe_allow_html=True)

        if st.session_state.pop("hd_export_requested", False):
            _export_rows = []
            for _key, _label, _, _row, _state in _hd_records:
                _export_item = {"Category": _label, "Status Group": _state}
                _export_item.update({str(column): _row.get(column) for column in _row.index})
                _export_rows.append(_export_item)
            if _export_rows:
                _export_csv = pd.DataFrame(_export_rows).to_csv(index=False).encode("utf-8-sig")
                st.download_button("ดาวน์โหลด Hardware Export", _export_csv, "hardware_assets.csv", "text/csv", key="hd_export_download")

    # -------------------------------------------------------
    # 📊 Overview Dashboard
    # -------------------------------------------------------
    if main_menu == "📊 Overview Dashboard":
        st.markdown("""
        <style>
        /* Dashboard-only visual system. All reusable classes use the db- prefix. */
        .db-shell{color:#172554;padding:0 0 24px}.db-topbar{display:flex;justify-content:flex-end;align-items:center;gap:12px;margin:0 0 14px}
        .db-top-card{height:40px;display:flex;align-items:center;gap:9px;padding:0 14px;border:1px solid #E2E8F0;border-radius:13px;background:#FFF;box-shadow:0 5px 16px rgba(15,23,42,.05);font-size:12px;font-weight:700;color:#334155}.db-top-card svg{width:16px;height:16px;stroke:#334155}
        .db-notify{position:relative;width:40px;padding:0;justify-content:center}.db-notify-badge{position:absolute;right:-3px;top:-6px;display:flex;align-items:center;justify-content:center;width:18px;height:18px;border:2px solid #FFF;border-radius:50%;background:#EF4444;color:#FFF;font-size:9px}
        .db-hero{position:relative;min-height:170px;display:flex;align-items:center;justify-content:space-between;overflow:hidden;padding:26px 38px;margin-bottom:20px;border-radius:28px;background:linear-gradient(118deg,#2563EB 0%,#6366F1 54%,#8B5CF6 100%);box-shadow:0 14px 34px rgba(79,70,229,.20);color:#FFF}
        .db-hero:before,.db-hero:after{content:'';position:absolute;border-radius:50%;border:1px solid rgba(255,255,255,.17)}.db-hero:before{width:560px;height:190px;right:90px;bottom:-125px}.db-hero:after{width:430px;height:150px;right:10px;bottom:-100px}
        .db-hero-copy{position:relative;z-index:2}.db-hero-badge{display:inline-flex;padding:6px 12px;margin-bottom:10px;border-radius:999px;background:rgba(255,255,255,.13);font-size:11px;font-weight:800;letter-spacing:.02em}
        .db-hero-title{font-size:34px;font-weight:800;line-height:1.15;letter-spacing:-.035em}.db-hero-subtitle{margin-top:6px;font-size:15px;color:rgba(255,255,255,.92)}.db-hero-status{display:flex;align-items:center;gap:12px;margin-top:14px;font-size:12px;color:rgba(255,255,255,.88)}.db-live-dot{width:8px;height:8px;border-radius:50%;background:#4ADE80;box-shadow:0 0 0 4px rgba(74,222,128,.14)}
        .db-hero-visual{position:relative;z-index:2;width:330px;height:130px;padding-right:8px}.db-hero-art{width:100%;height:100%;overflow:visible;filter:drop-shadow(0 16px 22px rgba(30,41,59,.18))}
        .db-section-title{margin:0 0 10px;font-size:18px;font-weight:800;color:#172554;letter-spacing:-.015em}.db-overview-grid{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));gap:12px;margin-bottom:14px}
        .db-metric-card{position:relative;min-width:0;height:120px;padding:15px;border:1px solid #E2E8F0;border-radius:18px;background:#FFF;box-shadow:0 6px 18px rgba(15,23,42,.045);overflow:hidden}.db-metric-head{display:flex;align-items:center;gap:10px}.db-metric-icon{display:flex;align-items:center;justify-content:center;flex:0 0 40px;width:40px;height:40px;border-radius:50%}.db-metric-icon svg{width:20px;height:20px;stroke:currentColor}.db-metric-label{min-width:0;font-size:13px;font-weight:700;color:#334155;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.db-metric-value{margin:5px 0 0 50px;font-size:30px;font-weight:800;line-height:1;color:#172554}.db-metric-foot{display:flex;align-items:flex-end;justify-content:space-between;margin:5px 0 0 50px}.db-metric-sub{font-size:11px;color:#64748B}.db-metric-delta{font-size:11px;font-weight:700;color:#94A3B8}.db-metric-delta:empty{display:none}
        .db-main-grid{display:grid;grid-template-columns:minmax(0,3fr) minmax(300px,2fr);gap:14px;margin-top:14px}.db-bottom-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:14px}.db-panel{min-width:0;padding:18px;border:1px solid #E2E8F0;border-radius:18px;background:#FFF;box-shadow:0 7px 22px rgba(15,23,42,.045)}.db-panel-head{display:flex;align-items:center;justify-content:space-between;margin-bottom:12px}.db-panel-title{font-size:16px;font-weight:800;color:#172554}.db-panel-link{font-size:11px;font-weight:700;color:#2563EB}
        .db-action-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:8px}.db-action-card{position:relative;min-height:108px;padding:13px;border:1px solid #E2E8F0;border-radius:16px;background:#FFF;transition:.18s ease}.db-action-card:hover{transform:translateY(-2px);border-color:#C7D2FE;box-shadow:0 10px 22px rgba(79,70,229,.09)}.db-action-icon{display:flex;align-items:center;justify-content:center;width:44px;height:44px;border-radius:50%;background:linear-gradient(135deg,#2563EB,#8B5CF6);color:#FFF}.db-action-icon svg{width:21px;height:21px;stroke:currentColor}.db-action-title{margin-top:8px;padding-right:24px;font-size:14px;font-weight:800;color:#172554}.db-action-desc{margin-top:3px;padding-right:18px;font-size:11px;line-height:1.4;color:#64748B}.db-action-arrow{position:absolute;right:11px;bottom:11px;display:flex;align-items:center;justify-content:center;width:22px;height:22px;border:1px solid #C7D2FE;border-radius:50%;background:#FFF;color:#4F46E5;font-size:12px;font-weight:800}
        .db-health-list,.db-attention-list,.db-activity-list{border:1px solid #E8EDF4;border-radius:14px;overflow:hidden}.db-health-row,.db-list-row{display:flex;align-items:center;gap:11px;min-height:56px;padding:9px 11px;border-bottom:1px solid #E8EDF4}.db-health-row:last-child,.db-list-row:last-child{border-bottom:0}.db-row-icon{display:flex;align-items:center;justify-content:center;flex:0 0 38px;width:38px;height:38px;border:1px solid rgba(255,255,255,.85);border-radius:13px;box-shadow:0 6px 14px rgba(15,23,42,.06)}.db-row-icon svg{width:21px;height:21px;stroke:currentColor}.db-row-copy{min-width:0;flex:1}.db-row-title{font-size:12px;font-weight:800;color:#24324A}.db-row-sub{margin-top:2px;font-size:10px;color:#64748B;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.db-status-online{padding:5px 10px;border-radius:999px;background:#E7F8EE;color:#16A34A;font-size:10px;font-weight:800}.db-status-warning{padding:5px 10px;border-radius:999px;background:#FFF4E5;color:#EA580C;font-size:10px;font-weight:800}.db-time{font-size:10px;color:#64748B;white-space:nowrap}.db-empty-state{padding:18px;text-align:center;color:#94A3B8;font-size:12px}
        [class*="st-key-db_quick_panel"]{padding:18px;border:1px solid #E2E8F0;border-radius:18px;background:#FFF;box-shadow:0 7px 22px rgba(15,23,42,.045)}
        [class*="st-key-db_quick_panel"] [data-testid="stVerticalBlock"]{gap:8px!important}
        [class*="st-key-db_card_action_"] .stButton>button{position:relative!important;display:flex!important;align-items:flex-start!important;justify-content:flex-start!important;width:100%!important;min-height:126px!important;padding:16px 48px 16px 16px!important;border:1px solid #E2E8F0!important;border-radius:16px!important;background:#FFF!important;color:#172554!important;box-shadow:none!important;text-align:left!important;white-space:pre-line!important;transition:transform .18s ease,box-shadow .18s ease,border-color .18s ease!important}
        [class*="st-key-db_card_action_"] .stButton>button:after{content:'›';position:absolute;right:13px;bottom:13px;display:flex;align-items:center;justify-content:center;width:24px;height:24px;border:1px solid #C7D2FE;border-radius:50%;background:#FFF;color:#4F46E5;font-size:15px;font-weight:800}
        [class*="st-key-db_card_action_"] .stButton>button:hover{transform:translateY(-2px)!important;border-color:#A5B4FC!important;background:#FFF!important;color:#172554!important;box-shadow:0 10px 22px rgba(79,70,229,.09)!important}
        [class*="st-key-db_card_action_"] .stButton>button p{width:100%!important;margin:0!important;white-space:pre-line!important;text-align:left!important;font-size:12px!important;line-height:1.55!important;color:#64748B!important;font-weight:500!important}
        [class*="st-key-db_card_action_"] .stButton>button p:first-line{font-size:15px!important;font-weight:800!important;color:#172554!important}
        [data-testid="stAppViewContainer"]{background:radial-gradient(circle at 8% 8%,rgba(56,189,248,.26),transparent 28rem),radial-gradient(circle at 92% 88%,rgba(139,92,246,.23),transparent 31rem),linear-gradient(145deg,#EFF8FF,#EEF2FF 48%,#F5F3FF)!important}section[data-testid="stMain"]{background:transparent!important}section[data-testid="stMain"] [data-testid="stMainBlockContainer"]{padding-top:6px!important}
        @media(max-width:1100px){.db-overview-grid{grid-template-columns:repeat(3,1fr)}.db-main-grid,.db-bottom-grid{grid-template-columns:1fr}.db-hero-visual{opacity:.75}.db-action-grid{grid-template-columns:repeat(3,1fr)}}
        @media(max-width:700px){.db-topbar{justify-content:flex-start}.db-hero{min-height:170px;padding:23px}.db-hero-title{font-size:27px}.db-hero-visual{display:none}.db-overview-grid{grid-template-columns:repeat(2,1fr)}.db-action-grid{grid-template-columns:1fr}.db-main-grid{grid-template-columns:1fr}.db-panel{padding:14px}}
        @media(max-width:430px){.db-overview-grid{grid-template-columns:1fr}.db-top-card span.db-date-label{display:none}}
        </style>
        """, unsafe_allow_html=True)

        # Use a fixed Bangkok offset so Cloud/server UTC does not leak into the UI.
        _bangkok_tz = datetime.timezone(datetime.timedelta(hours=7), name="Asia/Bangkok")
        _dash_now = datetime.datetime.now(_bangkok_tz)
        _thai_months = ["", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
        _dash_date = f"{_dash_now.day} {_thai_months[_dash_now.month]} {_dash_now.year + 543}"
        _dash_time = _dash_now.strftime("%H:%M")

        with st.spinner("กำลังโหลดข้อมูล Dashboard..."):
            df_comp = load_sp_data("Computer Asset")
            df_mon = load_sp_data("Asset Monitor")
            df_prn = load_sp_data("Asset Printer")
            df_ink = load_sp_data(INK_STOCK_LIST)
            df_nas = load_nas_data()

        _comp_count = len(df_comp) if df_comp is not None else 0
        _mon_count = len(df_mon) if df_mon is not None else 0
        _prn_count = len(df_prn) if df_prn is not None else 0
        _nas_count = len(df_nas) if df_nas is not None else 0
        _password_count = int(nav_badges.get("password", 0) or 0)
        _low_ink_count = 0
        if df_ink is not None and not df_ink.empty and "Quantity" in df_ink.columns and "Min_Qty" in df_ink.columns:
            def _db_int(value):
                try:
                    return int(float(value))
                except (TypeError, ValueError):
                    return 0
            _low_ink_count = int(df_ink.apply(lambda row: _db_int(row.get("Quantity", 0)) <= _db_int(row.get("Min_Qty", INK_LOW_THRESHOLD)), axis=1).sum())

        _ad_ok = bool(_ad_agent_enabled() or _ldap_enabled())
        _nas_ok = df_nas is not None
        _sharepoint_ok = any(frame is not None for frame in (df_comp, df_mon, df_prn, df_ink))
        _unassigned_assets = 0
        if df_comp is not None and not df_comp.empty and "field_3" in df_comp.columns:
            _owners = df_comp["field_3"].fillna("").astype(str).str.strip()
            _unassigned_assets = int((_owners == "").sum())

        def _db_svg(kind):
            paths = {
                "computer": '<rect x="3" y="4" width="18" height="13" rx="2"/><path d="M8 21h8M12 17v4"/>',
                "monitor": '<rect x="3" y="4" width="18" height="13" rx="2"/><path d="M7 20h10"/>',
                "printer": '<path d="M6 9V3h12v6"/><rect x="6" y="14" width="12" height="7" rx="1"/><path d="M6 17H4a2 2 0 0 1-2-2v-4a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v4a2 2 0 0 1-2 2h-2"/>',
                "folder": '<path d="M3 6a2 2 0 0 1 2-2h5l2 2h7a2 2 0 0 1 2 2v10a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2Z"/>',
                "lock": '<rect x="4" y="10" width="16" height="11" rx="2"/><path d="M8 10V7a4 4 0 0 1 8 0v3M12 14v3"/>',
                "drop": '<path d="M12 2s7 7.2 7 13a7 7 0 0 1-14 0c0-5.8 7-13 7-13Z"/><path d="M9 16a3 3 0 0 0 3 2"/>',
                "sharepoint": '<circle cx="8" cy="12" r="5"/><circle cx="16.5" cy="7" r="3"/><circle cx="17" cy="17" r="3.5"/><path d="M11.5 9.5 14 8M12 14l2.2 1.4"/>',
                "graph": '<circle cx="12" cy="4" r="2.5"/><circle cx="5" cy="17" r="2.5"/><circle cx="19" cy="17" r="2.5"/><path d="m10.7 6.2-4.4 8.6M13.3 6.2l4.4 8.6M7.5 17h9"/>',
                "server": '<rect x="3" y="3" width="18" height="7" rx="2"/><rect x="3" y="14" width="18" height="7" rx="2"/><path d="M7 6.5h.01M7 17.5h.01M11 6.5h7M11 17.5h7"/>',
                "directory": '<circle cx="9" cy="8" r="3"/><circle cx="17" cy="9" r="2.3"/><path d="M3 20c0-3.4 2.7-6 6-6s6 2.6 6 6M14 15c3.6-.8 7 1.1 7 5"/>',
            }
            return f'<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true">{paths[kind]}</svg>'

        st.markdown(f"""
        <div class="db-shell">
          <div class="db-hero">
            <div class="db-hero-copy">
              <div class="db-hero-badge">IT CONTROL CENTER</div>
              <div class="db-hero-title">DocumentReportUnified</div>
              <div class="db-hero-subtitle">Enterprise IT Management Platform</div>
              <div class="db-hero-status"><span class="db-live-dot"></span><span>ระบบพร้อมใช้งาน</span><span>|</span><span>อัปเดตล่าสุด: {_dash_time}</span></div>
            </div>
            <div class="db-hero-visual">
              <svg class="db-hero-art" viewBox="0 0 330 130" fill="none" aria-hidden="true">
                <g opacity=".42" stroke="#FFF" stroke-width="5"><rect x="8" y="35" width="92" height="62" rx="7"/><path d="M39 111h31M54 97v14M27 78l17-16 14 10 21-24"/></g>
                <g transform="translate(116 8)"><path d="M49 2 91 20v32c0 31-18 51-42 65C25 103 7 83 7 52V20L49 2Z" fill="rgba(255,255,255,.22)" stroke="#FFF" stroke-width="7"/><path d="m29 57 13 13 28-31" stroke="#FFF" stroke-width="8" stroke-linecap="round" stroke-linejoin="round"/></g>
                <path d="M248 90h57a19 19 0 0 0 1-38 28 28 0 0 0-53-7 23 23 0 0 0-5 45Z" fill="rgba(255,255,255,.72)"/>
              </svg>
            </div>
          </div>
          <div class="db-section-title">Overview</div>
        """, unsafe_allow_html=True)

        _db_metrics = [
            (_db_svg("computer"), "Computers", _comp_count, "เครื่อง", "", "#F0EDFF", "#5B5FEF", ""),
            (_db_svg("monitor"), "Monitors", _mon_count, "จอภาพ", "", "#EAF5FF", "#1684E8", ""),
            (_db_svg("printer"), "Printers", _prn_count, "เครื่องพิมพ์", "", "#EAF5FF", "#1684E8", ""),
            (_db_svg("folder"), "NAS Shares", _nas_count, "แชร์", "", "#EAF8EF", "#16A34A", ""),
            (_db_svg("lock"), "Password Records", _password_count, "รายการ", "", "#F1EDFF", "#6D4AFF", ""),
            (_db_svg("drop"), "Ink Stock (Low)", _low_ink_count, "รายการ", "", "#EAF5FF", "#1684E8", ""),
        ]
        _metric_html = ['<div class="db-overview-grid">']
        for _icon, _label, _value, _sub, _delta, _icon_bg, _icon_color, _delta_class in _db_metrics:
            _metric_html.append(f'''<div class="db-metric-card"><div class="db-metric-head"><div class="db-metric-icon" style="background:{_icon_bg};color:{_icon_color}">{_icon}</div><div class="db-metric-label">{_label}</div></div><div class="db-metric-value">{_value:,}</div><div class="db-metric-foot"><span class="db-metric-sub">{_sub}</span><span class="db-metric-delta {_delta_class}">{_delta}</span></div></div>''')
        _metric_html.append('</div>')
        st.markdown("".join(_metric_html), unsafe_allow_html=True)

        _main_cols = st.columns([3, 2], gap="medium")
        with _main_cols[0]:
            with st.container(key="db_quick_panel"):
                st.markdown('<div class="db-panel-head"><div class="db-panel-title">Quick Actions</div></div>', unsafe_allow_html=True)
                _quick_actions = [
                    ("hardware_dashboard", "🖥️", "Hardware", "Computers, Monitors, Printers และอุปกรณ์ทั้งหมด"),
                    ("software_dashboard", "💿", "Software", "License, Office 365, Windows และบัญชีผู้ใช้"),
                    ("permission_dashboard", "🔐", "Permission", "NAS Permission และ AD / Firewall Policy"),
                    ("password", "🔑", "Password", "จัดการรหัสผ่านของระบบและบริการ"),
                    ("vendor_list", "🏢", "Vendor List", "ข้อมูลผู้ขาย ผู้ให้บริการ และช่องทางติดต่อ"),
                    ("ink_stock", "💧", "Ink Stock", "ตรวจสอบสต็อกหมึกพิมพ์และประวัติการใช้งาน"),
                ]
                _action_cols = st.columns(3, gap="small")
                for _idx, (_target, _icon, _title, _description) in enumerate(_quick_actions):
                    with _action_cols[_idx % 3]:
                        if st.button(
                            f"{_icon}  {_title}\n{_description}",
                            key=f"db_card_action_{_idx}_{_target}",
                            use_container_width=True,
                        ):
                            st.session_state.active_nav = _target
                            st.rerun()
        with _main_cols[1]:
            _health_items = [
                (_db_svg("sharepoint"), "SharePoint Online", "เชื่อมต่อ SharePoint และดึงข้อมูลสำเร็จ", _sharepoint_ok, "linear-gradient(135deg,#DDF8EA,#F0FDF7)", "#0F9D68"),
                (_db_svg("graph"), "Microsoft Graph", "เชื่อมต่อ Microsoft Graph API สำเร็จ", _sharepoint_ok, "linear-gradient(135deg,#E0ECFF,#F2F5FF)", "#2563EB"),
                (_db_svg("server"), "NAS Agent", f"เชื่อมต่อ NAS Agent · {_nas_count} shares", _nas_ok, "linear-gradient(135deg,#DCFCE7,#F0FDF4)", "#16A34A"),
                (_db_svg("directory"), "AD Agent", "ตรวจพบการตั้งค่า AD Agent / LDAP" if _ad_ok else "ยังไม่พบการตั้งค่า Agent", _ad_ok, "linear-gradient(135deg,#E0F2FE,#EEF2FF)", "#4F46E5"),
            ]
            _health_html = ['<div class="db-panel"><div class="db-panel-head"><div class="db-panel-title">System Health</div><div class="db-panel-link">สถานะปัจจุบัน</div></div><div class="db-health-list">']
            for _icon, _title, _sub, _ok, _bg, _color in _health_items:
                _status_class = "db-status-online" if _ok else "db-status-warning"
                _status_text = "Online" if _ok else "Warning"
                _health_html.append(f'<div class="db-health-row"><div class="db-row-icon" style="background:{_bg};color:{_color}">{_icon}</div><div class="db-row-copy"><div class="db-row-title">{_title}</div><div class="db-row-sub">{_sub}</div></div><div class="{_status_class}">{_status_text}</div></div>')
            _health_html.append('</div></div>')
            st.markdown("".join(_health_html), unsafe_allow_html=True)

        _attention_items = []
        if _low_ink_count:
            _attention_items.append(("♢", "หมึกพิมพ์ใกล้หมด", f"พบ {_low_ink_count} รายการ", f"{_low_ink_count} รายการ", "#FFECEF", "#F43F5E"))
        if not _nas_ok:
            _attention_items.append(("▤", "NAS Connection Issue", "ไม่สามารถโหลดข้อมูล NAS", "1 รายการ", "#FFF4E5", "#F59E0B"))
        if not _ad_ok:
            _attention_items.append(("!", "AD Agent Warning", "ยังไม่พบการตั้งค่า Agent", "1 รายการ", "#FFF4E5", "#F59E0B"))
        if _unassigned_assets:
            _attention_items.append(("i", "Asset ไม่มีผู้ใช้งาน", f"พบ {_unassigned_assets} รายการ", f"{_unassigned_assets} รายการ", "#EAF3FF", "#2563EB"))

        _bottom_cols = st.columns(2, gap="medium")
        with _bottom_cols[0]:
            _attention_html = ['<div class="db-panel"><div class="db-panel-head"><div class="db-panel-title">Needs Attention</div></div>']
            if _attention_items:
                _attention_html.append('<div class="db-attention-list">')
                for _icon, _title, _sub, _badge, _bg, _color in _attention_items:
                    _attention_html.append(f'<div class="db-list-row"><div class="db-row-icon" style="background:{_bg};color:{_color}">{_icon}</div><div class="db-row-copy"><div class="db-row-title">{_title}</div><div class="db-row-sub">{_sub}</div></div><div class="db-status-warning">{_badge}</div></div>')
                _attention_html.append('</div>')
            else:
                _attention_html.append('<div class="db-empty-state">ไม่พบรายการที่ต้องตรวจสอบ</div>')
            _attention_html.append('</div>')
            st.markdown("".join(_attention_html), unsafe_allow_html=True)
        with _bottom_cols[1]:
            st.markdown('<div class="db-panel"><div class="db-panel-head"><div class="db-panel-title">Recent Activity</div></div><div class="db-empty-state">ยังไม่มีข้อมูลกิจกรรมล่าสุด</div></div></div>', unsafe_allow_html=True)

        # Stop here so the legacy dashboard remains unreachable and other routes are untouched.
        st.stop()

        # ── Legacy operational panels (kept unreachable for safe rollback) ──
        st.markdown('<div class="dash-section"><div class="dash-section-title">Operational Overview</div><div class="dash-section-note">Inventory mix and items requiring action</div></div>', unsafe_allow_html=True)
        _ops_cols = st.columns([1.05, .95], gap="medium")
        _portfolio_total = max(_total_assets, 1)
        _comp_pct = round((_comp_count / _portfolio_total) * 100)
        _mon_pct = round((_mon_count / _portfolio_total) * 100)
        _prn_pct = round((_prn_count / _portfolio_total) * 100)
        with _ops_cols[0]:
            st.markdown(f"""
            <div class="portfolio-card">
                <div class="panel-title">Asset Portfolio</div>
                <div class="panel-sub">Distribution across managed asset types</div>
                <div class="portfolio-row"><div class="portfolio-name">💻 Computers</div><div class="portfolio-track"><div class="portfolio-fill" style="width:{_comp_pct}%;background:linear-gradient(90deg,#2563EB,#60A5FA);"></div></div><div class="portfolio-count">{_comp_count}</div></div>
                <div class="portfolio-row"><div class="portfolio-name">🖥️ Monitors</div><div class="portfolio-track"><div class="portfolio-fill" style="width:{_mon_pct}%;background:linear-gradient(90deg,#0891B2,#22D3EE);"></div></div><div class="portfolio-count">{_mon_count}</div></div>
                <div class="portfolio-row"><div class="portfolio-name">🖨️ Printers</div><div class="portfolio-track"><div class="portfolio-fill" style="width:{_prn_pct}%;background:linear-gradient(90deg,#7C3AED,#A78BFA);"></div></div><div class="portfolio-count">{_prn_count}</div></div>
                <div class="portfolio-row"><div class="portfolio-name">📂 NAS Shares</div><div class="portfolio-track"><div class="portfolio-fill" style="width:{min(_nas_count,100)}%;background:linear-gradient(90deg,#059669,#34D399);"></div></div><div class="portfolio-count">{_nas_count}</div></div>
            </div>
            """, unsafe_allow_html=True)
        with _ops_cols[1]:
            _integration_issues = int(not _firewall_ok) + int(not _ad_ok) + int(not _nas_ok)
            _action_total = _attention_assets + _low_ink_count + _integration_issues
            _action_footer = (
                "✓ All monitored areas are healthy. No immediate action required."
                if _action_total == 0 else f"{_action_total} total items should be reviewed by the IT team."
            )
            st.markdown(f"""
            <div class="action-card">
                <div class="panel-title">Action Center</div>
                <div class="panel-sub">Prioritized issues from live system data</div>
                <div class="action-row"><div class="action-main"><div class="action-icon" style="background:#FFF7ED;color:#D97706;">!</div><div><div class="action-title">Asset attention</div><div class="action-desc">Inactive or repair status</div></div></div><div class="action-badge" style="background:#FFF7ED;color:#C2410C;">{_attention_assets}</div></div>
                <div class="action-row"><div class="action-main"><div class="action-icon" style="background:#FEF2F2;color:#DC2626;">↓</div><div><div class="action-title">Low ink stock</div><div class="action-desc">At or below reorder threshold</div></div></div><div class="action-badge" style="background:#FEF2F2;color:#DC2626;">{_low_ink_count}</div></div>
                <div class="action-row"><div class="action-main"><div class="action-icon" style="background:#EEF2FF;color:#4F46E5;">◆</div><div><div class="action-title">Integration health</div><div class="action-desc">Firewall, AD and NAS connections</div></div></div><div class="action-badge" style="background:#EEF2FF;color:#4F46E5;">{_integration_issues}</div></div>
                <div class="action-footer">{_action_footer}</div>
            </div>
            """, unsafe_allow_html=True)

        # Dashboard is complete. Avoid rendering legacy dashboard sections below.
        st.stop()

        # ── Ink low-stock alert ──
        low_ink = df_ink.iloc[0:0] if df_ink is not None else None
        low_str = "ไม่มีรายการ"
        if not df_ink.empty and "Quantity" in df_ink.columns and "Min_Qty" in df_ink.columns:
            def _toi2(v):
                try: return int(v)
                except: return 0
            low_ink = df_ink[df_ink.apply(
                lambda r: _toi2(r.get("Quantity", 0)) <= _toi2(r.get("Min_Qty", INK_LOW_THRESHOLD)), axis=1
            )]
            if not low_ink.empty:
                low_str = " · ".join(
                    f"{r.get('Title','?')} ({r.get('Color','-')})" for _, r in low_ink.head(3).iterrows()
                )

        st.markdown('<div class="dash-section"><div class="dash-section-title">Overview & Attention</div><div class="dash-section-note">เฉพาะข้อมูลที่ช่วยตัดสินใจและต้องดำเนินการ</div></div>', unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            with st.container():
                st.markdown("**🏢 สัดส่วนอุปกรณ์ตามบริษัท**")
                if not df_comp.empty and 'field_1' in df_comp.columns:
                    cc = df_comp['field_1'].value_counts().reset_index()
                    cc.columns = ['Company', 'Count']
                    _total_co = cc['Company'].nunique()
                    fig = px.pie(cc, values='Count', names='Company', hole=0.50,
                        color_discrete_sequence=['#5b7ff5','#8b6cf7','#a78bfa','#c4b5fd','#ddd6fe','#ede9fe'])
                    fig.update_traces(
                        textposition='inside',
                        textinfo='percent+label',
                        insidetextfont=dict(size=11, color='white'),
                        marker=dict(line=dict(color='white', width=2))
                    )
                    fig.add_annotation(
                        text=f"รวมทั้งหมด<br><b>{_total_co} บริษัท</b>",
                        x=0.5, y=0.5, xref="paper", yref="paper",
                        showarrow=False,
                        font=dict(size=12, family='IBM Plex Sans Thai', color='#374151'),
                        align="center"
                    )
                    fig.update_layout(
                        margin=dict(t=10,b=10,l=10,r=10), height=300,
                        showlegend=True,
                        legend=dict(
                            orientation="v", x=1.02, y=0.5, xanchor="left",
                            font=dict(size=12, family='IBM Plex Sans Thai'),
                            bgcolor='rgba(0,0,0,0)',
                            itemsizing='constant',
                            traceorder='normal',
                        ),
                        paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                        font=dict(family='IBM Plex Sans Thai')
                    )
                    st.plotly_chart(fig, use_container_width=True, config= {'displayModeBar': False})
                    st.markdown('<div style="background:#f8f9fc;border-radius:8px;padding:8px 12px;font-size:0.75rem;color:#94a3b8;">ℹ️ แสดงสัดส่วนจำนวนอุปกรณ์ทั้งหมด จำแนกตามบริษัท</div>', unsafe_allow_html=True)
        with col2:
            _active_rate = round((_active_assets / _total_assets) * 100) if _total_assets else 0
            _health_message = (
                '<div class="attention-ok">✓ ไม่มีรายการเร่งด่วน ระบบอยู่ในสถานะปกติ</div>'
                if _attention_assets == 0 and _low_ink_count == 0 else ''
            )
            st.markdown(f"""
            <div class="attention-card">
                <div class="attention-title">รายการที่ต้องดำเนินการ</div>
                <div class="attention-sub">สรุปเฉพาะสิ่งที่ควรตรวจสอบในวันนี้</div>
                <div class="attention-item">
                    <div><div class="attention-label">อุปกรณ์ผิดปกติ</div><div class="attention-note">Inactive หรืออยู่ระหว่างซ่อม</div></div>
                    <div class="attention-value" style="background:#FFF7ED;color:#C2410C;">{_attention_assets}</div>
                </div>
                <div class="attention-item">
                    <div><div class="attention-label">หมึกต่ำกว่ากำหนด</div><div class="attention-note">{low_str}</div></div>
                    <div class="attention-value" style="background:#FEF2F2;color:#DC2626;">{_low_ink_count}</div>
                </div>
                <div class="attention-item">
                    <div><div class="attention-label">Asset Health</div><div class="attention-note">อัตราสินทรัพย์ที่พร้อมใช้งาน</div></div>
                    <div class="attention-value" style="background:#ECFDF5;color:#047857;">{_active_rate}%</div>
                </div>
                <div class="attention-item">
                    <div><div class="attention-label">NAS Coverage</div><div class="attention-note">Shared folders ที่อยู่ในการติดตาม</div></div>
                    <div class="attention-value" style="background:#EEF2FF;color:#4F46E5;">{_nas_count}</div>
                </div>
                {_health_message}
            </div>
            """, unsafe_allow_html=True)

    
    # -------------------------------------------------------
    # Module landing dashboards
    # -------------------------------------------------------
    elif main_menu == "🖥 Hardware Dashboard":
        _render_hardware_command_center()

    elif main_menu == "💿 Software Dashboard":
        _render_module_hub(
            "Software Dashboard",
            "จัดการบัญชี License และซอฟต์แวร์ที่ใช้งานในองค์กร",
            "💿",
            [
                ("software_group_email", "✉️", "Group E-mail", "บัญชีกลุ่มและอีเมลส่วนกลาง"),
                ("software_office365", "☁️", "Office 365", "Microsoft 365 และ License"),
                ("software_pdf", "📄", "PDF", "Adobe Acrobat และโปรแกรม PDF"),
                ("software_windows", "⊞", "Windows", "Windows License และ Activation"),
                ("software_offboarded", "👤", "พนักงานลาออก", "บัญชีและ License ที่ต้องดำเนินการ"),
            ],
        )

    elif main_menu == "🔐 Permission Dashboard":
        _render_module_hub(
            "Permission Dashboard",
            "ตรวจสอบสิทธิ์การเข้าถึงระบบและนโยบายเครือข่าย",
            "🔐",
            [
                ("user_perm", "📁", "NAS Permission", "ตรวจสอบสิทธิ์การเข้าถึง NAS Shares"),
                ("ad_policy", "🛡️", "AD / Firewall", "ตรวจสอบ Internet Policy จาก AD / Entra ID"),
            ],
        )

    elif main_menu == "💿 Software Module":
        _software_config = {
            "Group E-mail": ("✉️", ["group email", "group e mail", "group mail", "email group", "email"]),
            "Office 365": ("☁️", ["office 365", "office365", "microsoft 365", "m365"]),
            "PDF": ("📄", ["pdf", "adobe", "acrobat"]),
            "Windows": ("⊞", ["windows"]),
            "พนักงานลาออก": ("👤", ["พนักงานลาออก", "ลาออก", "resign", "offboard", "former"]),
        }
        _software_icon, _software_keywords = _software_config.get(_hw_sub_override, ("💿", [_hw_sub_override]))
        _render_password_sheet_module(_hw_sub_override, "ข้อมูล License และบัญชีจากแหล่งข้อมูลเดิม", _software_icon, _software_keywords)

    elif main_menu == "🏢 Vendor List":
        _render_password_sheet_module("Vendor List", "ข้อมูลผู้ขาย ผู้ให้บริการ และช่องทางติดต่อ", "🏢", ["vendor", "supplier", "ผู้ขาย"])

    # -------------------------------------------------------
    # 💻 Hardware Asset
    # -------------------------------------------------------
    elif main_menu == "💻 Hardware Asset":

        st.markdown("""
        <style>

        .asset-hero{
            background:linear-gradient(135deg,#111827 0%, #312e81 100%);
            border-radius:28px;
            padding:32px;
            margin-bottom:24px;
            color:white;
            position:relative;
            overflow:hidden;
            box-shadow:0 20px 50px rgba(99,102,241,.18);
}

        .asset-hero::before{
            content:'';
            position:absolute;
            right:-80px;
            top:-80px;
            width:260px;
            height:260px;
            border-radius:50%;
            background:rgba(255,255,255,.08);
}

        .asset-title{
            font-size:2rem;
            font-weight:800;
            letter-spacing:-1px;
            margin-bottom:6px;
}

        .asset-sub{
            color:rgba(255,255,255,.72);
            font-size:.95rem;
}

        .asset-stat-wrap{
            display:grid;
            grid-template-columns:repeat(4,1fr);
            gap:16px;
            margin-bottom:24px;
}

        .asset-stat{
            background:white;
            border-radius:20px;
            padding:18px;
            border:1px solid #e2e8f0;
            box-shadow:0 10px 30px rgba(15,23,42,.05);
}

        .asset-stat-label{
            font-size:.75rem;
            font-weight:700;
            text-transform:uppercase;
            color:#94a3b8;
            margin-bottom:8px;
}

        .asset-stat-value{
            font-size:2rem;
            font-weight:800;
            color:#0f172a;
            letter-spacing:-1px;
}

        .asset-card{
            background:white;
            border-radius:24px;
            padding:22px;
            border:1px solid #e2e8f0;
            transition:all .22s ease;
            box-shadow:0 8px 24px rgba(15,23,42,.04);
            margin-bottom:20px;
}

        .asset-card:hover{
            transform:translateY(-4px);
            box-shadow:0 20px 40px rgba(99,102,241,.14);
            border-color:#c7d2fe;
}

        .asset-top{
            display:flex;
            justify-content:space-between;
            align-items:flex-start;
            margin-bottom:18px;
}

        .asset-user{
            display:flex;
            gap:14px;
            align-items:center;
}

        .asset-avatar{
            width:54px;
            height:54px;
            border-radius:18px;
            background:linear-gradient(135deg,#6366f1,#8b5cf6);
            display:flex;
            align-items:center;
            justify-content:center;
            color:white;
            font-weight:800;
            font-size:1.1rem;
            flex-shrink:0;
}

        .asset-name{
            font-size:1rem;
            font-weight:700;
            color:#0f172a;
}

        .asset-company{
            font-size:.78rem;
            color:#94a3b8;
            margin-top:2px;
}

        .asset-badge{
            padding:7px 14px;
            border-radius:999px;
            font-size:.76rem;
            font-weight:700;
}

        .badge-active{
            background:#dcfce7;
            color:#15803d;
}

        .badge-inactive{
            background:#fee2e2;
            color:#b91c1c;
}

        .asset-info{
            display:flex;
            flex-direction:column;
            gap:12px;
            margin-bottom:18px;
}

        .asset-row{
            display:flex;
            align-items:center;
            gap:10px;
            color:#475569;
            font-size:.88rem;
}

        .asset-row strong{
            color:#0f172a;
            min-width:90px;
}

        @media(max-width:900px){
            .asset-stat-wrap{
                grid-template-columns:1fr 1fr;
}
}

        </style>
        """, unsafe_allow_html=True)

        sub = _hw_sub_override or "Computer Asset"
        hardware_name = sub.replace("Asset ", "")
        df_hw = load_sp_data(sub)

        if sub == "Computer Asset":
            # UI REVISION: CA-ENTERPRISE-2026-06-22-R3
            # Compact single-row filters, fixed right-aligned pagination, and no column manager.
            # UI OWNER: Computer Asset only. SharePoint and CRUD functions remain unchanged.
            st.markdown("""
            <div class="ca-page"></div>
            <style>
            .stApp:has(.ca-page) [data-testid="stMainBlockContainer"]{background:transparent!important;padding-top:6px!important}
            .stApp:has(.ca-page) [data-testid="stHeader"],.stApp:has(.ca-page) [data-testid="stToolbar"]{display:none!important}
            .stApp:has(.ca-page) [data-testid="stVerticalBlock"]{gap:.58rem}.ca-page{display:none}
            .ca-header{height:102px;box-sizing:border-box;display:flex;align-items:center;gap:16px;padding:17px 22px;margin-bottom:12px;background:#FFF;border:1px solid #E2E8F0;border-radius:20px;box-shadow:0 10px 26px rgba(15,23,42,.055)}
            .ca-header-icon{width:58px;height:58px;flex:0 0 58px;display:grid;place-items:center;color:#4F46E5;background:#F3F5FF;border:1px solid #E0E7FF;border-radius:17px}.ca-header-icon svg{width:33px;height:33px;stroke:currentColor}
            .ca-header h1{margin:0 0 4px!important;color:#0F172A!important;font-size:25px!important;font-weight:800;letter-spacing:-.035em}.ca-header p{margin:0!important;color:#64748B!important;font-size:12.5px!important}
            .ca-metric-grid{display:grid;grid-template-columns:repeat(6,minmax(0,1fr));gap:10px;margin-bottom:12px}.ca-card{height:106px;box-sizing:border-box;position:relative;padding:14px;background:#FFF;border:1px solid #E2E8F0;border-radius:17px;box-shadow:0 7px 18px rgba(15,23,42,.04);overflow:hidden}.ca-card:after{content:"";position:absolute;inset:auto 0 0;height:3px;background:var(--tone)}
            .ca-card-icon{position:absolute;right:12px;top:12px;width:42px;height:42px;display:grid;place-items:center;border-radius:50%;color:var(--tone);background:var(--soft)}.ca-card-icon svg{width:22px;height:22px;stroke:currentColor}.ca-card-label{max-width:calc(100% - 44px);min-height:27px;color:#475569;font-size:11px;line-height:1.25;font-weight:700;overflow:hidden}.ca-card-value{margin-top:1px;color:#0F172A;font-size:25px;line-height:1;font-weight:850;letter-spacing:-.04em}.ca-card-foot{position:absolute;left:14px;right:14px;bottom:11px;display:flex;justify-content:space-between;color:#64748B;font-size:10.5px}.ca-card-foot strong{color:var(--tone)}
            .ca-search-panel{display:none}.ca-search-panel-title{margin:2px 0 7px;padding:0 2px;color:#334155;font-size:12px;font-weight:800}.ca-filter-row{display:none}
            .stApp:has(.ca-page) .stTextInput input,.stApp:has(.ca-page) .stSelectbox div[data-baseweb="select"]>div{height:44px!important;min-height:44px!important;border:1px solid #DDE5EF!important;border-radius:12px!important;background:#FFF!important;font-size:12px!important}.stApp:has(.ca-page) .stButton>button,.stApp:has(.ca-page) .stDownloadButton>button{height:40px;min-height:40px;border-radius:11px;border-color:#E2E8F0;font-size:12px;font-weight:700}.stApp:has(.ca-page) button[kind="primary"]{color:#FFF!important;border:0!important;background:linear-gradient(135deg,#3B82F6,#7C3AED)!important;box-shadow:0 7px 16px rgba(99,102,241,.20)}
            .ca-action-bar{height:52px;display:flex;align-items:center;justify-content:space-between;padding:0 14px;background:#FFF;border:1px solid #E2E8F0;border-bottom:0;border-radius:18px 18px 0 0;color:#334155;font-size:12px;font-weight:750}.ca-action-title{display:flex;align-items:center;gap:8px}.ca-action-title span{width:30px;height:30px;display:grid;place-items:center;border-radius:9px;background:#EEF2FF;color:#4F46E5}
            .ca-table{background:#FFF;border:1px solid #E2E8F0;border-radius:0 0 17px 17px;box-shadow:0 8px 22px rgba(15,23,42,.045);overflow:hidden}.ca-table-scroll{overflow:auto;max-height:560px}.ca-table table{width:100%;min-width:1120px;border-collapse:separate;border-spacing:0;table-layout:fixed;font-size:11px;color:#334155}.ca-table th{position:sticky;top:0;z-index:2;height:44px;padding:0 11px;text-align:left;background:#F8FAFC;color:#475569;font-size:10.5px;font-weight:800;border-bottom:1px solid #E2E8F0;white-space:nowrap}.ca-table td{height:44px;box-sizing:border-box;padding:0 11px;border-bottom:1px solid #EDF2F7;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.ca-table tbody tr:hover td{background:#F8FAFF}
            .ca-status{display:inline-flex;align-items:center;gap:5px;padding:4px 8px;border-radius:999px;font-size:10px;font-weight:800}.ca-status:before{content:"";width:6px;height:6px;border-radius:50%;background:currentColor}.ca-status-online{background:#ECFDF5;color:#059669}.ca-status-offline{background:#FEF2F2;color:#DC2626}.ca-status-nouser{background:#FFF7ED;color:#D97706}.ca-row-actions{display:flex;gap:5px}.ca-row-action{width:25px;height:25px;display:grid;place-items:center;border:1px solid #E2E8F0;border-radius:8px;font-size:12px}.ca-view{color:#2563EB;background:#EFF6FF}.ca-edit{color:#7C3AED;background:#F5F3FF}.ca-delete{color:#EF4444;background:#FEF2F2}
            .ca-table-footer{height:52px;display:flex;align-items:center;justify-content:space-between;padding:0 13px;border-top:1px solid #E2E8F0;color:#64748B;font-size:11px}.ca-pages{display:flex;gap:5px}.ca-pages span{min-width:28px;height:28px;display:grid;place-items:center;border:1px solid #E2E8F0;border-radius:8px}.ca-pages .active{color:#FFF;border-color:#6366F1;background:linear-gradient(135deg,#6366F1,#7C3AED)}
            .stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid){padding:0!important;border:1px solid #E2E8F0!important;border-radius:0 0 17px 17px!important;background:#FFF!important;box-shadow:0 8px 22px rgba(15,23,42,.045)!important;overflow:hidden}
            .stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid)>div{padding:0!important}.ca-native-grid{display:none}.ca-native-head{height:42px;display:flex;align-items:center;padding:0 10px;color:#475569;background:#F8FAFC;font-size:11px;font-weight:800;white-space:nowrap}.ca-native-cell{height:42px;display:flex;align-items:center;padding:0 10px;color:#334155;font-size:11.5px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.ca-native-status{height:42px;display:flex;align-items:center;padding-left:5px}.ca-native-divider{height:1px;margin:0;background:#EDF2F7}.ca-native-footer{height:44px;display:flex;align-items:center;justify-content:space-between;padding:0 14px;color:#64748B;font-size:10.5px;border-top:1px solid #E2E8F0}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="stVerticalBlock"]{gap:0!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="stElementContainer"]{margin:0!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) .stButton{display:flex!important;align-items:center!important;justify-content:center!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) .stButton>button{width:28px!important;min-width:28px!important;max-width:28px!important;height:28px!important;min-height:28px!important;max-height:28px!important;padding:0!important;margin:0 auto!important;border-radius:8px!important;box-shadow:none!important;overflow:hidden!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) .stButton>button p{font-size:0!important;line-height:0!important;margin:0!important;color:transparent!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) .stButton>button span[data-testid="stIconMaterial"]{display:block!important;font-size:16px!important;line-height:1!important;color:currentColor!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="stHorizontalBlock"]{gap:0!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="column"]{border-right:1px solid #EDF2F7}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="column"]:last-child{border-right:0}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="column"]:last-child [data-testid="stHorizontalBlock"] [data-testid="column"]{border-right:0!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="column"]:last-child [data-testid="stHorizontalBlock"] [data-testid="column"]:nth-child(1) button{color:#2563EB!important;background:#EFF6FF!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="column"]:last-child [data-testid="stHorizontalBlock"] [data-testid="column"]:nth-child(2) button{color:#7C3AED!important;background:#F5F3FF!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="column"]:last-child [data-testid="stHorizontalBlock"] [data-testid="column"]:nth-child(3) button{color:#DC2626!important;background:#FEF2F2!important}
            .ca-action-marker{display:none}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(.ca-action-marker) .stButton{display:flex!important;align-items:center!important;justify-content:center!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(.ca-action-marker) .stButton>button{width:32px!important;min-width:32px!important;max-width:32px!important;height:32px!important;min-height:32px!important;max-height:32px!important;padding:0!important;margin:0 auto!important;border:1px solid transparent!important;border-radius:10px!important;box-shadow:none!important;overflow:hidden!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(.ca-action-marker) .stButton>button p{font-size:0!important;line-height:0!important;margin:0!important;color:transparent!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(.ca-action-marker) .stButton>button span[data-testid="stIconMaterial"]{display:block!important;font-size:17px!important;line-height:1!important;color:currentColor!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(.ca-action-view) button{color:#2563EB!important;background:#EFF6FF!important;border-color:#DBEAFE!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(.ca-action-edit) button{color:#7C3AED!important;background:#F5F3FF!important;border-color:#EDE9FE!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(.ca-action-delete) button{color:#DC2626!important;background:#FEF2F2!important;border-color:#FEE2E2!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(.ca-action-marker) button:hover{transform:translateY(-1px)!important;filter:saturate(1.15);box-shadow:0 4px 10px rgba(15,23,42,.08)!important}
            .ca-row-action-marker{display:none}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-marker) .stButton{display:flex!important;align-items:center!important;justify-content:center!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-marker) .stButton>button{width:32px!important;min-width:32px!important;max-width:32px!important;height:32px!important;min-height:32px!important;max-height:32px!important;padding:0!important;margin:0 auto!important;border:1px solid transparent!important;border-radius:10px!important;box-shadow:none!important;overflow:hidden!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-marker) .stButton>button p{font-size:0!important;line-height:0!important;margin:0!important;color:transparent!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-marker) .stButton>button span[data-testid="stIconMaterial"]{display:block!important;font-size:17px!important;line-height:1!important;color:currentColor!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-view) button{color:#2563EB!important;background:#EFF6FF!important;border-color:#DBEAFE!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-edit) button{color:#7C3AED!important;background:#F5F3FF!important;border-color:#EDE9FE!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-delete) button{color:#DC2626!important;background:#FEF2F2!important;border-color:#FEE2E2!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-marker) button:hover{transform:translateY(-1px)!important;filter:saturate(1.15);box-shadow:0 4px 10px rgba(15,23,42,.08)!important}
            .stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-marker) .stButton button{width:32px!important;min-width:32px!important;max-width:32px!important;height:32px!important;min-height:32px!important;max-height:32px!important;padding:0!important;margin:0 auto!important;border-radius:10px!important;box-shadow:none!important;overflow:hidden!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-marker) .stButton button p{font-size:0!important;line-height:0!important;margin:0!important;color:transparent!important}.stApp:has(.ca-page) [data-testid="stVerticalBlock"]:has(> [data-testid="stElementContainer"] .ca-row-action-marker) .stButton button span[data-testid="stIconMaterial"]{font-size:16px!important;line-height:1!important;color:currentColor!important}
            .stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="stButton"]>button{width:30px!important;min-width:30px!important;max-width:30px!important;height:30px!important;min-height:30px!important;max-height:30px!important;padding:0!important;margin:0 auto!important;border-radius:9px!important;box-shadow:none!important;overflow:hidden!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="stButton"]>button p{font-size:0!important;line-height:0!important;margin:0!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="stButton"]>button span[data-testid="stIconMaterial"]{font-size:16px!important;line-height:1!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="column"]:last-child [data-testid="stHorizontalBlock"]{gap:5px!important;justify-content:center!important}.stApp:has(.ca-page) div[data-testid="stVerticalBlockBorderWrapper"]:has(.ca-native-grid) [data-testid="column"]:last-child [data-testid="stHorizontalBlock"]>[data-testid="column"]{width:30px!important;min-width:30px!important;flex:0 0 30px!important}
            .stApp:has(.ca-page) [class*="st-key-ca_view_"] .stButton>button{color:#2563EB!important;background:#EFF6FF!important;border-color:#DBEAFE!important}.stApp:has(.ca-page) [class*="st-key-ca_edit_"] .stButton>button{color:#7C3AED!important;background:#F5F3FF!important;border-color:#EDE9FE!important}.stApp:has(.ca-page) [class*="st-key-ca_delete_"] .stButton>button{color:#DC2626!important;background:#FEF2F2!important;border-color:#FEE2E2!important}
            .stApp:has(.ca-page) [class*="st-key-ca_first"] .stButton>button,.stApp:has(.ca-page) [class*="st-key-ca_prev"] .stButton>button,.stApp:has(.ca-page) [class*="st-key-ca_current"] .stButton>button,.stApp:has(.ca-page) [class*="st-key-ca_next"] .stButton>button,.stApp:has(.ca-page) [class*="st-key-ca_last"] .stButton>button{height:36px!important;min-height:36px!important;border-radius:10px!important}
            .ca-analytics{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-top:12px}.ca-chart-card,.ca-recent-card{height:235px;box-sizing:border-box;padding:15px;background:#FFF;border:1px solid #E2E8F0;border-radius:17px;box-shadow:0 7px 18px rgba(15,23,42,.04);overflow:hidden}.ca-chart-title{margin-bottom:11px;color:#0F172A;font-size:12.5px;font-weight:800}.ca-donut-layout{display:flex;align-items:center;gap:12px;height:164px}.ca-donut{width:112px;height:112px;flex:0 0 112px;border-radius:50%;position:relative;background:conic-gradient(var(--d1) 0 var(--p1),var(--d2) var(--p1) var(--p2),var(--d3) var(--p2) 100%)}.ca-donut:after{content:"";position:absolute;inset:22px;border-radius:50%;background:#FFF}.ca-legend{display:grid;gap:9px;min-width:0}.ca-legend-row{display:grid;grid-template-columns:8px 1fr auto;align-items:center;gap:7px;color:#64748B;font-size:10px}.ca-legend-row i{width:8px;height:8px;border-radius:50%}.ca-legend-row b{color:#334155;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
            .ca-bars{display:grid;gap:12px}.ca-bar-row{display:grid;grid-template-columns:78px 1fr 24px;align-items:center;gap:8px;color:#475569;font-size:10px}.ca-bar-track{height:7px;background:#EEF2FF;border-radius:99px;overflow:hidden}.ca-bar-fill{height:100%;border-radius:99px;background:linear-gradient(90deg,#3B82F6,#8B5CF6)}.ca-bar-row strong{text-align:right}.ca-recent-list{display:grid;gap:11px}.ca-recent-item{display:grid;grid-template-columns:1fr auto;gap:8px;padding-bottom:9px;border-bottom:1px solid #F1F5F9;font-size:10px}.ca-recent-name{color:#334155;font-weight:750}.ca-recent-state,.ca-recent-time{color:#94A3B8}
            @media(max-width:1180px){.ca-metric-grid{grid-template-columns:repeat(3,minmax(0,1fr))}.ca-analytics{grid-template-columns:repeat(2,minmax(0,1fr))}}@media(max-width:700px){.ca-header{height:auto;min-height:110px;padding:18px}.ca-header-icon{width:56px;height:56px;flex-basis:56px}.ca-header h1{font-size:23px!important}.ca-metric-grid,.ca-analytics{grid-template-columns:1fr}.ca-chart-card,.ca-recent-card{height:auto;min-height:245px}.ca-table-footer{height:auto;padding:11px;gap:8px;align-items:flex-start;flex-direction:column}}
            </style>""", unsafe_allow_html=True)

            def _ca_value(row, *names, default="-"):
                for name in names:
                    value = row.get(name, None)
                    if value is not None and str(value).strip() not in ("", "nan", "None", "NaT"):
                        return str(value).strip()
                return default

            def _ca_status(row):
                user = _ca_value(row, "field_3", "User", "Employee", default="")
                raw = _ca_value(row, "Status", "ComputerStatus", default="Active").lower()
                if not user or user == "-": return "No User", "nouser"
                if raw in ("inactive", "offline", "repair", "เสีย", "ซ่อม"): return "Offline", "offline"
                return "Online", "online"

            _ca_esc = lambda value: html.escape(str(value), quote=True)
            _ca_pct = lambda value,total: (float(value)/float(total)*100) if total else 0
            def _ca_os_key(row):
                raw_os = _ca_value(row,"field_10",default="")
                normalized_os = re.sub(r"[^a-z0-9]+","",str(raw_os).casefold())
                if "windows11" in normalized_os or normalized_os.startswith("win11"): return "Windows 11"
                if "windows10" in normalized_os or normalized_os.startswith("win10"): return "Windows 10"
                if "windows7" in normalized_os or normalized_os.startswith("win7"): return "Windows 7"
                return "Other"
            _ca_total = len(df_hw)
            _ca_online = sum(_ca_status(r)[1] == "online" for _,r in df_hw.iterrows())
            _ca_offline = sum(_ca_status(r)[1] == "offline" for _,r in df_hw.iterrows())
            _ca_nouser = sum(_ca_status(r)[1] == "nouser" for _,r in df_hw.iterrows())
            _ca_win11 = sum(_ca_os_key(r)=="Windows 11" for _,r in df_hw.iterrows())
            _ca_win10 = sum(_ca_os_key(r)=="Windows 10" for _,r in df_hw.iterrows())
            _ca_attention = _ca_offline + _ca_nouser
            _ca_monitor = '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="4" width="18" height="13" rx="2"/><path d="M8 21h8M12 17v4"/></svg>'
            _ca_icons = [
                _ca_monitor,
                '<svg viewBox="0 0 24 24" fill="none" stroke-width="2"><circle cx="12" cy="12" r="9"/><path d="m8 12 2.5 2.5L16 9"/></svg>',
                '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.9"><path d="M4 9c4.7-4 11.3-4 16 0M7 12c3-2.5 7-2.5 10 0M10 15c1.2-1 2.8-1 4 0"/><circle cx="12" cy="19" r="1"/></svg>',
                '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.9"><circle cx="12" cy="8" r="3.5"/><path d="M5 20c0-4 3-7 7-7s7 3 7 7"/></svg>',
                '<svg viewBox="0 0 24 24" fill="currentColor" stroke="none"><path d="m3 5 8-1.1v7.6H3V5Zm9-1.2L21 2.5v9h-9V3.8ZM3 12.5h8v7.6L3 19v-6.5Zm9 0h9v9L12 20.2v-7.7Z"/></svg>',
                '<svg viewBox="0 0 24 24" fill="none" stroke-width="2"><circle cx="12" cy="12" r="9"/><path d="M12 7v6M12 17h.01"/></svg>'
            ]
            _ca_icons[5] = _ca_icons[4]
            st.markdown(f'<section class="ca-header"><div class="ca-header-icon">{_ca_monitor}</div><div><h1>Computer Asset</h1><p>จัดการข้อมูลคอมพิวเตอร์ทั้งหมดในองค์กร</p></div></section>',unsafe_allow_html=True)
            _ca_metrics=[("คอมพิวเตอร์ทั้งหมด",_ca_total,"#2563EB","#EFF6FF"),("ใช้งานปกติ",_ca_online,"#10B981","#ECFDF5"),("Offline",_ca_offline,"#F59E0B","#FFF7ED"),("ไม่มีผู้ใช้งาน",_ca_nouser,"#8B5CF6","#F5F3FF"),("Windows 11",_ca_win11,"#38BDF8","#F0F9FF"),("Windows 10",_ca_win10,"#3B82F6","#EFF6FF")]
            st.markdown('<div class="ca-metric-grid">'+''.join(f'<div class="ca-card" style="--tone:{tone};--soft:{soft}"><div class="ca-card-label">{label}</div><div class="ca-card-value">{value:,}</div><div class="ca-card-icon">{_ca_icons[i]}</div><div class="ca-card-foot"><span>เครื่อง</span><strong>{_ca_pct(value,_ca_total):.2f}%</strong></div></div>' for i,(label,value,tone,soft) in enumerate(_ca_metrics))+'</div>',unsafe_allow_html=True)

            _ca_departments=sorted({_ca_value(r,"field_4") for _,r in df_hw.iterrows()})
            st.markdown('<div class="ca-search-panel-title">ค้นหาและกรองข้อมูล</div>',unsafe_allow_html=True)
            f1,f2,f3,f4,f5=st.columns([3.1,1.25,1.35,1,1.2])
            with f1: _ca_search=st.text_input("ค้นหา",placeholder="ค้นหา Computer, User, LoginAccount, Serial",label_visibility="collapsed",key="ca_search")
            with f2: _ca_sf=st.selectbox("Status",["ทั้งหมด","Online","Offline","No User"],label_visibility="collapsed",key="ca_status")
            with f3: _ca_df=st.selectbox("Department",["ทั้งหมด"]+_ca_departments,label_visibility="collapsed",key="ca_department")
            with f4: st.button("⌕ ค้นหา",use_container_width=True,type="primary",key="ca_search_button")
            with f5:
                if st.button("↻ ล้างตัวกรอง",use_container_width=True,key="ca_reset"):
                    for _ca_key in ("ca_search","ca_status","ca_department"):
                        st.session_state.pop(_ca_key,None)
                    st.rerun()
            _ca_filtered=df_hw.copy()
            if _ca_search: _ca_filtered=_ca_filtered[_ca_filtered.astype(str).apply(lambda c:c.str.contains(_ca_search,case=False,na=False)).any(axis=1)]
            if _ca_sf!="ทั้งหมด": _ca_filtered=_ca_filtered[_ca_filtered.apply(lambda r:_ca_status(r)[0]==_ca_sf,axis=1)]
            if _ca_df!="ทั้งหมด": _ca_filtered=_ca_filtered[_ca_filtered.apply(lambda r:_ca_value(r,"field_4")==_ca_df,axis=1)]
            a1,a2,a3,_ca_action_space=st.columns([1.25,.9,1.8,3.2])
            with a1:
                if admin_mode and st.button("＋ เพิ่มคอมพิวเตอร์",use_container_width=True,type="primary",key="ca_add"): add_computer_dialog(sub)
            with a2: st.download_button("⇩ Export",_ca_filtered.to_csv(index=False).encode("utf-8-sig"),"computer_assets.csv","text/csv",use_container_width=True,key="ca_export")
            _ca_column_defs={"computer":"Computer Name","user":"User","login":"LoginAccount","department":"Department","os":"OS","model":"Model","serial":"Serial Number","status":"Status"}
            _ca_visible_columns=list(_ca_column_defs)
            with a3: _ca_sort=st.selectbox("เรียงข้อมูล",["Computer Name A–Z","Computer Name Z–A"],label_visibility="collapsed",key="ca_sort")

            _ca_records=[]
            for idx,row in _ca_filtered.iterrows():
                status,status_class=_ca_status(row)
                _ca_records.append((idx,row,{"computer":_ca_value(row,"field_6"),"user":_ca_value(row,"field_3"),"login":_ca_value(row,"LoginAccount"),"department":_ca_value(row,"field_4"),"os":_ca_value(row,"field_10"),"model":_ca_value(row,"field_7"),"serial":_ca_value(row,"field_8"),"status":status,"status_class":status_class,"seen":_ca_value(row,"LastSeen","Last Seen","Modified")}))
            _ca_records.sort(key=lambda x:x[2]["computer"].lower(),reverse=_ca_sort=="Computer Name Z–A")
            _ca_page_size=10
            _ca_page_count=max(1,(len(_ca_records)+_ca_page_size-1)//_ca_page_size)
            _ca_page=max(1,min(st.session_state.get("ca_page",1),_ca_page_count))
            st.session_state["ca_page"]=_ca_page
            _ca_start=(_ca_page-1)*_ca_page_size
            _ca_slice=_ca_records[_ca_start:_ca_start+_ca_page_size]
            st.markdown('<div class="ca-action-bar"><div class="ca-action-title"><span>▦</span>รายการคอมพิวเตอร์</div><div>Enterprise Data Grid</div></div>',unsafe_allow_html=True)
            _ca_from=(_ca_page-1)*_ca_page_size+1 if _ca_records else 0; _ca_to=min(_ca_page*_ca_page_size,len(_ca_records))
            with st.container(border=True):
                st.markdown('<div class="ca-native-grid"></div>',unsafe_allow_html=True)
                _ca_widths=[1.06,1.0,1.05,.96,.9,1.1,1.0,.78,1.05]
                _ca_head=st.columns(_ca_widths,gap="small")
                for _ca_col,_ca_label in zip(_ca_head,["Computer Name","User","LoginAccount","Department","OS","Model","Serial Number","Status","Action"]):
                    with _ca_col: st.markdown(f'<div class="ca-native-head">{_ca_label}</div>',unsafe_allow_html=True)
                if not _ca_slice:
                    st.info("ไม่พบข้อมูลตามเงื่อนไข")
                for _ca_idx,_ca_row,_ca_data in _ca_slice:
                    _ca_cols=st.columns(_ca_widths,gap="small",vertical_alignment="center")
                    _ca_values=[_ca_data["computer"],_ca_data["user"],_ca_data["login"],_ca_data["department"],_ca_data["os"],_ca_data["model"],_ca_data["serial"]]
                    for _ca_col,_ca_value_text in zip(_ca_cols[:7],_ca_values):
                        with _ca_col: st.markdown(f'<div class="ca-native-cell" title="{_ca_esc(_ca_value_text)}">{_ca_esc(_ca_value_text)}</div>',unsafe_allow_html=True)
                    with _ca_cols[7]:
                        st.markdown(f'<div class="ca-native-status"><span class="ca-status ca-status-{_ca_data["status_class"]}">{_ca_esc(_ca_data["status"])}</span></div>',unsafe_allow_html=True)
                    with _ca_cols[8]:
                        _ca_b1,_ca_b2,_ca_b3=st.columns(3,gap="small")
                        with _ca_b1:
                            st.markdown('<span class="ca-row-action-marker ca-row-action-view"></span>',unsafe_allow_html=True)
                            if st.button(" ",icon=":material/visibility:",key=f"ca_view_{_ca_idx}",help="ดูรายละเอียด"): show_pop_computer(_ca_row.to_dict())
                        with _ca_b2:
                            st.markdown('<span class="ca-row-action-marker ca-row-action-edit"></span>',unsafe_allow_html=True)
                            if admin_mode and st.button(" ",icon=":material/edit:",key=f"ca_edit_{_ca_idx}",help="แก้ไข"): edit_computer_dialog(_ca_row.to_dict(),sub)
                        with _ca_b3:
                            st.markdown('<span class="ca-row-action-marker ca-row-action-delete"></span>',unsafe_allow_html=True)
                            if admin_mode and st.button(" ",icon=":material/delete:",key=f"ca_delete_{_ca_idx}",help="ลบ"): delete_computer_dialog(_ca_row.to_dict(),sub)
                    st.markdown('<div class="ca-native-divider"></div>',unsafe_allow_html=True)
                st.markdown(f'<div class="ca-native-footer"><span>แสดง {_ca_from} ถึง {_ca_to} จาก {len(_ca_records)} รายการ</span><span>หน้า {_ca_page} / {_ca_page_count}</span></div>',unsafe_allow_html=True)

            _ca_nav=st.columns([7,.52,.52,.52,.52,.52])
            with _ca_nav[1]:
                if st.button("«",use_container_width=True,key="ca_first",disabled=_ca_page<=1): st.session_state["ca_page"]=1; st.rerun()
            with _ca_nav[2]:
                if st.button("‹",use_container_width=True,key="ca_prev",disabled=_ca_page<=1): st.session_state["ca_page"]=_ca_page-1; st.rerun()
            with _ca_nav[3]:
                st.button(str(_ca_page),use_container_width=True,type="primary",key="ca_current",disabled=True)
            with _ca_nav[4]:
                if st.button("›",use_container_width=True,key="ca_next",disabled=_ca_page>=_ca_page_count): st.session_state["ca_page"]=_ca_page+1; st.rerun()
            with _ca_nav[5]:
                if st.button("»",use_container_width=True,key="ca_last",disabled=_ca_page>=_ca_page_count): st.session_state["ca_page"]=_ca_page_count; st.rerun()

            _ca_types={"Desktop":0,"All-in-One":0,"Notebook":0}; _ca_windows={"Windows 11":0,"Windows 10":0,"Windows 7":0}; _ca_depts={}
            for _,r in df_hw.iterrows():
                model=_ca_value(r,"field_7","Model",default="").lower(); kind="Notebook" if any(x in model for x in ("notebook","laptop","thinkpad","latitude")) else ("All-in-One" if any(x in model for x in ("all-in-one","aio")) else "Desktop"); _ca_types[kind]+=1
                os_key=_ca_os_key(r)
                if os_key in _ca_windows: _ca_windows[os_key]+=1
                dept=_ca_value(r,"field_4",default="ไม่ระบุ"); _ca_depts[dept]=_ca_depts.get(dept,0)+1
            def _ca_donut(title,data,colors):
                total=max(sum(data.values()),1); values=list(data.values()); p1=values[0]/total*100; p2=(values[0]+values[1])/total*100; legend=''.join(f'<div class="ca-legend-row"><i style="background:{colors[i]}"></i><b>{_ca_esc(k)}</b><span>{v}</span></div>' for i,(k,v) in enumerate(data.items())); return f'<div class="ca-chart-card"><div class="ca-chart-title">{title}</div><div class="ca-donut-layout"><div class="ca-donut" style="--d1:{colors[0]};--d2:{colors[1]};--d3:{colors[2]};--p1:{p1:.2f}%;--p2:{p2:.2f}%"></div><div class="ca-legend">{legend}</div></div></div>'
            _ca_top=sorted(_ca_depts.items(),key=lambda x:x[1],reverse=True)[:5]; _ca_max=max([v for _,v in _ca_top] or [1]); _ca_bars=''.join(f'<div class="ca-bar-row"><span>{_ca_esc(k)}</span><div class="ca-bar-track"><div class="ca-bar-fill" style="width:{v/_ca_max*100:.1f}%"></div></div><strong>{v}</strong></div>' for k,v in _ca_top); _ca_recent=''.join(f'<div class="ca-recent-item"><div><div class="ca-recent-name">{_ca_esc(d["computer"])}</div><div class="ca-recent-state">{_ca_esc(d["status"])}</div></div><div class="ca-recent-time">{_ca_esc(d["seen"])}</div></div>' for _,_,d in sorted(_ca_records,key=lambda x:x[2]["seen"],reverse=True)[:5])
            st.markdown('<div class="ca-analytics">'+_ca_donut("ประเภทเครื่อง",_ca_types,["#4F46E5","#38BDF8","#A855F7"])+_ca_donut("Windows Version",_ca_windows,["#2563EB","#3B82F6","#22C1C3"])+f'<div class="ca-chart-card"><div class="ca-chart-title">Top 5 Department</div><div class="ca-bars">{_ca_bars}</div></div><div class="ca-recent-card"><div class="ca-chart-title">◷ อัปเดตล่าสุด</div><div class="ca-recent-list">{_ca_recent}</div></div></div>',unsafe_allow_html=True)
            st.stop()

        st.markdown(f"""
        <div class="asset-hero">
            <div class="asset-title">💻 {hardware_name}</div>
            <div class="asset-sub">
                ระบบจัดการ{hardware_name}และทรัพย์สิน IT ทั้งหมด
            </div>
        </div>
        """, unsafe_allow_html=True)

        total_assets = len(df_hw)
        active_assets = len(df_hw[df_hw["Status"] == "Active"]) if not df_hw.empty else 0
        inactive_assets = len(df_hw[df_hw["Status"] == "Inactive"]) if not df_hw.empty else 0
        repair_assets = len(df_hw[df_hw["Status"] == "Repair"]) if not df_hw.empty else 0

        # ใช้ Streamlit metric แทน HTML เพื่อป้องกัน HTML render เป็น text
        m1, m2, m3, m4 = st.columns(4)

        with m1:
            st.metric("TOTAL ASSETS", total_assets)

        with m2:
            st.metric("ACTIVE", active_assets)

        with m3:
            st.metric("INACTIVE", inactive_assets)

        with m4:
            st.metric("REPAIR", repair_assets)

        col_search, col_add = st.columns([0.82, 0.18])

        with col_search:
            search = st.text_input(
                "",
                placeholder="🔍 ค้นหาชื่อพนักงาน, Hostname, Model, S/N...",
                label_visibility="collapsed"
            )

        with col_add:
            if admin_mode:
                if st.button("➕ เพิ่มคอมพิวเตอร์", use_container_width=True, type="primary"):
                    add_computer_dialog(sub)

        if search and not df_hw.empty:
            df_hw = df_hw[df_hw.astype(str).apply(
                lambda x: x.str.contains(search, case=False)
            ).any(axis=1)]

        cols = st.columns(3)

        for i, (idx, row) in enumerate(df_hw.iterrows()):

            with cols[i % 3]:

                name = row.get("field_3", "Unknown")
                initials = "".join([x[0] for x in name.split()[:2]]).upper()

                status = row.get("Status", "Active")
                badge_class = "badge-active" if status == "Active" else "badge-inactive"

                with st.container(border=True):
                    st.markdown(f"### 👤 {name}")
                    st.caption(f"🏢 {row.get('field_1','-')}  |  Status: {status}")
                    st.write(f"💻 Hostname: {row.get('field_6','-')}")
                    st.write(f"🏷️ Model: {row.get('field_7','-')}")
                    st.write(f"💾 RAM: {row.get('field_13','-')}")
                    st.write(f"🔢 Serial: {row.get('field_8','-')}")

                if admin_mode:
                    b1, b2 = st.columns(2)

                    with b1:
                        if st.button("🔍 ดูข้อมูล", key=f"view_{idx}", use_container_width=True):
                            show_pop_computer(row.to_dict())

                    with b2:
                        if st.button("✏️ แก้ไข", key=f"edit_{idx}", use_container_width=True):
                            edit_computer_dialog(row.to_dict(), sub)
                else:
                    st.caption("🔒 ดูรายละเอียดเพิ่มเติมได้เฉพาะผู้ดูแลระบบ")



    # -------------------------------------------------------
    # 🌐 AD / Firewall Policy
    # -------------------------------------------------------
    elif main_menu == "🌐 AD / Firewall Policy":
        # UI OWNER: AD / Firewall Policy only.  All selectors are scoped by
        # .adp-page-marker so this theme cannot leak into other Streamlit pages.
        st.markdown("""
        <div class="adp-page-marker"></div>
        <style>
        .stApp:has(.adp-page-marker) [data-testid="stMainBlockContainer"] {
            background:#F7F9FD;
            padding-top:0!important;
        }
        .stApp:has(.adp-page-marker) [data-testid="stHeader"],
        .stApp:has(.adp-page-marker) [data-testid="stToolbar"] {
            display:none!important;
        }
        .adp-hero {
            height:132px; min-height:132px; box-sizing:border-box; padding:22px 26px;
            margin:0 0 14px; border-radius:18px; color:#fff;
            display:flex; align-items:center; gap:20px; overflow:hidden; position:relative;
            background:
              radial-gradient(circle at 88% 18%,rgba(56,189,248,.38),transparent 21%),
              radial-gradient(circle at 78% 120%,rgba(139,92,246,.72),transparent 36%),
              linear-gradient(128deg,#1E40AF 0%,#3949C6 50%,#6D3DEB 100%);
            border:1px solid rgba(255,255,255,.24);
            box-shadow:0 16px 36px rgba(49,46,129,.18);
        }
        .adp-hero:after {
            content:""; position:absolute; width:210px; height:210px; right:35px; top:-84px;
            border:1px solid rgba(255,255,255,.14); border-radius:50%;
            box-shadow:0 0 0 26px rgba(255,255,255,.035),0 0 0 58px rgba(255,255,255,.025);
        }
        .adp-hero-icon {
            width:64px; height:64px; flex:0 0 64px; border-radius:16px;
            display:grid; place-items:center; font-size:30px;
            background:linear-gradient(145deg,rgba(255,255,255,.26),rgba(255,255,255,.12));
            border:1px solid rgba(255,255,255,.18); box-shadow:inset 0 1px 0 rgba(255,255,255,.22);
        }
        .adp-hero-icon svg{width:38px;height:38px;display:block;filter:drop-shadow(0 4px 8px rgba(15,23,42,.18))}
        .adp-hero-copy {position:relative; z-index:1; min-width:0}
        .adp-hero h1 {font-size:28px!important; line-height:1.15; margin:0 0 9px!important; color:#fff!important; letter-spacing:-.025em}
        .adp-hero h1 a,.adp-hero h1 svg{display:none!important}
        .adp-hero p,.stMarkdown .adp-hero p {font-size:13px!important; line-height:1.5; margin:0!important; color:rgba(255,255,255,.88)!important}
        .adp-hero-art {position:absolute;z-index:1;right:34px;top:7px;width:230px;height:118px;color:#BAE6FD;opacity:.94}
        .adp-hero-art svg{display:block;width:100%;height:100%;overflow:visible;filter:drop-shadow(0 10px 18px rgba(30,64,175,.20))}
        .adp-globe {position:absolute;right:76px;top:0;width:80px;height:80px;border:3px solid rgba(125,211,252,.58);border-radius:50%;background:linear-gradient(90deg,transparent 46%,rgba(125,211,252,.42) 47%,rgba(125,211,252,.42) 53%,transparent 54%)}
        .adp-globe:before,.adp-globe:after {content:"";position:absolute;left:8px;right:8px;border-top:3px solid rgba(125,211,252,.55)}
        .adp-globe:before{top:27px}.adp-globe:after{top:55px}
        .adp-wall {position:absolute;right:0;bottom:4px;display:grid;grid-template-columns:repeat(3,30px);gap:4px;transform:skewY(-2deg)}
        .adp-wall i{height:22px;border-radius:4px;background:linear-gradient(145deg,#7DD3FC,#60A5FA);box-shadow:0 4px 10px rgba(30,64,175,.24)}
        .adp-lock {position:absolute;right:18px;bottom:0;font-size:35px;filter:drop-shadow(0 5px 8px rgba(30,64,175,.24))}
        .adp-info-banner {
            min-height:46px; box-sizing:border-box; display:flex; align-items:center; gap:10px;
            margin:0 0 10px; padding:8px 14px; border-radius:13px;
            color:#405174; font-size:12.5px; line-height:1.45;
            background:linear-gradient(90deg,#F1F5FF 0%,#F8FAFF 100%); border:1px solid #D8E1F5;
            box-shadow:0 2px 8px rgba(79,70,229,.025);
        }
        .adp-info-icon {width:18px;height:18px;color:#4F46E5;flex:0 0 18px;display:grid;place-items:center}
        .adp-info-icon svg{width:17px;height:17px;display:block;stroke:currentColor}
        .adp-info-banner code {font-size:11px; color:#4338CA; background:#E0E7FF; padding:2px 6px; border-radius:6px}
        .adp-stat-grid {display:grid; grid-template-columns:repeat(auto-fit,minmax(210px,1fr)); gap:12px; margin:14px 0 8px}
        .adp-stat-card {
            height:94px; box-sizing:border-box; display:flex; align-items:center; gap:12px;
            padding:14px 16px; min-width:0; background:#FFFFFF; border:1px solid #E2E8F0;
            border-radius:14px; box-shadow:0 7px 18px rgba(30,41,59,.055);
        }
        .adp-stat-icon {width:38px;height:38px;flex:0 0 38px;border-radius:12px;display:grid;place-items:center;font-size:18px;background:linear-gradient(145deg,#EEF2FF,#E8EDFF);color:#5B5FF0;border:1px solid #E1E7FF}
        .adp-stat-icon svg{width:21px;height:21px;display:block;stroke:currentColor}
        .adp-stat-grid .adp-stat-card:nth-child(1) .adp-stat-icon{background:linear-gradient(145deg,#F3F0FF,#ECE8FF);color:#6D5DF8;border-color:#E5DEFF}
        .adp-stat-grid .adp-stat-card:nth-child(2) .adp-stat-icon{background:linear-gradient(145deg,#EFF6FF,#E7F0FF);color:#2563EB;border-color:#D9E8FF}
        .adp-stat-grid .adp-stat-card:nth-child(3) .adp-stat-icon{background:linear-gradient(145deg,#F5F1FF,#EEE9FF);color:#7C3AED;border-color:#E7DEFF}
        .adp-stat-grid .adp-stat-card:nth-child(4) .adp-stat-icon{background:linear-gradient(145deg,#ECFDF5,#E5F8EF);color:#059669;border-color:#D4F1E3}
        .adp-stat-copy {min-width:0;overflow:hidden}
        .adp-stat-label {font-size:10.5px;line-height:1.2;font-weight:750;letter-spacing:.06em;text-transform:uppercase;color:#64748B;margin-bottom:6px}
        .adp-stat-value {font-size:18px;line-height:1.25;font-weight:720;color:#16213A;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
        .adp-summary-grid {display:grid;grid-template-columns:1fr 1fr;gap:12px;margin:10px 0 12px}
        .adp-summary-card {height:90px;box-sizing:border-box;padding:14px 18px;background:#fff;border:1px solid #E2E8F0;border-radius:14px;box-shadow:0 6px 16px rgba(30,41,59,.04)}
        .adp-summary-card:first-child {background:linear-gradient(135deg,#FFFFFF 0%,#F7F5FF 100%);border-color:#DCD7FE}
        .adp-summary-card:last-child {background:linear-gradient(135deg,#FFFFFF 0%,#F0F9FF 100%);border-color:#D4EAF7}
        .adp-summary-card{position:relative;overflow:hidden}
        .adp-summary-icon{position:absolute;right:17px;top:17px;width:40px;height:40px;display:grid;place-items:center;border-radius:13px;background:#F0EDFF;color:#6366F1}
        .adp-summary-card:last-child .adp-summary-icon{background:#EAF7FF;color:#0284C7}
        .adp-summary-icon svg{width:23px;height:23px;stroke:currentColor}
        .adp-summary-label {font-size:10.5px;font-weight:750;letter-spacing:.06em;color:#4F46E5;text-transform:uppercase}
        .adp-summary-value {font-size:22px;line-height:1.15;font-weight:800;color:#1E3A8A;margin-top:6px}
        .adp-summary-note {font-size:10.5px;color:#64748B;margin-top:1px}
        .adp-table-title {display:flex;align-items:center;gap:8px;font-size:13px;font-weight:750;color:#1E293B;margin:0 0 9px}
        .adp-table-title .adp-stat-icon svg{width:14px;height:14px}
        .adp-field-label {font-size:11px;font-weight:700;color:#334155;margin:0 0 5px}
        .adp-table-wrap {background:#fff;border:1px solid #DEE5EF;border-radius:15px;padding:13px 14px 14px;box-shadow:0 8px 22px rgba(30,41,59,.05);margin-top:2px;overflow:hidden}
        .adp-policy-table {width:100%;border-collapse:separate;border-spacing:0;table-layout:fixed;border:1px solid #E2E8F0;border-radius:10px;overflow:hidden;font-size:10.5px;color:#334155}
        .adp-policy-table th {height:34px;box-sizing:border-box;padding:6px 8px;text-align:left;background:#F5F7FB;color:#52617A;font-size:10px;font-weight:800;border-right:1px solid #E2E8F0;border-bottom:1px solid #DCE3ED;white-space:nowrap}
        .adp-policy-table td {height:36px;box-sizing:border-box;padding:6px 8px;border-right:1px solid #EDF1F5;border-bottom:1px solid #E8EDF3;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;vertical-align:middle}
        .adp-policy-table tr:last-child td{border-bottom:0}.adp-policy-table th:last-child,.adp-policy-table td:last-child{border-right:0}
        .adp-source-pill {display:inline-flex;align-items:center;justify-content:center;min-width:26px;height:18px;padding:0 6px;border-radius:999px;background:#EDE9FE;color:#4F46E5;font-size:9.5px;font-weight:800}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab-list"] {gap:8px;padding:0;background:transparent;border-radius:13px;width:min(600px,100%);max-width:100%;margin-bottom:10px}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"] {height:58px;flex:1;position:relative;justify-content:flex-start;padding:0 12px 0 46px;border:1px solid #DDE4EE;border-radius:12px;background:#FFFFFF;color:#64748B;font-size:12px;font-weight:750;box-shadow:0 4px 12px rgba(30,41,59,.035)}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"] p {font-size:12px!important;font-weight:750!important;color:inherit!important;margin:0 0 15px!important}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]:before {content:"";position:absolute;left:15px;top:18px;width:20px;height:20px;background:currentColor;opacity:.9;mask-repeat:no-repeat;mask-position:center;mask-size:contain;-webkit-mask-repeat:no-repeat;-webkit-mask-position:center;-webkit-mask-size:contain}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]:nth-child(1):before{mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Ccircle cx='9' cy='8' r='3' fill='black'/%3E%3Cpath d='M3 19c0-3.3 2.7-6 6-6s6 2.7 6 6v1H3zM16 8h5v2h-5zM17.5 5.5h2v7h-2z' fill='black'/%3E%3C/svg%3E");-webkit-mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Ccircle cx='9' cy='8' r='3' fill='black'/%3E%3Cpath d='M3 19c0-3.3 2.7-6 6-6s6 2.7 6 6v1H3zM16 8h5v2h-5zM17.5 5.5h2v7h-2z' fill='black'/%3E%3C/svg%3E")}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]:nth-child(2):before{mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M10.5 3a7.5 7.5 0 1 0 4.7 13.3L20 21l1-1-4.7-4.8A7.5 7.5 0 0 0 10.5 3zm0 2a5.5 5.5 0 1 1 0 11 5.5 5.5 0 0 1 0-11z' fill='black'/%3E%3C/svg%3E");-webkit-mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M10.5 3a7.5 7.5 0 1 0 4.7 13.3L20 21l1-1-4.7-4.8A7.5 7.5 0 0 0 10.5 3zm0 2a5.5 5.5 0 1 1 0 11 5.5 5.5 0 0 1 0-11z' fill='black'/%3E%3C/svg%3E")}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]:nth-child(3):before{mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M4 3h6v6H4V3zm10 0h6v6h-6V3zM4 15h6v6H4v-6zm10 0h6v6h-6v-6zM10 5h4v2h-4V5zm-3 4h2v6H7V9zm8 0h2v6h-2V9zm-5 9h4v2h-4v-2z' fill='black'/%3E%3C/svg%3E");-webkit-mask-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24'%3E%3Cpath d='M4 3h6v6H4V3zm10 0h6v6h-6V3zM4 15h6v6H4v-6zm10 0h6v6h-6v-6zM10 5h4v2h-4V5zm-3 4h2v6H7V9zm8 0h2v6h-2V9zm-5 9h4v2h-4v-2z' fill='black'/%3E%3C/svg%3E")}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]:after {position:absolute;left:46px;top:34px;color:#94A3B8;font-size:9.5px;font-weight:550;white-space:nowrap}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]:nth-child(1):after{content:"ค้นหานโยบายของผู้ใช้"}.stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]:nth-child(2):after{content:"ค้นหาผู้ใช้ตาม Policy"}.stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]:nth-child(3):after{content:"ตรวจสอบการแม็ป Policy"}
        .stApp:has(.adp-page-marker) .stTabs [aria-selected="true"] {background:linear-gradient(180deg,#FFFFFF 0%,#F6F5FF 100%)!important;color:#4F46E5!important;border-color:transparent!important;border-radius:12px!important;box-shadow:0 7px 16px rgba(79,70,229,.11)!important;z-index:2}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab-panel"] {padding-top:0!important}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab-highlight"],
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab-border"] {display:none!important}
        .stApp:has(.adp-page-marker) .stTextInput input,
        .stApp:has(.adp-page-marker) .stSelectbox div[data-baseweb="select"]>div {height:42px!important;min-height:42px!important;border-radius:10px!important;border-color:#DCE3ED!important;font-size:12px!important}
        .stApp:has(.adp-page-marker) .stButton>button {height:42px!important;min-height:42px!important;border-radius:10px!important;font-size:12px!important}
        .stApp:has(.adp-page-marker) .stButton>button p {color:inherit!important;font-size:inherit!important}
        .stApp:has(.adp-page-marker) div[data-testid="stVerticalBlockBorderWrapper"] {border:1px solid #DCE3ED!important;border-radius:14px!important;background:#FFFFFF!important;box-shadow:0 7px 18px rgba(30,41,59,.045)!important}
        .stApp:has(.adp-page-marker) div[data-testid="stVerticalBlockBorderWrapper"]>div {padding:12px 14px!important}
        .adp-search-panel-marker{display:none}
        .stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab-panel"] div[data-testid="stVerticalBlockBorderWrapper"]:has(.adp-search-panel-marker){position:relative;z-index:1;margin-top:0;border:0!important;border-radius:14px!important;background:#FFFFFF!important;box-shadow:0 8px 22px rgba(30,41,59,.06)!important}
        .adp-table-footer-marker{display:flex;align-items:center;gap:7px;height:30px;margin:5px 2px 0;padding:0 4px;color:#64748B;font-size:10.5px}
        .adp-table-footer-marker svg{width:15px;height:15px;stroke:#7EA5DF}
        .stApp:has(.adp-page-marker) [data-testid="stDataFrame"] {border:1px solid #E2E8F0!important;border-radius:12px!important;box-shadow:none!important}
        .stApp:has(.adp-page-marker) [data-testid="stExpander"] {background:#FFFFFF!important;border:1px solid #DEE5EF!important;border-radius:12px!important;box-shadow:0 4px 12px rgba(30,41,59,.035);overflow:hidden;margin-top:3px}
        .stApp:has(.adp-page-marker) [data-testid="stExpander"] summary {min-height:40px;padding-top:2px;padding-bottom:2px;font-size:12px;font-weight:700;color:#334155;background:#FFFFFF!important}
        .stApp:has(.adp-page-marker) [data-testid="stExpander"] summary:hover{background:#F8FAFF!important}
        .stApp:has(.adp-page-marker) [data-testid="stExpander"] summary p {font-size:12px!important;color:#334155!important}
        .stApp:has(.adp-page-marker) [data-testid="stVerticalBlock"] {gap:.58rem}
        @media(max-width:900px){.adp-stat-grid{grid-template-columns:repeat(2,minmax(0,1fr))}.adp-hero{height:132px;min-height:132px}.adp-hero:after,.adp-hero-art{display:none}.adp-policy-table th:nth-child(5),.adp-policy-table td:nth-child(5),.adp-policy-table th:nth-child(6),.adp-policy-table td:nth-child(6){display:none}}
        @media(max-width:640px){.adp-hero{padding:22px 18px;gap:14px}.adp-hero-icon{width:54px;height:54px;flex-basis:54px;font-size:27px}.adp-hero h1{font-size:28px!important}.adp-hero p{font-size:14px}.adp-stat-grid,.adp-summary-grid{grid-template-columns:1fr}.adp-stat-card{height:108px}.stApp:has(.adp-page-marker) .stTabs [data-baseweb="tab"]{height:50px;padding:0 9px;font-size:12px}}
        </style>
        <section class="adp-hero">
          <div class="adp-hero-icon">
            <svg viewBox="0 0 48 48" aria-hidden="true"><defs><linearGradient id="adpShield" x1="0" y1="0" x2="1" y2="1"><stop stop-color="#E0F2FE"/><stop offset="1" stop-color="#93C5FD"/></linearGradient></defs><path d="M24 4 40 11v11c0 10.2-6.5 18.3-16 22C14.5 40.3 8 32.2 8 22V11L24 4Z" fill="url(#adpShield)" stroke="#BFDBFE" stroke-width="2"/><path d="m17.2 24.2 4.4 4.4 9.6-10" fill="none" stroke="#1D4ED8" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/></svg>
          </div>
          <div class="adp-hero-copy">
            <h1>AD / Firewall Policy</h1>
            <p>ตรวจสอบ Internet Policy ที่ผู้ใช้หรือ Policy Group ได้รับจาก AD / Entra ID Group</p>
          </div>
          <div class="adp-hero-art" aria-hidden="true">
            <svg viewBox="0 0 260 130"><defs><linearGradient id="adpBrick" x1="0" y1="0" x2="1" y2="1"><stop stop-color="#7DD3FC"/><stop offset="1" stop-color="#60A5FA"/></linearGradient><linearGradient id="adpMiniShield" x1="0" y1="0" x2="1" y2="1"><stop stop-color="#BAE6FD"/><stop offset="1" stop-color="#818CF8"/></linearGradient></defs><g fill="none" stroke="#7DD3FC" stroke-width="3" opacity=".62"><circle cx="69" cy="57" r="43"/><path d="M27 57h84M69 14c-14 13-20 27-20 43s6 30 20 43M69 14c14 13 20 27 20 43s-6 30-20 43M34 35h70M34 79h70"/></g><g fill="url(#adpBrick)" stroke="#93C5FD" stroke-width="1"><rect x="102" y="48" width="43" height="24" rx="5"/><rect x="149" y="48" width="43" height="24" rx="5"/><rect x="196" y="48" width="43" height="24" rx="5"/><rect x="114" y="76" width="43" height="24" rx="5"/><rect x="161" y="76" width="43" height="24" rx="5"/><rect x="208" y="76" width="31" height="24" rx="5"/></g><g transform="translate(174 16)"><path d="M24 1 45 10v14c0 13-8.5 23.5-21 28C11.5 47.5 3 37 3 24V10L24 1Z" fill="url(#adpMiniShield)" stroke="#BAE6FD" stroke-width="2"/><path d="m15 25 6 6 13-14" fill="none" stroke="#E0F2FE" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/></g><g transform="translate(184 73)"><path d="M8 14V9a12 12 0 0 1 24 0v5" fill="none" stroke="#BFDBFE" stroke-width="4"/><rect x="3" y="13" width="34" height="29" rx="6" fill="#4F46E5" stroke="#93C5FD" stroke-width="2"/><circle cx="20" cy="25" r="4" fill="#BAE6FD"/><path d="M20 29v6" stroke="#BAE6FD" stroke-width="3" stroke-linecap="round"/></g></svg>
          </div>
        </section>
        <div class="adp-info-banner"><span class="adp-info-icon"><svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round"><circle cx="12" cy="12" r="9"/><path d="M12 11v5M12 8h.01"/></svg></span><span>ระบบอ่าน Group Membership จาก AD Agent / LDAP / Microsoft Graph แล้วแปลงกลุ่มที่ขึ้นต้นด้วย <code>FW_</code>, <code>Firewall_</code> หรือ <code>Internet_</code> เป็น Internet Policy</span></div>
        """, unsafe_allow_html=True)

        ADP_ICONS = {
            "user": '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="8" r="3.5"/><path d="M5 20c0-4 3.1-7 7-7s7 3 7 7"/></svg>',
            "account": '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="5" width="18" height="14" rx="3"/><circle cx="9" cy="11" r="2.2"/><path d="M5.8 16c.7-1.8 1.8-2.7 3.2-2.7s2.5.9 3.2 2.7M15 10h3M15 14h3"/></svg>',
            "mail": '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="5" width="18" height="14" rx="2.5"/><path d="m4 7 8 6 8-6"/></svg>',
            "server": '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="18" height="7" rx="2"/><rect x="3" y="14" width="18" height="7" rx="2"/><path d="M7 6.5h.01M7 17.5h.01M11 6.5h7M11 17.5h7"/></svg>',
            "shield": '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="M12 2 20 5.5V11c0 5.1-3.2 9.1-8 11-4.8-1.9-8-5.9-8-11V5.5L12 2Z"/><path d="m8.5 12 2.2 2.2 4.8-5"/></svg>',
            "groups": '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><circle cx="9" cy="8" r="3"/><circle cx="17" cy="9" r="2.3"/><path d="M3 20c0-4 2.7-7 6-7s6 3 6 7M15 14c3.2 0 5 2.2 5 5"/></svg>',
            "link": '<svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="m9.5 14.5 5-5M7.5 17.5l-1 1a3.5 3.5 0 0 1-5-5l4-4a3.5 3.5 0 0 1 5 0M16.5 6.5l1-1a3.5 3.5 0 0 1 5 5l-4 4a3.5 3.5 0 0 1-5 0"/></svg>',
        }

        def adp_stat_card(label, value, icon="shield"):
            safe_label = html.escape(str(label))
            safe_value = html.escape(str(value if value not in (None, "") else "-"))
            icon_svg = ADP_ICONS.get(icon, ADP_ICONS["shield"])
            return f'''<div class="adp-stat-card"><div class="adp-stat-icon">{icon_svg}</div><div class="adp-stat-copy"><div class="adp-stat-label">{safe_label}</div><div class="adp-stat-value" title="{safe_value}">{safe_value}</div></div></div>'''

        def adp_stat_grid(items):
            st.markdown('<div class="adp-stat-grid">' + ''.join(adp_stat_card(*item) for item in items) + '</div>', unsafe_allow_html=True)

        def adp_summary_grid(policy_count, group_count):
            st.markdown(f'''<div class="adp-summary-grid">
              <div class="adp-summary-card"><div class="adp-summary-icon">{ADP_ICONS['shield']}</div><div class="adp-summary-label">Internet Policies</div><div class="adp-summary-value">{int(policy_count)}</div><div class="adp-summary-note">นโยบายที่ได้รับ</div></div>
              <div class="adp-summary-card"><div class="adp-summary-icon">{ADP_ICONS['groups']}</div><div class="adp-summary-label">AD Groups</div><div class="adp-summary-value">{int(group_count)}</div><div class="adp-summary-note">กลุ่มที่เป็นสมาชิก</div></div>
            </div>''', unsafe_allow_html=True)

        def adp_policy_table(rows):
            columns = [
                ("Policy Internet", "Policy Internet", "14%"),
                ("AD Group", "AD Group", "14%"),
                ("Policy Name", "Policy Name", "14%"),
                ("Internet Level", "Internet Level", "10%"),
                ("Allowed", "Allowed", "18%"),
                ("Blocked", "Blocked", "11%"),
                ("Firewall Rule", "Firewall", "11%"),
                ("Source", "แหล่งที่มา", "8%"),
            ]
            colgroup = ''.join(f'<col style="width:{width}">' for _, _, width in columns)
            header = ''.join(f'<th>{html.escape(title)}</th>' for _, title, _ in columns)
            body_rows = []
            for row in rows:
                cells = []
                for key, _, _ in columns:
                    value = str(row.get(key, "-") or "-")
                    safe_value = html.escape(value)
                    if key == "Source":
                        # Policies in this view originate from the user's AD group
                        # membership; keep the full mapping source in the tooltip.
                        cells.append(f'<td title="{safe_value}"><span class="adp-source-pill">AD</span></td>')
                    else:
                        cells.append(f'<td title="{safe_value}">{safe_value}</td>')
                body_rows.append('<tr>' + ''.join(cells) + '</tr>')
            table = f'''<div class="adp-table-wrap"><div class="adp-table-title"><span class="adp-stat-icon" style="width:24px;height:24px;flex-basis:24px;border-radius:8px">{ADP_ICONS['link']}</span><span>Internet Policy ที่ได้รับ</span></div><table class="adp-policy-table"><colgroup>{colgroup}</colgroup><thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>'''
            st.markdown(table, unsafe_allow_html=True)

        tab_user, tab_policy, tab_map = st.tabs([
            "ค้นหา User",
            "ค้นหา Policy",
            "Policy Mapping",
        ])

        with tab_user:
            default_identity = st.session_state.get("user_email", "")
            with st.container(border=True):
                st.markdown('<div class="adp-search-panel-marker"></div>', unsafe_allow_html=True)
                st.markdown('<div class="adp-field-label">User / Email / UPN</div>', unsafe_allow_html=True)
                col_identity, col_lookup, col_clear = st.columns([0.66, 0.17, 0.17])
                with col_identity:
                    user_identity = st.text_input(
                        "User / Email / UPN",
                        value=default_identity,
                        placeholder="เช่น supranee.ch หรือ user@company.com",
                        key="ad_policy_user_identity",
                        label_visibility="collapsed",
                    )
                with col_lookup:
                    lookup_clicked = st.button("ตรวจสอบ Policy", type="primary", use_container_width=True, key="ad_policy_lookup")
                with col_clear:
                    if st.button("ล้าง Cache AD", use_container_width=True, key="ad_policy_clear_cache"):
                        ldap_find_user.clear()
                        get_ldap_group_names_for_user.clear()
                        get_ad_agent_policy_summary.clear()
                        get_ad_agent_policy_users.clear()
                        graph_find_user.clear()
                        get_ad_group_names_for_user.clear()
                        load_firewall_policy_mapping.clear()
                        st.rerun()

            if lookup_clicked or user_identity:
                if not user_identity.strip():
                    st.warning("กรุณาระบุ User / Email / UPN")
                else:
                    with st.spinner("กำลังดึงข้อมูลจาก AD / Entra ID..."):
                        policy_summary = get_user_internet_policy_summary(user_identity)
                        user_obj = policy_summary.get("user", {}) if policy_summary.get("ok") else {}

                    if user_obj:
                        adp_stat_grid([
                            ("Display Name", user_obj.get("displayName", "-") or "-", "user"),
                            ("Account", user_obj.get("sAMAccountName") or user_obj.get("userPrincipalName") or "-", "account"),
                            ("Mail", user_obj.get("mail") or "-", "mail"),
                            ("Source", policy_summary.get("source", "-"), "server"),
                        ])
                    else:
                        st.warning("ไม่พบ User นี้จาก AD LDAP / AD Agent / Microsoft Graph หรือตั้งค่า source ยังไม่ครบ")

                    if policy_summary.get("ok"):
                        policies = policy_summary.get("policies", [])
                        groups = policy_summary.get("groups", [])

                        adp_summary_grid(len(policies), len(groups))

                        if policies:
                            adp_policy_table(policies)
                            adp_now = datetime.datetime.now(
                                datetime.timezone(datetime.timedelta(hours=7))
                            )
                            adp_thai_months = [
                                "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
                                "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค.",
                            ]
                            adp_updated_text = (
                                f"{adp_now.day} {adp_thai_months[adp_now.month - 1]} "
                                f"{adp_now.year} {adp_now:%H:%M}"
                            )
                            st.markdown(
                                f'''<div class="adp-table-footer-marker">
                                <svg viewBox="0 0 24 24" fill="none" stroke-width="1.8" stroke-linecap="round"><circle cx="12" cy="12" r="9"/><path d="M12 11v5M12 8h.01"/></svg>
                                <span>อัปเดตล่าสุด: {adp_updated_text}</span></div>''',
                                unsafe_allow_html=True,
                            )
                        else:
                            st.info("ไม่พบ Internet Policy Group สำหรับ User นี้")
                    else:
                        st.error(f"ยังดึงข้อมูล AD / Entra ID ไม่ได้: {policy_summary.get('error', '')}")

        with tab_policy:
            map_rows = load_firewall_policy_mapping()
            available_policies = sorted({
                str(row.get("AD Group", "")).strip()
                for row in map_rows
                if str(row.get("AD Group", "")).strip()
            })

            c_policy, c_manual = st.columns([0.42, 0.58])
            with c_policy:
                selected_policy = st.selectbox(
                    "เลือก Policy",
                    available_policies if available_policies else sorted(FW_POLICY_MAP.keys()),
                    key="ad_policy_selected_policy",
                )
            with c_manual:
                manual_policy = st.text_input(
                    "หรือพิมพ์ชื่อ Policy เอง",
                    placeholder="เช่น FW_Officer_A",
                    key="ad_policy_manual_policy",
                )

            policy_query = manual_policy.strip() or selected_policy

            col_search, col_cache = st.columns([0.52, 0.48])
            with col_search:
                policy_lookup_clicked = st.button("ค้นหา User", type="primary", use_container_width=True, key="ad_policy_policy_lookup")
            with col_cache:
                if st.button("ล้าง Cache Policy", use_container_width=True, key="ad_policy_policy_clear"):
                    get_ad_agent_policy_users.clear()
                    load_firewall_policy_mapping.clear()
                    st.rerun()

            if policy_lookup_clicked or manual_policy.strip():
                with st.spinner(f"กำลังค้นหา User ที่ใช้ {policy_query}..."):
                    policy_users = get_policy_users_summary(policy_query)

                if policy_users.get("ok"):
                    users = policy_users.get("users", []) or []
                    adp_stat_grid([
                        ("Policy", policy_users.get("policy", policy_query), "shield"),
                        ("Users", len(users), "user"),
                        ("Source", policy_users.get("source", "-"), "server"),
                    ])

                    desc = policy_users.get("description") or FW_POLICY_MAP.get(policy_query, "")
                    if desc:
                        st.success(f"{policy_query}: {desc}")

                    if users:
                        users_df = pd.DataFrame(users)
                        preferred_cols = [
                            "displayName",
                            "sAMAccountName",
                            "userPrincipalName",
                            "mail",
                            "department",
                            "title",
                            "company",
                        ]
                        show_cols = [c for c in preferred_cols if c in users_df.columns]
                        if show_cols:
                            users_df = users_df[show_cols]
                        users_df = users_df.rename(columns={
                            "displayName": "Display Name",
                            "sAMAccountName": "Account",
                            "userPrincipalName": "UPN",
                            "mail": "Mail",
                            "department": "Department",
                            "title": "Title",
                            "company": "Company",
                        })

                        st.dataframe(users_df, use_container_width=True, hide_index=True)
                        st.download_button(
                            "Export Users CSV",
                            data=users_df.to_csv(index=False).encode("utf-8-sig"),
                            file_name=f"{policy_query}_users.csv",
                            mime="text/csv",
                            use_container_width=True,
                        )
                    else:
                        st.info("ไม่พบ User ที่อยู่ใน Policy นี้")
                else:
                    st.error(f"ยังดึงรายชื่อ User ของ Policy นี้ไม่ได้: {policy_users.get('error', '')}")

            st.caption("หมายเหตุ: การค้นหา Policy ต้องใช้ AD Agent endpoint `/policy-users` เช่น `https://ad-agent.poonyaruk.co.th/policy-users?policy=FW_Officer_A`")

        with tab_map:
            st.caption(f"อ่าน mapping จาก SharePoint List: {FIREWALL_POLICY_MAPPING_LIST} ถ้าไม่มีหรือว่าง ระบบจะใช้ Default Mapping ในโค้ด")
            map_rows = [
                row for row in load_firewall_policy_mapping()
                if str(row.get("AD Group", "")).strip().casefold() != "fw_supervisor_b"
            ]
            map_df = pd.DataFrame(map_rows)
            if not map_df.empty:
                st.dataframe(map_df, use_container_width=True, hide_index=True)
                st.download_button(
                    "Export Policy Mapping CSV",
                    data=map_df.to_csv(index=False).encode("utf-8-sig"),
                    file_name="firewall_policy_mapping.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            else:
                st.info("ยังไม่มี Policy Mapping")
            st.caption(f"Policy prefixes: {', '.join(FW_POLICY_PREFIXES)}")


    # -------------------------------------------------------
    # 🖨️ Stock หมึกพิมพ์
    # -------------------------------------------------------
    elif main_menu == "⚙ Administration":
        _admin_pages = {
            "admin_users": ("👥", "Users", "จัดการผู้ใช้และสิทธิ์การเข้าถึง"),
            "admin_settings": ("⚙", "Settings", "การตั้งค่าระบบและการเชื่อมต่อ"),
            "admin_logs": ("📜", "Activity Logs", "บันทึกกิจกรรมและการตรวจสอบ"),
        }
        _ap = _admin_pages.get(_nav, ("⚙", "Administration", ""))
        page_header(_ap[0], _ap[1], _ap[2])
        st.markdown("""
        <div style="background:rgba(255,255,255,.92);backdrop-filter:blur(16px);border-radius:16px;
            padding:2rem 2.2rem;border:1px solid #e2e8f0;box-shadow:0 8px 32px rgba(99,102,241,.08);">
            <div style="font-size:2.5rem;margin-bottom:12px;">🚧</div>
            <h3 style="color:#201f1e;margin:0 0 8px;font-size:1.1rem;">Coming soon</h3>
            <p style="color:#605e5c;margin:0;font-size:0.9rem;">
                ส่วนนี้อยู่ระหว่างพัฒนา — ฟีเจอร์จะเปิดใช้งานในรุ่นถัดไป
            </p>
        </div>
        """, unsafe_allow_html=True)

    elif main_menu == "🖨️ Stock หมึกพิมพ์":
        if _nav == "consumables":
            page_header("📁", "Consumables", "วัสดุสิ้นเปลืองและประวัติการเบิกจ่าย")
        else:
            page_header("🖨️", "Stock หมึกพิมพ์", "ระบบติดตามและเบิกจ่ายหมึกพิมพ์")

        df_ink = load_sp_data(INK_STOCK_LIST)

        # ---- Low Stock Alert Banner ----
        if not df_ink.empty and "Quantity" in df_ink.columns and "Min_Qty" in df_ink.columns:
            def _to_int(v, default=0):
                try: return int(v)
                except: return default
            low_mask = df_ink.apply(
                lambda r: _to_int(r.get("Quantity", 0)) <= _to_int(r.get("Min_Qty", INK_LOW_THRESHOLD)), axis=1
            )
            low_items = df_ink[low_mask]
            if not low_items.empty:
                low_names = ", ".join(
                    f"{r.get('Title','?')} ({r.get('Color','-')})"
                    for _, r in low_items.iterrows()
                )
                st.warning(f"⚠️ **สต็อกต่ำ / หมด:** {low_names}")

        st.markdown("---")

        # ---- Sidebar filters ----
        _ink_filter_cols = st.columns(3, gap="small")
        with _ink_filter_cols[0]:
            comp_filter = st.selectbox("🏢 บริษัท:", ["ALL"] + COMPANY_OPTIONS, key="ink_company_filter")
        with _ink_filter_cols[1]:
            color_filter = st.selectbox("🎨 สี:", ["ALL"] + INK_COLOR_OPTIONS, key="ink_color_filter")
        with _ink_filter_cols[2]:
            show_history = st.checkbox("📜 แสดงประวัติการเบิก", value=show_ink_history_only, key="ink_show_history")

        # ---- Admin: add button ----
        if admin_mode:
            col_title2, col_add2 = st.columns([0.8, 0.2])
            with col_add2:
                if st.button("➕ เพิ่มหมึกใหม่", use_container_width=True, type="primary"):
                    add_ink_dialog()

        # ---- Summary Metrics ----
        if not df_ink.empty:
            def _toi(v):
                try: return int(v)
                except: return 0
            total_items = len(df_ink)
            total_qty   = sum(_toi(r) for r in df_ink.get("Quantity", pd.Series([])))
            low_count   = int(low_mask.sum()) if not df_ink.empty and "Quantity" in df_ink.columns else 0
            out_count   = int((df_ink["Quantity"].apply(_toi) == 0).sum()) if "Quantity" in df_ink.columns else 0
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("📦 รายการทั้งหมด", total_items)
            m2.metric("🔢 จำนวนรวม", total_qty, help="รวมทุกสี/รุ่น")
            m3.metric("⚠️ ใกล้หมด", low_count)
            m4.metric("❌ หมดสต็อก", out_count)
            st.markdown("---")

        # ---- Filter + Search ----
        if not df_ink.empty:
            df_show = df_ink.copy()
            if comp_filter != "ALL" and "Company" in df_show.columns:
                df_show = df_show[(df_show["Company"] == comp_filter) | (df_show["Company"] == "ALL")]
            if color_filter != "ALL" and "Color" in df_show.columns:
                df_show = df_show[df_show["Color"] == color_filter]
            search_ink = st.text_input("🔍 ค้นหารุ่นหมึก หรือ รุ่นเครื่องพิมพ์...")
            if search_ink:
                df_show = df_show[df_show.astype(str).apply(
                    lambda x: x.str.contains(search_ink, case=False)
                ).any(axis=1)]

            if df_show.empty:
                st.info("ไม่พบรายการที่ตรงกัน")
            else:
                ink_cols = st.columns(2)
                requester = st.session_state.get("user_name", "Unknown")
                for i, (idx, row) in enumerate(df_show.iterrows()):
                    with ink_cols[i % 2]:
                        render_ink_card(row, idx, admin_mode, requester)
        else:
            st.info("ยังไม่มีข้อมูลสต็อกหมึก — กด ➕ เพิ่มหมึกใหม่ เพื่อเริ่มต้น")
            if not admin_mode:
                st.caption("(ต้องใช้สิทธิ์ Admin ในการเพิ่มข้อมูล)")

        # ---- History Section ----
        if show_history:
            st.markdown("---")
            st.subheader("📜 ประวัติการเบิก/เพิ่มสต็อก")
            df_hist = load_sp_data(INK_HISTORY_LIST)
            if not df_hist.empty:
                display_cols = [c for c in ["Timestamp", "Ink_Title", "Color", "Action", "Qty_Change", "Requester", "Note"] if c in df_hist.columns]
                rename_map   = {k: v for k, v in INK_HISTORY_FIELDS.items() if k in display_cols}
                df_hist_show = df_hist[display_cols].rename(columns=rename_map)
                if "วันเวลา" in df_hist_show.columns:
                    df_hist_show = df_hist_show.sort_values("วันเวลา", ascending=False)

                # Export history
                col_hist_title, col_hist_export = st.columns([0.8, 0.2])
                with col_hist_export:
                    buf_hist = io.StringIO()
                    df_hist_show.to_csv(buf_hist, index=False, encoding="utf-8-sig")
                    st.download_button(
                        "📥 Export CSV", buf_hist.getvalue(),
                        "ink_history.csv", "text/csv", use_container_width=True
                    )
                st.dataframe(df_hist_show, use_container_width=True, hide_index=True)
            else:
                st.info("ยังไม่มีประวัติการเบิก/เพิ่มสต็อก")

    # -------------------------------------------------------
    # 📂 NAS Drive Check
    # -------------------------------------------------------
    elif main_menu == "📂 NAS Drive Check":
        page_header("📂", "NAS Permission Analyzer", "ตรวจสอบสิทธิ์การเข้าถึง Share บน Synology NAS")
        st.info("🔒 ข้อมูลสิทธิ์ NAS เป็น Read-only — กรุณาแก้ไขผ่าน Synology DSM โดยตรง")
        st.markdown("---")

        if 'nas_df' not in st.session_state or st.session_state.nas_df is None:
            with st.spinner("กำลังดึงข้อมูลจาก NAS..."):
                df_result = load_nas_data()
                st.session_state.nas_df = df_result if df_result is not None else None
                if df_result is None:
                    st.error("ไม่สามารถเชื่อมต่อ NAS ได้")

        if st.session_state.nas_df is not None:
            display_df = st.session_state.nas_df.copy()

            st.markdown("""
            <style>
            .nas-hero-search{
                background:linear-gradient(135deg,#ffffff,#f8faff);
                border:1px solid #e2e8f0;
                border-radius:22px;
                padding:22px;
                margin-bottom:20px;
                box-shadow:0 10px 30px rgba(99,102,241,.08);
}

            .nas-modern-card{
                background:linear-gradient(180deg,#ffffff 0%, #fafbff 100%);
                border:1px solid #e2e8f0;
                border-radius:24px;
                padding:22px;
                margin-bottom:18px;
                box-shadow:0 8px 30px rgba(15,23,42,.05);
                transition:all .2s ease;
}

            .nas-modern-card:hover{
                transform:translateY(-2px);
                box-shadow:0 18px 40px rgba(99,102,241,.12);
}

            /* ป้องกัน block ว่างจาก markdown div */
            .nas-modern-card:empty{
                display:none !important;
                padding:0 !important;
                margin:0 !important;
                border:none !important;
}

            .nas-card-title{
                font-size:1.15rem;
                font-weight:800;
                color:#312e81;
                margin-top:8px;
}

            .nas-mini-stat{
                display:inline-flex;
                align-items:center;
                gap:6px;
                padding:7px 12px;
                border-radius:999px;
                background:#eef2ff;
                color:#4f46e5;
                font-size:.76rem;
                font-weight:700;
                margin-right:8px;
                margin-top:10px;
}

            .nas-section-title{
                font-size:.72rem;
                font-weight:800;
                letter-spacing:1px;
                text-transform:uppercase;
                color:#94a3b8;
                margin-bottom:10px;
}

            .nas-user-pill{
                display:inline-flex;
                align-items:center;
                padding:6px 12px;
                border-radius:999px;
                font-size:.76rem;
                font-weight:700;
                margin:4px;
}

            .nas-user-rw{
                background:rgba(99,102,241,.14);
                color:#4338ca;
                border:1px solid rgba(99,102,241,.20);
}

            .nas-user-ro{
                background:#dcfce7;
                color:#166534;
                border:1px solid #86efac;
}
            </style>
            """, unsafe_allow_html=True)

            total_shares = len(display_df)

            total_rw = 0
            total_ro = 0

            for _, row in display_df.iterrows():
                if row["Matched Employees"]:
                    for staff in row["Matched Employees"].split(", "):
                        if "Read/Write" in staff:
                            total_rw += 1
                        else:
                            total_ro += 1

            s1, s2, s3 = st.columns(3)
            with s1:
                st.metric("📁 Shares", total_shares)
            with s2:
                st.metric("🔐 Read / Write", total_rw)
            with s3:
                st.metric("👁 Read Only", total_ro)

            

            col_s, col_r = st.columns([0.82, 0.18])

            with col_s:
                search_term = st.text_input(
                    "🔎 Search User / Share Drive",
                    "",
                    placeholder="ค้นหา Share Drive หรือ Username..."
                )

            with col_r:


                if st.button("🔄 Refresh", use_container_width=True):
                    load_nas_data.clear()
                    st.session_state.nas_df = load_nas_data()
                    st.rerun()

            

            if search_term:
                # แสดง Internet Policy ของ User ที่ค้นหา จาก AD / Firewall Group
                # ถ้าค้นหาเป็นชื่อ Share Drive อย่างเดียว อาจไม่พบ User ใน AD ซึ่งระบบจะแจ้งแบบไม่ทำให้หน้าล่ม
                policy_summary = get_user_internet_policy_summary(search_term)
                if policy_summary.get("ok") and policy_summary.get("policies"):
                    st.markdown("### 🌐 Internet Policy จาก AD Group")
                    st.dataframe(
                        pd.DataFrame(policy_summary["policies"]),
                        use_container_width=True,
                        hide_index=True,
                    )
                    with st.expander("ดู AD Groups ทั้งหมดของ User นี้"):
                        st.write(", ".join(policy_summary.get("groups", [])) or "-")
                elif policy_summary.get("ok"):
                    st.info("🌐 ไม่พบ AD Group ที่ตรงกับ Internet Policy เช่น FW_Officer_B / FW_IT สำหรับคำค้นหานี้")
                else:
                    st.warning(f"🌐 ยังดึง Internet Policy จาก AD ไม่ได้: {policy_summary.get('error', '')}")

                display_df = display_df[
                    display_df["Share"].str.contains(search_term, case=False, na=False) |
                    display_df["Matched Employees"].str.contains(search_term, case=False, na=False) |
                    display_df["ACL Tags (Raw)"].str.contains(search_term, case=False, na=False)
                ]

            for idx, row in display_df.iterrows():

                rw_users = []
                ro_users = []

                raw_acl = str(row['ACL Tags (Raw)'])

                if raw_acl and raw_acl != "nan":

                    for item in [x.strip() for x in raw_acl.split(',')]:

                        m = re.search(r"^(.*?)\s*\((Read(?:/Write)?)\)", item)

                        if m:

                            entity = _clean_nas_principal(m.group(1))

                            permission = m.group(2)

                            if permission == "Read/Write":
                                rw_users.append(entity)
                            else:
                                ro_users.append(entity)

                st.markdown('<div class="nas-modern-card">', unsafe_allow_html=True)

                top1, top2 = st.columns([0.75, 0.25])

                with top1:
                    card_html = f"<div style='display:flex;align-items:center;gap:16px;'>" \
                        f"<div style='width:74px;height:74px;border-radius:20px;background:linear-gradient(135deg,#ede9fe,#c4b5fd);display:flex;align-items:center;justify-content:center;font-size:2rem;border:1px solid #c4b5fd;'>📁</div>" \
                        f"<div>" \
                        f"<div class='nas-card-title'>{row['Share']}</div>" \
                        f"<div class='nas-mini-stat'>👥 {len(rw_users)+len(ro_users)} Users</div>" \
                        f"<div class='nas-mini-stat'>🔐 {len(rw_users)} RW</div>" \
                        f"<div class='nas-mini-stat'>👁 {len(ro_users)} Read</div>" \
                        f"</div></div>"

                    st.markdown(card_html, unsafe_allow_html=True)

                with top2:

                    if st.button("🔎 รายละเอียด", key=f"acl_{idx}", use_container_width=True):

                        @st.dialog("📜 Authorized Users/Groups")
                        def show_acl_pop(raw):
                            if raw:
                                parsed = []
                                for item in [t.strip() for t in raw.split(',')]:
                                    m = re.search(r"^(.*?)\s*\((Read(?:/Write)?)\)", item)
                                    if m:
                                        e = _clean_nas_principal(m.group(1))
                                        if e:
                                            policy_summary = get_user_internet_policy_summary(e)
                                            policies = policy_summary.get("policies", [])
                                            parsed.append({
                                                "Entity": e,
                                                "Permission": m.group(2),
                                                "Policy Internet": format_policy_names(policies),
                                                "Policy Description": format_policy_descriptions(policies),
                                                "AD Groups": ", ".join(policy_summary.get("groups", [])) if policy_summary.get("ok") else f"Error: {policy_summary.get('error', '')}",
                                            })

                                if parsed:
                                    parsed_df = pd.DataFrame(parsed)

                                    tab_perm, tab_policy, tab_raw = st.tabs([
                                        "NAS Permission",
                                        "Internet Policy",
                                        "Raw ACL",
                                    ])

                                    with tab_perm:
                                        st.dataframe(
                                            parsed_df[["Entity", "Permission"]],
                                            use_container_width=True,
                                            hide_index=True
                                        )

                                    with tab_policy:
                                        policy_df = parsed_df[[
                                            "Entity",
                                            "Policy Internet",
                                            "Policy Description",
                                            "AD Groups",
                                        ]]
                                        st.dataframe(
                                            policy_df,
                                            use_container_width=True,
                                            hide_index=True
                                        )
                                        st.caption("ดึงจาก AD / Entra ID Group ที่ขึ้นต้นด้วย FW_ แล้วเทียบกับ FW_POLICY_MAP ในโค้ด")

                                    with tab_raw:
                                        st.code(raw)
                                else:
                                    st.info("ไม่พบรายการ ACL ที่อ่านได้")
                                    with st.expander("Raw ACL"):
                                        st.code(raw)

                        show_acl_pop(row['ACL Tags (Raw)'])

                
                if not rw_users and not ro_users:
                    st.info("ไม่พบผู้ใช้งาน")
                else:
                    st.markdown(
                        f"""
                        <div style='margin-top:10px;padding:14px 18px;border-radius:16px;
                        background:linear-gradient(135deg,#eef2ff,#f8faff);
                        border:1px solid #dbeafe;
                        color:#475569;
                        font-size:0.92rem;
                        font-weight:600;'>
                        🔐 พบผู้ใช้งานที่มีสิทธิ์ทั้งหมด <b>{len(rw_users)+len(ro_users)}</b> รายการ
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

                # ปิด div ของ nas-modern-card เพื่อป้องกัน layout เพี้ยน/เกิดช่องว่างสีขาว
                st.markdown('</div>', unsafe_allow_html=True)



            
            # --------------------------------------------------
            # Export CSV / Excel
            # --------------------------------------------------
            export_rows = []

            for _, row in display_df.iterrows():

                raw_acl = str(row.get('ACL Tags (Raw)', ''))

                if not raw_acl or raw_acl == 'nan':
                    continue

                for item in [x.strip() for x in raw_acl.split(',')]:

                    m = re.search(r"^(.*?)\s*\((Read(?:/Write)?)\)", item)

                    if not m:
                        continue

                    entity = _clean_nas_principal(m.group(1))
                    permission = m.group(2)

                    if search_term:
                        keyword = search_term.lower().strip()

                        if (
                            keyword not in row['Share'].lower() and
                            keyword not in entity.lower()
                        ):
                            continue

                    export_rows.append({
                        "Share Drive": row['Share'],
                        "Name": entity,
                        "Permission": permission
})

            export_df = pd.DataFrame(export_rows)

            if not export_df.empty:
                export_df = export_df.drop_duplicates()
                export_df = export_df.sort_values(
                    by=["Share Drive", "Name"]
                )

            csv_buf = io.StringIO()

            export_df.to_csv(
                csv_buf,
                index=False,
                encoding='utf-8-sig'
            )

            excel_buf = io.BytesIO()

            with pd.ExcelWriter(excel_buf, engine='openpyxl') as writer:

                export_df.to_excel(
                    writer,
                    index=False,
                    sheet_name='NAS Permissions'
                )

                ws = writer.sheets['NAS Permissions']

                for col in ws.columns:

                    max_length = 0
                    column = col[0].column_letter

                    for cell in col:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    ws.column_dimensions[column].width = min(max_length + 5, 50)

                ws.freeze_panes = 'A2'
                ws.auto_filter.ref = ws.dimensions

            if admin_mode:
                st.divider()

                ex1, ex2 = st.columns(2)

                with ex1:
                    st.download_button(
                        "📥 Export CSV",
                        csv_buf.getvalue(),
                        "nas_acl_report.csv",
                        "text/csv",
                        use_container_width=True
                    )

                with ex2:
                    st.download_button(
                        "📊 Export Excel",
                        data=excel_buf.getvalue(),
                        file_name="nas_acl_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
            else:
                st.info("🔒 Export ข้อมูล NAS ได้เฉพาะผู้ดูแลระบบ")


    # -------------------------------------------------------
    # 🔑 Password Information
    # -------------------------------------------------------
    elif main_menu == "🔑 Password Information":
        page_header("🔑", "Password Manager", "ข้อมูล Credentials และรหัสผ่านระบบ")

        with st.spinner("กำลังโหลด..."):
            pw_result = load_password_excel()
            pw_sheets, drive_id = pw_result if isinstance(pw_result, tuple) else ({}, None)

        if not pw_sheets or "_error" in pw_sheets:
            err = pw_sheets.get("_error", "ไม่ทราบสาเหตุ") if pw_sheets else "ไม่ได้รับข้อมูลจาก SharePoint"
            st.error(f"❌ ไม่สามารถโหลดไฟล์ได้")
            st.code(err, language="text")
            st.info(f"💡 ตรวจสอบ: ชื่อโฟลเดอร์ = '{SHAREPOINT_FOLDER}' / ชื่อไฟล์ = '{PASSWORD_FILE_NAME}'")
        else:
            sheet_names = list(pw_sheets.keys())
            selected_sheet = st.selectbox("📋 หมวดหมู่:", sheet_names, key="pw_sheet_select")
            df_pw = pw_sheets[selected_sheet].copy()
            sheet_icon = get_sheet_icon(selected_sheet)

            col_title, col_add = st.columns([0.8, 0.2])
            with col_title:
                st.subheader(f"{sheet_icon} {selected_sheet}")
                st.caption(f"พบข้อมูลทั้งหมด {len(df_pw)} รายการ")
            with col_add:
                if admin_mode:
                    st.write("##")
                    if st.button("➕ เพิ่มรายการ", use_container_width=True, type="primary"):
                        add_password_dialog(selected_sheet, df_pw, drive_id, pw_sheets)

            st.markdown("---")

            if df_pw.empty:
                st.warning("ไม่มีข้อมูลในชีทนี้")
            else:
                card_cols = st.columns(2)
                for idx, row in df_pw.iterrows():
                    with card_cols[idx % 2]:
