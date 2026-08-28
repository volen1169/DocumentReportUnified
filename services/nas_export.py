"""Pure transformation and serialization helpers for NAS permission exports."""

import io
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_COMPANY_ALIASES = {
    "optimaltechcoltd": "OPT",
    "poonyarukpattanacoltd": "PRP",
    "puritylabcoltd": "PLC",
    "evergloryinternationalcoltd": "EGI",
    "siamwinindustrycoltd": "SWI",
}

_PERMISSION_RANK = {"R": 1, "R/W": 2, "Deny": 3}

# -----------------------------------------------------------------------------
# EXPORT CONFIG
# ปรับรายชื่อที่ไม่ต้องการ Export และเพิ่มคอลัมน์ใหม่ได้จากจุดเดียว
# -----------------------------------------------------------------------------
# ตัวอย่าง:
# _NAS_EXPORT_EXCLUDED_NAMES = {"Administrator", "Guest", "Test.User"}
_NAS_EXPORT_EXCLUDED_NAMES = {"acc01","ActiveBackup","admin","Administrator","administrators","backup","Domain Admins","Domain Users",
                              "EGI_WH","Enterprise Admins","Epicor","erplife1","Firealarm","fortigate","HR","HRHO","HRBR","HRBP","IT","it01","it02","it03",
                              "it04","Local Admin","MD","MicroTap","OPT_PL","OPT_SC","OPT_SF","OPT_WH","OPTWAREHOUSE","Pafun.Ath","PLC_WH",
                             "PRP_AC","PRP_HR","PRP_IT","SWI_AC","SWI_MD","SWI_PD","SWI_Plan","SWI_SC","SWI_WH","Patcharin.Su","Veerapat.Ch","Voratat.Ch",}

# เพิ่มคอลัมน์แบบค่าคงที่ให้ทั้ง CSV และ Excel
# ตัวอย่าง:
# _NAS_EXPORT_EXTRA_COLUMNS = {"Status": "Active", "Remark": ""}
_NAS_EXPORT_EXTRA_COLUMNS = {}

# -----------------------------------------------------------------------------
# EXPORT-ONLY REQUIRED NAS PERMISSIONS
# มีผลเฉพาะ CSV / Excel ที่ Export ออกมา ไม่ได้แก้ ACL จริงบน NAS
# -----------------------------------------------------------------------------
_NAS_EXPORT_REQUIRED_GLOBAL_SHARES = {
    "OPG_Data_Center": "R/W",
}

_NAS_EXPORT_REQUIRED_COMPANY_SHARES = {
    "EGI": {"EGI_Data_Center": "R/W"},
    "OPT": {"OPT_Data_Center": "R/W"},
    "PLC": {"PLC_Data_Center": "R/W"},
    "SWI": {"SWI_Data_Center": "R/W"},
}

# OPG_Information_Technology:
# อนุญาต R/W เฉพาะรายชื่อด้านล่างเท่านั้น ส่วนคนอื่นต้องเป็นค่าว่างใน Export
_NAS_EXPORT_OPG_IT_RW_USERS = {
    "teerapat.po",
    "cholticha.ma",
    "itsupport",
    "it_network",
    "sompong.po",
}

# OPT_ISO:
# พนักงาน OPT ทุกคนได้ R ยกเว้นรายชื่อด้านล่างให้เป็นค่าว่างใน Export
_NAS_EXPORT_OPT_ISO_EXCLUDED_USERS = {
    "teerapat.po",
    "cholticha.ma",
    "sompong.po",
    "sirinapa.ru",
    "supanee.na",
    "surattana.ch",
    "sirikorn.ph",
}

_SHARE_HEADER_COLORS = {
    "EGI_": ("166534", "FFFFFF"),
    "OPG_": ("F97316", "FFFFFF"),
    "OPT_": ("93C5FD", "0F172A"),
    "PLC_": ("67E8F9", "0F172A"),
    "SWI_": ("86EFAC", "0F172A"),
}


def _nas_company_abbreviation(value):
    """Convert the full AD company name to the agreed abbreviation."""
    raw = str(value or "").strip()
    # Ignore punctuation/spacing/case so Co.,Ltd. and Co., Ltd. both match.
    key = re.sub(r"[^a-z0-9]", "", raw.casefold())
    return _COMPANY_ALIASES.get(key, raw or "-")


def _nas_first_value(data, *keys):
    if not isinstance(data, dict):
        return ""
    folded = {str(k).casefold(): v for k, v in data.items()}
    for key in keys:
        value = data.get(key)
        if value in (None, ""):
            value = folded.get(str(key).casefold())
        if isinstance(value, list):
            value = value[0] if value else ""
        if value not in (None, ""):
            return str(value).strip()
    return ""


def resolve_nas_export_profile(
    entity,
    *,
    clean_principal,
    policy_lookup,
    policy_formatter,
) -> dict:
    """Resolve Company, Department, and Firewall Policy for an ACL entity."""
    clean_entity = clean_principal(entity)
    candidates = []
    for candidate in (
        clean_entity,
        clean_entity.split("@")[0] if "@" in clean_entity else "",
        clean_entity.replace(" ", ".") if " " in clean_entity else "",
    ):
        candidate = str(candidate or "").strip()
        if candidate and candidate.casefold() not in {x.casefold() for x in candidates}:
            candidates.append(candidate)

    last_error = ""
    for candidate in candidates:
        summary = policy_lookup(candidate)
        if not summary.get("ok"):
            last_error = summary.get("error", "")
            continue
        user = summary.get("user") or {}
        company = _nas_first_value(user, "company", "companyName", "Company", "CompanyName")
        department = _nas_first_value(user, "department", "Department", "departmentName")
        job_title = _nas_first_value(user, "title", "jobTitle", "Title", "JobTitle")
        policy_names = policy_formatter(summary.get("policies", []))
        return {
            "Company": _nas_company_abbreviation(company),
            "Division": department or "-",
            "Position": job_title or "-",
            "Firewall Policy": policy_names or "-",
        }
    return {
        "Company": "-",
        "Division": "-",
        "Position": "-",
        "Firewall Policy": "-",
        "Error": last_error,
    }


def build_nas_export_dataframe(
    display_df,
    *,
    clean_principal,
    profile_lookup,
    exclude_names=None,
    extra_columns=None,
):
    """Build the user-by-share NAS permission matrix in baseline order.

    Parameters
    ----------
    exclude_names : iterable[str] | None
        Exact principal names to remove from the final CSV/Excel export.
        Matching is case-insensitive. If omitted, _NAS_EXPORT_EXCLUDED_NAMES is used.
    extra_columns : dict | None
        Extra constant columns to insert after Department. If omitted,
        _NAS_EXPORT_EXTRA_COLUMNS is used.
    """
    prepared_export = prepare_nas_export_permissions(
        display_df,
        clean_principal=clean_principal,
    )
    return build_nas_export_dataframe_from_permissions(
        prepared_export,
        profile_lookup=profile_lookup,
        exclude_names=exclude_names,
        extra_columns=extra_columns,
    )


def prepare_nas_export_permissions(
    display_df,
    *,
    clean_principal,
):
    """Parse the share and permission mapping before profile enrichment."""
    share_names = []
    permission_by_user = {}

    for _, export_source_row in display_df.iterrows():
        share_name = str(export_source_row.get("Share", "")).strip()
        if not share_name:
            continue
        if share_name not in share_names:
            share_names.append(share_name)
        raw_acl = str(export_source_row.get("ACL Tags (Raw)", "") or "")
        if not raw_acl or raw_acl.casefold() in ("nan", "none"):
            continue
        for item in [part.strip() for part in raw_acl.split(",")]:
            match = re.search(r"^(.*?)\s*\((Read(?:/Write)?|Deny)\)", item, flags=re.I)
            if not match:
                continue
            entity = clean_principal(match.group(1))
            if not entity:
                continue
            raw_permission = match.group(2).casefold()
            permission = "Deny" if raw_permission == "deny" else ("R/W" if "write" in raw_permission else "R")
            user_key = entity.casefold()
            user_record = permission_by_user.setdefault(user_key, {"Name": entity, "Shares": {}})
            old_permission = user_record["Shares"].get(share_name, "")
            if _PERMISSION_RANK.get(permission, 0) > _PERMISSION_RANK.get(old_permission, 0):
                user_record["Shares"][share_name] = permission

    return share_names, permission_by_user


def _nas_export_required_share_names():
    """Return all mandatory export-only share columns in stable order."""
    required = list(_NAS_EXPORT_REQUIRED_GLOBAL_SHARES.keys())
    for company_rules in _NAS_EXPORT_REQUIRED_COMPANY_SHARES.values():
        for share_name in company_rules:
            if share_name not in required:
                required.append(share_name)

    if "OPG_Information_Technology" not in required:
        required.append("OPG_Information_Technology")

    if "OPT_ISO" not in required:
        required.append("OPT_ISO")

    return required


def _apply_nas_export_required_permissions(export_record, company, user_name):
    """Apply mandatory export-only permissions based on Company and user."""
    for share_name, permission in _NAS_EXPORT_REQUIRED_GLOBAL_SHARES.items():
        export_record[share_name] = permission

    company_key = str(company or "").strip().upper()
    for share_name, permission in _NAS_EXPORT_REQUIRED_COMPANY_SHARES.get(company_key, {}).items():
        export_record[share_name] = permission

    # OPG_Information_Technology is restricted to the approved users only.
    # Everyone else must be blank in the exported report, regardless of source ACL.
    normalized_user = str(user_name or "").strip().casefold()
    export_record["OPG_Information_Technology"] = (
        "R/W" if normalized_user in _NAS_EXPORT_OPG_IT_RW_USERS else ""
    )

    # OPT_ISO: พนักงาน OPT ทุกคนได้ R ยกเว้นผู้ใช้ที่ระบุไว้ให้เป็นค่าว่าง
    if company_key == "OPT":
        export_record["OPT_ISO"] = (
            ""
            if normalized_user in _NAS_EXPORT_OPT_ISO_EXCLUDED_USERS
            else "R"
        )
    else:
        export_record["OPT_ISO"] = ""

    return export_record


def build_nas_export_dataframe_from_permissions(
    prepared_export,
    *,
    profile_lookup,
    exclude_names=None,
    extra_columns=None,
):
    """Enrich a prepared permission mapping and build the export DataFrame.

    Filtering and extra columns are applied here so CSV and Excel always receive
    the exact same final DataFrame.
    """
    share_names, permission_by_user = prepared_export

    # เพิ่ม Mandatory Share เป็นคอลัมน์เสมอ แม้ source ACL ปัจจุบันยังไม่มีคอลัมน์นั้น
    share_names = list(share_names)
    for required_share in _nas_export_required_share_names():
        if required_share not in share_names:
            share_names.append(required_share)

    excluded = _NAS_EXPORT_EXCLUDED_NAMES if exclude_names is None else exclude_names
    excluded_keys = {
        str(name or "").strip().casefold()
        for name in excluded
        if str(name or "").strip()
    }

    configured_extra_columns = (
        _NAS_EXPORT_EXTRA_COLUMNS if extra_columns is None else extra_columns
    )
    if not isinstance(configured_extra_columns, dict):
        raise TypeError("extra_columns must be a dict, for example {'Status': 'Active'}")

    matrix_rows = []
    for user_record in sorted(permission_by_user.values(), key=lambda item: item["Name"].casefold()):
        user_name = str(user_record.get("Name", "")).strip()

        # Remove names that should not appear in either CSV or Excel.
        if user_name.casefold() in excluded_keys:
            continue

        profile = profile_lookup(user_name)
        export_record = {
            "Name": user_name,
            "Position": profile.get("Position", "-"),
            "Division": profile.get("Division", "-"),
            "Company": profile.get("Company", "-"),
        }

        # Extra columns are inserted after Department.
        for column_name, value in configured_extra_columns.items():
            export_record[str(column_name)] = value

        export_record.update({share: user_record["Shares"].get(share, "") for share in share_names})

        # บังคับสิทธิ์เฉพาะในรายงาน Export ตามกฎของบริษัท
        # ค่า R/W นี้ตั้งใจให้ override ค่าเดิม (R / Deny / ว่าง) ในไฟล์ Export
        export_record = _apply_nas_export_required_permissions(
            export_record,
            profile.get("Company", "-"),
            user_name,
        )

        export_record["Firewall Policy"] = profile.get("Firewall Policy", "-")
        matrix_rows.append(export_record)

    export_columns = (
        ["Name", "Position", "Division", "Company"]
        + [str(column_name) for column_name in configured_extra_columns]
        + share_names
        + ["Firewall Policy"]
    )

    export_df = pd.DataFrame(matrix_rows, columns=export_columns)

    # FINAL EXPORT GUARD:
    # บังคับ OPG_Information_Technology อีกครั้งที่ DataFrame สุดท้าย
    # เพื่อป้องกันค่า ACL เดิม R/W หลุดกลับเข้ามาใน CSV / Excel
    # ไม่ว่าค่าใน source NAS จะเป็นอะไร คนที่ไม่อยู่ใน allowlist ต้องว่างเสมอ
    if "OPG_Information_Technology" in export_df.columns and "Name" in export_df.columns:
        allowed_it_users = {
            str(name).strip().casefold()
            for name in _NAS_EXPORT_OPG_IT_RW_USERS
        }

        export_df["OPG_Information_Technology"] = export_df["Name"].apply(
            lambda value: (
                "R/W"
                if str(value or "").strip().casefold() in allowed_it_users
                else ""
            )
        )

    # Sort Company A-Z first, then Name A-Z within each company.
    if not export_df.empty:
        export_df = export_df.sort_values(
            by=["Company", "Name"],
            key=lambda col: col.astype(str).str.casefold(),
            kind="stable",
        ).reset_index(drop=True)

    # Generate No. after sorting so numbering follows the final export order.
    export_df.insert(0, "No.", range(1, len(export_df) + 1))
    return export_df


def build_nas_csv(export_df) -> bytes:
    """Serialize a NAS export DataFrame as UTF-8 CSV with BOM."""
    return export_df.to_csv(index=False).encode("utf-8-sig")


def build_nas_excel(export_df) -> bytes:
    """Serialize NAS permissions plus Policy Mapping as a formatted workbook."""
    excel_buf = io.BytesIO()
    export_columns = list(export_df.columns)

    with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="NAS Permissions")
        ws = writer.sheets["NAS Permissions"]
        thin = Side(style="thin", color="000000")
        default_fill, default_font_color = "4472C4", "FFFFFF"
        metadata_columns = {
            "No.",
            "Name",
            "Position",
            "Division",
            "Company",
            "Firewall Policy",
            *[str(column_name) for column_name in _NAS_EXPORT_EXTRA_COLUMNS],
        }

        for col_index, column_name in enumerate(export_columns, start=1):
            cell = ws.cell(row=1, column=col_index)
            is_share_column = column_name not in metadata_columns

            fill_color, font_color = default_fill, default_font_color
            if is_share_column:
                for prefix, colors in _SHARE_HEADER_COLORS.items():
                    if str(column_name).upper().startswith(prefix):
                        fill_color, font_color = colors
                        break

            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name="Kanit", size=10, bold=True, color=font_color)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                text_rotation=90 if is_share_column else 0,
                wrap_text=True,
            )
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for row_index in range(2, ws.max_row + 1):
                body_cell = ws.cell(row=row_index, column=col_index)
                body_cell.font = Font(name="Kanit", size=10)
                body_cell.alignment = Alignment(
                    horizontal="left" if column_name == "Name" else "center",
                    vertical="center",
                )
                body_cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            if column_name == "No.":
                column_width = 7
            elif is_share_column:
                column_width = 8
            elif column_name == "Name":
                column_width = 24
            elif column_name in ("Position", "Division"):
                column_width = 22
            elif column_name == "Company":
                column_width = 12
            else:
                column_width = min(max(len(str(column_name)) + 3, 14), 34)

            ws.column_dimensions[get_column_letter(col_index)].width = column_width

        # บังคับกรอบตาราง NAS Permissions ให้ครบทุกช่อง
        # ครอบคลุมตั้งแต่ Header ถึงข้อมูลแถว/คอลัมน์สุดท้าย
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin,
                )

        ws.row_dimensions[1].height = 120
        ws.freeze_panes = "F2"
        ws.auto_filter.ref = ws.dimensions

        # ---------------------------------------------------------------------
        # Sheet 2: Policy Mapping
        # Static policy reference sheet based on the company's mapping format.
        # ---------------------------------------------------------------------
        policy_ws = writer.book.create_sheet("Policy Mapping")
        # ให้ Policy Mapping เป็น Sheet แรก และ NAS Permissions เป็น Sheet ถัดไป
        writer.book._sheets.remove(policy_ws)
        writer.book._sheets.insert(0, policy_ws)

        # Column widths similar to the reference workbook.
        policy_ws.column_dimensions["A"].width = 4
        policy_ws.column_dimensions["B"].width = 24
        policy_ws.column_dimensions["C"].width = 20
        policy_ws.column_dimensions["D"].width = 34
        policy_ws.column_dimensions["E"].width = 26
        policy_ws.column_dimensions["F"].width = 4

        title_font = Font(name="Kanit", size=11, bold=True, underline="single")
        normal_font = Font(name="Kanit", size=10)
        table_header_font = Font(name="Kanit", size=10, bold=True)
        table_header_fill = PatternFill("solid", fgColor="B7B7B7")

        # Procedure section
        policy_ws["B2"] = "Procedure"
        policy_ws["B2"].font = title_font
        policy_ws["B3"] = "ใบแจ้งขอ/ใบแจ้งพนักงานใหม่ > file: check list user > file: permission share drive"
        policy_ws["B3"].font = normal_font

        # Firewall policy section
        policy_ws["B5"] = "Policy Internet - Firewall"
        policy_ws["B5"].font = title_font

        headers = ["Policy Name", "Level", "Description", "Limit"]
        for idx, header in enumerate(headers, start=2):
            cell = policy_ws.cell(row=7, column=idx, value=header)
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        firewall_rows = [
            ["MD", "MD", "ใช้ได้เฉพาะ MD", "Unlimit"],
            ["Conference", "Conference", "ใช้ได้เฉพาะเครื่อง Conference", "Unlimit + Bypass Authen"],
            ["IT", "IT", "ใช้ได้เฉพาะ IT", "400Mbps"],
            ["Officer_A", "C1-C5", "Allow All", "400Mbps"],
            ["Officer_B", "C1-C5", "Block All", "400Mbps"],
            ["Officer_C", "C1-C5", "Allow Youtube", "400Mbps"],
            ["Officer_D", "C1-C5", "Allow Facebook", "400Mbps"],
            ["Officer_E", "C1-C5", "Allow Youtube + Facebook", "400Mbps"],
            ["Supervisor_A", "C6-C10", "Allow All", "600Mbps"],
            ["Supervisor_B", "C6-C10", "Block All", "600Mbps"],
            ["Supervisor_C", "C6-C10", "Allow Youtube", "600Mbps"],
            ["Supervisor_C_Allow_Google_AI_Studio", "C6-C10", "Allow AI", "600Mbps"],
            ["Supervisor_D", "C6-C10", "Allow Facebook", "600Mbps"],
            ["Supervisor_E", "C6-C10", "Allow Youtube + Facebook", "600Mbps"],
        ]

        for row_offset, values in enumerate(firewall_rows, start=8):
            for col_offset, value in enumerate(values, start=2):
                cell = policy_ws.cell(row=row_offset, column=col_offset, value=value)
                cell.font = normal_font
                cell.alignment = Alignment(
                    horizontal="center" if col_offset in (3, 5) else "left",
                    vertical="center",
                )
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ใส่กรอบสีดำให้ตาราง Firewall ครบทุกช่อง
        # คำนวณแถวสุดท้ายจากจำนวน Policy จริง เพื่อไม่ให้ตกหล่นเมื่อเพิ่ม Policy ใหม่
        firewall_last_row = 7 + len(firewall_rows)
        for row in policy_ws.iter_rows(
            min_row=7,
            max_row=firewall_last_row,
            min_col=2,
            max_col=5,
        ):
            for cell in row:
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=thin,
                )

        # Bitdefender section header, matching the supplied reference.
        policy_ws["B22"] = "Policy Internet - Bitdefender"
        policy_ws["B22"].font = title_font

        # Keep the sheet visually similar to the reference.
        policy_ws.sheet_view.showGridLines = True
        policy_ws.freeze_panes = None

    excel_buf.seek(0)
    return excel_buf.getvalue()
