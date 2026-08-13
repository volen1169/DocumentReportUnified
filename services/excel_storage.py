"""Excel workbook storage helpers."""

import io

import pandas as pd
import requests
import streamlit as st
from openpyxl import load_workbook

from services.microsoft_graph import GRAPH_URL, get_access_token, get_sp_site_id

SHAREPOINT_FOLDER = "Update IT documents"

PASSWORD_FILE_NAME = "Password.xlsx"

SOFTWARE_FILE_MAP = {
    "Group Email": "Software_Group_Email.xlsx",
    "Office 365": "Software_Office365.xlsx",
    "PDF": "Software_PDF.xlsx",
    "Windows": "Software_Windows.xlsx",
    "Offboarded Employees": "Software_Offboarded_Employees.xlsx",
    "Antivirus Security": "Software_Antivirus_Security.xlsx",
    "Developer Tools": "Software_Developer_Tools.xlsx",
}

def parse_password_sheet(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return pd.DataFrame()
    raw_headers = list(all_rows[0])
    valid_col_idx = [i for i, h in enumerate(raw_headers) if h is not None]
    headers = [str(raw_headers[i]).strip() for i in valid_col_idx]
    data = []
    for row in all_rows[1:]:
        vals = [row[i] if i < len(row) else None for i in valid_col_idx]
        non_null = [v for v in vals if v is not None and str(v).strip() != '']
        if non_null:
            data.append([str(v).strip() if v is not None else None for v in vals])
    if not data:
        return pd.DataFrame()
    return pd.DataFrame(data, columns=headers)

@st.cache_data(ttl=1800)
def load_password_excel():
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    try:
        site_id = get_sp_site_id()
        drive_res = requests.get(f"{GRAPH_URL}/sites/{site_id}/drive", headers=headers).json()
        drive_id = drive_res.get("id")
        file_url = f"{GRAPH_URL}/drives/{drive_id}/root:/{SHAREPOINT_FOLDER}/{PASSWORD_FILE_NAME}:/content"
        file_res = requests.get(file_url, headers=headers)
        if file_res.status_code == 200:
            wb = load_workbook(io.BytesIO(file_res.content))
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                df = parse_password_sheet(ws)
                if not df.empty:
                    sheets[sheet_name] = df
            return sheets, drive_id
        return {"_error": f"HTTP {file_res.status_code} - ไม่พบไฟล์ที่ path: {SHAREPOINT_FOLDER}/{PASSWORD_FILE_NAME}", "_url": file_url}, drive_id
    except Exception as e:
        return {"_error": str(e)}, None

@st.cache_data(ttl=60)
def load_software_excels():
    """Load each software category from its own SharePoint Excel workbook.

    Missing or unreadable files are reported separately and never replaced with
    sample rows. All non-empty worksheets inside a workbook are combined into
    the workbook's canonical software category.
    """
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    software_sheets = {}
    errors = {}
    try:
        site_id = get_sp_site_id()
        drive_res = requests.get(
            f"{GRAPH_URL}/sites/{site_id}/drive",
            headers=headers,
            timeout=30,
        )
        drive_res.raise_for_status()
        drive_id = drive_res.json().get("id")
        if not drive_id:
            return {}, {"SharePoint Drive": "ไม่พบ Drive ID"}

        for category_name, file_name in SOFTWARE_FILE_MAP.items():
            file_url = f"{GRAPH_URL}/drives/{drive_id}/root:/{SHAREPOINT_FOLDER}/{file_name}:/content"
            try:
                file_res = requests.get(file_url, headers=headers, timeout=45)
                if file_res.status_code != 200:
                    errors[file_name] = f"HTTP {file_res.status_code}"
                    continue
                workbook = load_workbook(io.BytesIO(file_res.content), data_only=True)
                frames = []
                for worksheet in workbook.worksheets:
                    frame = parse_password_sheet(worksheet)
                    if not frame.empty:
                        frame = frame.copy()
                        frame["Source File"] = file_name
                        frame["Source Sheet"] = worksheet.title
                        frame["Source Row"] = [int(i) + 2 for i in frame.index]
                        frames.append(frame)
                if frames:
                    software_sheets[category_name] = pd.concat(frames, ignore_index=True, sort=False)
                else:
                    software_sheets[category_name] = pd.DataFrame()
            except Exception as file_error:
                errors[file_name] = str(file_error)
        return software_sheets, errors
    except Exception as error:
        return {}, {"SharePoint": str(error)}

def upload_password_excel(drive_id, sheets_dict):
    """
    สร้าง Excel ใหม่จาก sheets_dict แล้วอัปโหลดทับไฟล์เดิมบน SharePoint
    """
    token = get_access_token()
    headers_auth = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/octet-stream'}

    # ดาวน์โหลดไฟล์เดิมมาก่อนเพื่อรักษา format
    site_id = get_sp_site_id()
    dl_headers = {'Authorization': f'Bearer {token}'}
    file_res = requests.get(
        f"{GRAPH_URL}/drives/{drive_id}/root:/{SHAREPOINT_FOLDER}/{PASSWORD_FILE_NAME}:/content",
        headers=dl_headers
    )
    wb = load_workbook(io.BytesIO(file_res.content))

    for sheet_name, df in sheets_dict.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        # ลบ data rows เดิม (เก็บ header row 1)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

        # เขียนข้อมูลใหม่
        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val if val not in ('None', 'nan', '') else None)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    upload_res = requests.put(
        f"{GRAPH_URL}/drives/{drive_id}/root:/{SHAREPOINT_FOLDER}/{PASSWORD_FILE_NAME}:/content",
        headers=headers_auth,
        data=output.getvalue()
    )
    return upload_res.status_code in (200, 201)

def _software_workbook_path(category_name: str) -> str:
    file_name = SOFTWARE_FILE_MAP.get(category_name)
    if not file_name:
        raise ValueError(f"ไม่พบไฟล์ Software สำหรับหมวด {category_name}")
    return f"{SHAREPOINT_FOLDER}/{file_name}"

def _get_sharepoint_drive_id():
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    site_id = get_sp_site_id()
    drive_res = requests.get(f"{GRAPH_URL}/sites/{site_id}/drive", headers=headers, timeout=30)
    drive_res.raise_for_status()
    drive_id = drive_res.json().get("id")
    if not drive_id:
        raise RuntimeError("ไม่พบ SharePoint Drive ID")
    return drive_id

def load_software_workbook(category_name: str):
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}
    drive_id = _get_sharepoint_drive_id()
    workbook_path = _software_workbook_path(category_name)
    file_res = requests.get(
        f"{GRAPH_URL}/drives/{drive_id}/root:/{workbook_path}:/content",
        headers=headers,
        timeout=45,
    )
    file_res.raise_for_status()
    return load_workbook(io.BytesIO(file_res.content)), drive_id, workbook_path

def upload_software_workbook(category_name: str, workbook):
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/octet-stream"}
    drive_id = _get_sharepoint_drive_id()
    workbook_path = _software_workbook_path(category_name)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    upload_res = requests.put(
        f"{GRAPH_URL}/drives/{drive_id}/root:/{workbook_path}:/content",
        headers=headers,
        data=output.getvalue(),
        timeout=45,
    )
    return upload_res.status_code in (200, 201), upload_res.text[:500]

def _software_sheet_names(category_name: str):
    try:
        workbook, _, _ = load_software_workbook(category_name)
        return list(workbook.sheetnames)
    except Exception:
        return []

def _software_row_to_sheet_position(row, default_sheet=""):
    sheet_name = str(row.get("Source Sheet", "") or default_sheet or "").strip()
    source_row = row.get("Source Row", None) if hasattr(row, "get") else None
    if source_row is not None and str(source_row).strip().lower() not in ("", "nan", "none"):
        try:
            return sheet_name, int(float(source_row))
        except Exception:
            pass
    try:
        idx = int(row.name)
    except Exception:
        idx = int(row.get("_row_index", 0) or 0) if hasattr(row, "get") else 0
    return sheet_name, idx + 2

def _software_ws_headers(ws):
    headers = []
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip():
            headers.append(str(cell.value).strip())
    return headers

def _software_form_value(value):
    if value is None:
        return ""
    text = str(value)
    return "" if text.lower() in ("nan", "nat", "none") else text
